"""
Utilidad para importar materiales y clientes desde archivos Excel
Valida los encabezados y formatos antes de importar
"""

import openpyxl
import logging
from typing import List, Dict, Tuple
from pathlib import Path

logger = logging.getLogger(__name__)


class ExcelImporterError(Exception):
    """Excepción para errores de importación de Excel"""
    pass


class ExcelImporter:
    """
    Importador de datos desde Excel para Lactalis Ventas

    ENCABEZADOS REQUERIDOS:
    - Materiales: CODIGO, DESCRIPCION, SOCIEDAD
    - Clientes: Cód.Padre, Nombre Código Padre, NIT
    """

    # Encabezados esperados
    MATERIALES_HEADERS = ['CODIGO', 'DESCRIPCION', 'SOCIEDAD']
    CLIENTES_HEADERS = ['Cód.Padre', 'Nombre Código Padre', 'NIT']

    # Mapeo de variantes de encabezados (normalización)
    MATERIALES_HEADERS_VARIANTS = {
        'codigo': 'CODIGO',
        'código': 'CODIGO',
        'code': 'CODIGO',
        'descripcion': 'DESCRIPCION',
        'descripción': 'DESCRIPCION',
        'description': 'DESCRIPCION',
        'sociedad': 'SOCIEDAD',
        'company': 'SOCIEDAD'
    }

    CLIENTES_HEADERS_VARIANTS = {
        'cod.padre': 'Cód.Padre',
        'cod padre': 'Cód.Padre',
        'codigo padre': 'Cód.Padre',
        'código padre': 'Cód.Padre',
        'nombre codigo padre': 'Nombre Código Padre',
        'nombre código padre': 'Nombre Código Padre',
        'nombre': 'Nombre Código Padre',
        'nit': 'NIT'
    }

    @staticmethod
    def importar_materiales_desde_excel(archivo_path: str) -> List[Dict]:
        """
        Importa materiales desde un archivo Excel

        Args:
            archivo_path: Ruta al archivo Excel

        Returns:
            Lista de diccionarios con materiales

        Raises:
            ExcelImporterError: Si hay problemas con el archivo o formato
        """
        try:
            # Verificar que el archivo existe
            if not Path(archivo_path).exists():
                raise ExcelImporterError(f"Archivo no encontrado: {archivo_path}")

            # Leer el archivo Excel
            logger.info(f"Leyendo archivo de materiales: {archivo_path}")
            wb = openpyxl.load_workbook(archivo_path, read_only=True)
            ws = wb.active

            # Leer encabezados (primera fila)
            headers_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            if not headers_row:
                raise ExcelImporterError("El archivo está vacío")

            # Normalizar encabezados
            headers_originales = [str(h).strip() if h else '' for h in headers_row]
            logger.debug(f"Encabezados encontrados: {headers_originales}")

            # Mapear encabezados con variantes
            columnas_mapeadas = {}
            headers_normalizados = []
            for col in headers_originales:
                col_lower = col.lower().strip()
                if col_lower in ExcelImporter.MATERIALES_HEADERS_VARIANTS:
                    col_normalizado = ExcelImporter.MATERIALES_HEADERS_VARIANTS[col_lower]
                    columnas_mapeadas[col] = col_normalizado
                    headers_normalizados.append(col_normalizado)
                else:
                    headers_normalizados.append(col)

            if columnas_mapeadas:
                logger.info(f"Columnas normalizadas: {columnas_mapeadas}")

            # Verificar encabezados requeridos
            headers_faltantes = [h for h in ExcelImporter.MATERIALES_HEADERS if h not in headers_normalizados]
            if headers_faltantes:
                raise ExcelImporterError(
                    f"Encabezados faltantes: {headers_faltantes}. "
                    f"Se esperan: {ExcelImporter.MATERIALES_HEADERS}"
                )

            # Obtener índices de columnas
            idx_codigo = headers_normalizados.index('CODIGO')
            idx_descripcion = headers_normalizados.index('DESCRIPCION')
            idx_sociedad = headers_normalizados.index('SOCIEDAD')

            # Mapeo de nombres de empresas a NITs
            SOCIEDAD_MAP = {
                'parmalat': '800245795',  # Lactalis
                'lactalis': '800245795',
                'proleche': '890903711',  # Procesadora de Leches
                'procesadora de leches': '890903711',
            }

            # Leer datos (desde fila 2 en adelante)
            materiales = []
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row or all(cell is None or str(cell).strip() == '' for cell in row):
                    continue  # Saltar filas vacías

                # Extraer valores
                codigo = str(row[idx_codigo]).strip() if row[idx_codigo] is not None else ''
                descripcion = str(row[idx_descripcion]).strip() if row[idx_descripcion] is not None else ''
                sociedad_raw = str(row[idx_sociedad]).strip() if row[idx_sociedad] is not None else ''

                # Convertir nombre de empresa a NIT
                sociedad_lower = sociedad_raw.lower()
                sociedad_nit = SOCIEDAD_MAP.get(sociedad_lower, sociedad_raw)

                # Si no se encontró en el mapeo y no parece un NIT numérico, loguear advertencia
                if sociedad_nit == sociedad_raw and not sociedad_raw.isdigit():
                    if sociedad_raw:  # Solo advertir si no está vacío
                        logger.warning(
                            f"Fila {row_num}: Sociedad '{sociedad_raw}' no reconocida. "
                            f"Se esperaba 'Parmalat' o 'Proleche'. Se usará tal cual."
                        )

                material = {
                    'codigo': codigo,
                    'descripcion': descripcion,
                    'sociedad': sociedad_nit
                }

                # Filtrar filas vacías
                if material['codigo'] or material['descripcion'] or material['sociedad']:
                    materiales.append(material)

            wb.close()
            logger.info(f"Materiales extraídos: {len(materiales)}")
            return materiales

        except ExcelImporterError:
            raise
        except Exception as e:
            logger.error(f"Error importando materiales: {str(e)}")
            raise ExcelImporterError(f"Error leyendo archivo: {str(e)}")

    @staticmethod
    def importar_clientes_desde_excel(archivo_path: str) -> List[Dict]:
        """
        Importa clientes desde un archivo Excel

        Args:
            archivo_path: Ruta al archivo Excel

        Returns:
            Lista de diccionarios con clientes

        Raises:
            ExcelImporterError: Si hay problemas con el archivo o formato
        """
        try:
            # Verificar que el archivo existe
            if not Path(archivo_path).exists():
                raise ExcelImporterError(f"Archivo no encontrado: {archivo_path}")

            # Leer el archivo Excel
            logger.info(f"Leyendo archivo de clientes: {archivo_path}")
            wb = openpyxl.load_workbook(archivo_path, read_only=True)
            ws = wb.active

            # Leer encabezados (primera fila)
            headers_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            if not headers_row:
                raise ExcelImporterError("El archivo está vacío")

            # Normalizar encabezados
            headers_originales = [str(h).strip() if h else '' for h in headers_row]
            logger.debug(f"Encabezados encontrados: {headers_originales}")

            # Mapear encabezados con variantes
            columnas_mapeadas = {}
            headers_normalizados = []
            for col in headers_originales:
                col_lower = col.lower().strip()
                if col_lower in ExcelImporter.CLIENTES_HEADERS_VARIANTS:
                    col_normalizado = ExcelImporter.CLIENTES_HEADERS_VARIANTS[col_lower]
                    columnas_mapeadas[col] = col_normalizado
                    headers_normalizados.append(col_normalizado)
                else:
                    headers_normalizados.append(col)

            if columnas_mapeadas:
                logger.info(f"Columnas normalizadas: {columnas_mapeadas}")

            # Verificar encabezados requeridos
            headers_faltantes = [h for h in ExcelImporter.CLIENTES_HEADERS if h not in headers_normalizados]
            if headers_faltantes:
                raise ExcelImporterError(
                    f"Encabezados faltantes: {headers_faltantes}. "
                    f"Se esperan: {ExcelImporter.CLIENTES_HEADERS}"
                )

            # Obtener índices de columnas
            idx_cod_padre = headers_normalizados.index('Cód.Padre')
            idx_nombre = headers_normalizados.index('Nombre Código Padre')
            idx_nit = headers_normalizados.index('NIT')

            # Leer datos (desde fila 2 en adelante)
            clientes = []
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row or all(cell is None or str(cell).strip() == '' for cell in row):
                    continue  # Saltar filas vacías

                # Extraer valores
                cod_padre = str(row[idx_cod_padre]).strip() if row[idx_cod_padre] is not None else ''
                nombre = str(row[idx_nombre]).strip() if row[idx_nombre] is not None else ''
                nit = str(row[idx_nit]).strip() if row[idx_nit] is not None else ''

                cliente = {
                    'cod_padre': cod_padre,
                    'nombre_codigo_padre': nombre,
                    'nit': nit
                }

                # Filtrar filas vacías
                if cliente['cod_padre'] or cliente['nombre_codigo_padre']:
                    clientes.append(cliente)

            wb.close()
            logger.info(f"Clientes extraídos: {len(clientes)}")
            return clientes

        except ExcelImporterError:
            raise
        except Exception as e:
            logger.error(f"Error importando clientes: {str(e)}")
            raise ExcelImporterError(f"Error leyendo archivo: {str(e)}")

    @staticmethod
    def validar_archivo_materiales(archivo_path: str) -> Tuple[bool, str]:
        """
        Valida que un archivo Excel tenga el formato correcto para materiales

        Args:
            archivo_path: Ruta al archivo Excel

        Returns:
            Tupla (es_valido, mensaje)
        """
        try:
            if not Path(archivo_path).exists():
                return False, f"Archivo no encontrado: {archivo_path}"

            wb = openpyxl.load_workbook(archivo_path, read_only=True)
            ws = wb.active

            # Leer solo encabezados
            headers_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            if not headers_row:
                wb.close()
                return False, "El archivo está vacío"

            headers_originales = [str(h).strip() if h else '' for h in headers_row]

            # Intentar mapear encabezados
            columnas_mapeadas = {}
            headers_normalizados = []
            for col in headers_originales:
                col_lower = col.lower().strip()
                if col_lower in ExcelImporter.MATERIALES_HEADERS_VARIANTS:
                    col_normalizado = ExcelImporter.MATERIALES_HEADERS_VARIANTS[col_lower]
                    columnas_mapeadas[col] = col_normalizado
                    headers_normalizados.append(col_normalizado)
                else:
                    headers_normalizados.append(col)

            headers_faltantes = [h for h in ExcelImporter.MATERIALES_HEADERS if h not in headers_normalizados]

            wb.close()

            if headers_faltantes:
                return False, f"Encabezados faltantes: {', '.join(headers_faltantes)}"

            return True, "Formato válido"

        except Exception as e:
            return False, f"Error validando archivo: {str(e)}"

    @staticmethod
    def validar_archivo_clientes(archivo_path: str) -> Tuple[bool, str]:
        """
        Valida que un archivo Excel tenga el formato correcto para clientes

        Args:
            archivo_path: Ruta al archivo Excel

        Returns:
            Tupla (es_valido, mensaje)
        """
        try:
            if not Path(archivo_path).exists():
                return False, f"Archivo no encontrado: {archivo_path}"

            wb = openpyxl.load_workbook(archivo_path, read_only=True)
            ws = wb.active

            # Leer solo encabezados
            headers_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            if not headers_row:
                wb.close()
                return False, "El archivo está vacío"

            headers_originales = [str(h).strip() if h else '' for h in headers_row]

            # Intentar mapear encabezados
            columnas_mapeadas = {}
            headers_normalizados = []
            for col in headers_originales:
                col_lower = col.lower().strip()
                if col_lower in ExcelImporter.CLIENTES_HEADERS_VARIANTS:
                    col_normalizado = ExcelImporter.CLIENTES_HEADERS_VARIANTS[col_lower]
                    columnas_mapeadas[col] = col_normalizado
                    headers_normalizados.append(col_normalizado)
                else:
                    headers_normalizados.append(col)

            headers_faltantes = [h for h in ExcelImporter.CLIENTES_HEADERS if h not in headers_normalizados]

            wb.close()

            if headers_faltantes:
                return False, f"Encabezados faltantes: {', '.join(headers_faltantes)}"

            return True, "Formato válido"

        except Exception as e:
            return False, f"Error validando archivo: {str(e)}"
