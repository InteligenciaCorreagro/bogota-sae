"""
Procesador específico para LACTALIS COMPRAS
Lee archivos ZIP y XML, extrae datos y genera Excel en formato REGGIS
"""

import zipfile
import logging
import openpyxl
from pathlib import Path
from typing import List, Dict, Optional
from openpyxl.styles import Font, PatternFill, Alignment

from config.constants import REGGIS_HEADERS, NAMESPACES, get_data_output_path
from extractors.lactalis_extractor import FacturaExtractorLactalis

logger = logging.getLogger(__name__)


class ProcesadorLactalis:
    """
    Procesador específico para LACTALIS COMPRAS

    Funcionalidad:
    1. Lee carpeta con archivos ZIP y/o XML
    2. Por cada ZIP, extrae el archivo XML
    3. Por cada XML suelto, lo procesa directamente
    4. Procesa todos los XML con FacturaExtractorLactalis
    5. Genera archivo Excel con formato REGGIS
    """

    def __init__(self, carpeta_archivos: Path, plantilla_excel: Path):
        """
        Inicializa el procesador

        Args:
            carpeta_archivos: Path a carpeta con archivos ZIP y/o XML
            plantilla_excel: Path a plantilla Excel base
        """
        self.carpeta_archivos = carpeta_archivos
        self.plantilla_excel = plantilla_excel
        self.carpeta_salida = None

    def procesar(self) -> Path:
        """
        Procesa todos los archivos ZIP y XML en la carpeta

        Returns:
            Path a la carpeta de salida con resultados
        """
        logger.info(f"Iniciando procesamiento de LACTALIS COMPRAS desde: {self.carpeta_archivos}")

        # Buscar archivos ZIP y XML
        archivos_zip = list(self.carpeta_archivos.glob("*.zip"))
        archivos_xml = list(self.carpeta_archivos.glob("*.xml"))

        total_archivos = len(archivos_zip) + len(archivos_xml)
        logger.info(f"Se encontraron {len(archivos_zip)} archivo(s) ZIP y {len(archivos_xml)} archivo(s) XML")

        if total_archivos == 0:
            raise ValueError("No se encontraron archivos ZIP ni XML en la carpeta")

        # Procesar todos los archivos
        todas_lineas = []
        archivos_procesados = 0
        archivos_error = 0

        # Procesar ZIPs
        for zip_file in archivos_zip:
            try:
                lineas = self.procesar_zip(zip_file)
                todas_lineas.extend(lineas)
                archivos_procesados += 1
                logger.info(f"[OK] Procesado ZIP: {zip_file.name} - {len(lineas)} lineas")
            except Exception as e:
                archivos_error += 1
                logger.error(f"[ERROR] Procesando ZIP {zip_file.name}: {str(e)}")

        # Procesar XMLs sueltos
        for xml_file in archivos_xml:
            try:
                lineas = self.procesar_xml(xml_file)
                todas_lineas.extend(lineas)
                archivos_procesados += 1
                logger.info(f"[OK] Procesado XML: {xml_file.name} - {len(lineas)} lineas")
            except Exception as e:
                archivos_error += 1
                logger.error(f"[ERROR] Procesando XML {xml_file.name}: {str(e)}")

        logger.info(
            f"Procesamiento completado: {archivos_procesados} exitosos, "
            f"{archivos_error} con errores, "
            f"{len(todas_lineas)} lineas totales"
        )

        if not todas_lineas:
            raise ValueError(
                f"No se pudo extraer ninguna linea de los archivos.\n"
                f"Archivos procesados: {archivos_procesados}\n"
                f"Archivos con error: {archivos_error}\n\n"
                f"Revise los logs para mas detalles."
            )

        # Crear carpeta de salida
        self.carpeta_salida = self.crear_carpeta_salida()

        # Escribir Excel
        archivo_salida = self.escribir_reggis(todas_lineas)

        logger.info(f"Archivo Excel generado: {archivo_salida}")
        return self.carpeta_salida

    def procesar_xml(self, xml_path: Path) -> List[Dict]:
        """
        Procesa un archivo XML directamente

        Args:
            xml_path: Path al archivo XML

        Returns:
            Lista de líneas extraídas del XML
        """
        try:
            # Leer contenido del XML
            with open(xml_path, 'r', encoding='utf-8', errors='ignore') as f:
                xml_content = f.read()

            # Verificar si es un AttachedDocument y extraer la factura
            invoice_xml = self.extraer_invoice_de_attached_document(xml_content)

            if invoice_xml:
                # Es un AttachedDocument, procesar el XML interno
                logger.debug(f"{xml_path.name}: Es un AttachedDocument, extrayendo factura interna")
                extractor = FacturaExtractorLactalis(invoice_xml, xml_path.name)
            else:
                # Es un XML de factura directo
                logger.debug(f"{xml_path.name}: Es una factura directa")
                extractor = FacturaExtractorLactalis(xml_content, xml_path.name)

            lineas = extractor.extraer_datos()
            return lineas

        except Exception as e:
            logger.error(f"Error procesando XML {xml_path.name}: {str(e)}", exc_info=True)
            raise

    def extraer_invoice_de_attached_document(self, xml_content: str) -> Optional[str]:
        """
        Extrae el XML de la factura desde un documento adjunto (AttachedDocument)

        Args:
            xml_content: Contenido XML completo

        Returns:
            XML de la factura si es un AttachedDocument, None si no lo es
        """
        try:
            import xml.etree.ElementTree as ET
            root = ET.fromstring(xml_content)

            # Buscar en ExternalReference/Description (formato SEABOARD/LACTALIS)
            description = root.find('.//cac:ExternalReference/cbc:Description', NAMESPACES)

            if description is not None and description.text:
                return description.text.strip()

            # No es un AttachedDocument
            return None

        except Exception as e:
            logger.debug(f"No es un AttachedDocument: {str(e)}")
            return None

    def procesar_zip(self, zip_path: Path) -> List[Dict]:
        """
        Procesa un archivo ZIP extrayendo el XML y procesándolo

        Args:
            zip_path: Path al archivo ZIP

        Returns:
            Lista de líneas extraídas del XML
        """
        lineas = []

        try:
            with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                # Listar archivos en el ZIP
                archivos_en_zip = zip_ref.namelist()
                logger.debug(f"Archivos en {zip_path.name}: {archivos_en_zip}")

                # Buscar archivo XML (puede haber PDF también)
                xml_files = [f for f in archivos_en_zip if f.lower().endswith('.xml')]

                if not xml_files:
                    logger.warning(f"No se encontro archivo XML en {zip_path.name}")
                    return []

                # Procesar el primer XML encontrado
                xml_filename = xml_files[0]
                logger.debug(f"Extrayendo XML: {xml_filename}")

                # Leer contenido del XML directamente del ZIP
                with zip_ref.open(xml_filename) as xml_file:
                    xml_content = xml_file.read().decode('utf-8', errors='ignore')

                # Verificar si es un AttachedDocument y extraer la factura
                invoice_xml = self.extraer_invoice_de_attached_document(xml_content)

                if invoice_xml:
                    # Es un AttachedDocument, procesar el XML interno
                    logger.debug(f"{zip_path.name}: Es un AttachedDocument, extrayendo factura interna")
                    extractor = FacturaExtractorLactalis(invoice_xml, f"{zip_path.name}/{xml_filename}")
                else:
                    # Es un XML de factura directo
                    logger.debug(f"{zip_path.name}: Es una factura directa")
                    extractor = FacturaExtractorLactalis(xml_content, f"{zip_path.name}/{xml_filename}")

                lineas = extractor.extraer_datos()

                if len(xml_files) > 1:
                    logger.warning(f"Se encontraron {len(xml_files)} XMLs en {zip_path.name}, solo se proceso el primero")

        except zipfile.BadZipFile:
            logger.error(f"Archivo ZIP corrupto: {zip_path.name}")
            raise
        except Exception as e:
            logger.error(f"Error procesando ZIP {zip_path.name}: {str(e)}", exc_info=True)
            raise

        return lineas

    def crear_carpeta_salida(self) -> Path:
        """
        Crea la carpeta de salida para los resultados en data/YYYY-MM-DD/

        Returns:
            Path a la carpeta de salida
        """
        # Usar helper que crea estructura data/YYYY-MM-DD/
        return get_data_output_path()

    def escribir_reggis(self, lineas: List[Dict]) -> Path:
        """
        Escribe las líneas procesadas en un archivo Excel formato REGGIS

        Args:
            lineas: Lista de diccionarios con datos de cada línea

        Returns:
            Path al archivo Excel generado
        """
        # Cargar plantilla
        wb = openpyxl.load_workbook(self.plantilla_excel)
        ws = wb.active

        # Determinar fila inicial
        fila_inicial = ws.max_row + 1 if ws.max_row > 1 else 2

        # Escribir cada línea
        for idx, linea in enumerate(lineas, start=fila_inicial):
            ws.cell(row=idx, column=1, value=linea['numero_factura'])
            ws.cell(row=idx, column=2, value=linea['nombre_producto'])
            ws.cell(row=idx, column=3, value=linea['codigo_subyacente'])
            ws.cell(row=idx, column=4, value=linea['unidad_medida'])
            ws.cell(row=idx, column=5, value=linea['cantidad'])
            ws.cell(row=idx, column=6, value=linea['precio_unitario'])
            ws.cell(row=idx, column=7, value=linea['fecha_factura'])
            ws.cell(row=idx, column=8, value=linea['fecha_pago'])
            ws.cell(row=idx, column=9, value=linea['nit_comprador'])
            ws.cell(row=idx, column=10, value=linea['nombre_comprador'])
            ws.cell(row=idx, column=11, value=linea['nit_vendedor'])
            ws.cell(row=idx, column=12, value=linea['nombre_vendedor'])
            ws.cell(row=idx, column=13, value=linea['principal'])
            ws.cell(row=idx, column=14, value=linea['municipio'])
            ws.cell(row=idx, column=15, value=linea['iva'])
            ws.cell(row=idx, column=16, value=linea['descripcion'])
            ws.cell(row=idx, column=17, value=linea['activa_factura'])
            ws.cell(row=idx, column=18, value=linea['activa_bodega'])
            ws.cell(row=idx, column=19, value=linea['incentivo'])
            ws.cell(row=idx, column=20, value=linea['cantidad_original'])
            ws.cell(row=idx, column=21, value=linea['moneda'])
            ws.cell(row=idx, column=22, value=linea['total_sin_iva'])
            ws.cell(row=idx, column=23, value=linea['total_iva'])
            ws.cell(row=idx, column=24, value=linea['total_con_iva'])

        # Generar nombre de archivo de salida
        archivo_salida = self.carpeta_salida / f"LACTALIS_COMPRAS_{self.carpeta_archivos.name}.xlsx"

        # Guardar
        wb.save(archivo_salida)
        logger.info(f"Excel guardado: {archivo_salida}")

        return archivo_salida

    def crear_plantilla_base(self, ruta: Path):
        """
        Crea una plantilla base de Excel con formato REGGIS

        Args:
            ruta: Path donde crear la plantilla
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Datos"

        # Estilos de encabezado
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        # Escribir encabezados
        for col, header in enumerate(REGGIS_HEADERS, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        wb.save(ruta)
        logger.info(f"Plantilla creada: {ruta}")
