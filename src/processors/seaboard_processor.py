"""
Procesador específico para SEABOARD
"""

import xml.etree.ElementTree as ET
import openpyxl
import logging
from pathlib import Path
from typing import Dict, List, Optional

from config.constants import NAMESPACES
from extractors.seaboard_extractor import FacturaExtractorSeaboard

logger = logging.getLogger(__name__)


class ProcesadorSeaboard:
    """Procesador específico para SEABOARD"""

    def __init__(self, carpeta_xml: Path, plantilla_excel: Path):
        self.carpeta_xml = carpeta_xml
        self.plantilla_excel = plantilla_excel
        self.carpeta_salida = None

    def extraer_invoice_de_attached_document(self, xml_path: Path) -> Optional[str]:
        """Extrae el XML de la factura desde un documento adjunto"""
        try:
            tree = ET.parse(xml_path)
            root = tree.getroot()
            description = root.find('.//cac:ExternalReference/cbc:Description', NAMESPACES)

            if description is not None and description.text:
                return description.text.strip()
            return None
        except Exception as e:
            logger.error(f"Error al extraer factura de {xml_path.name}: {str(e)}")
            return None

    def procesar_archivos_xml(self) -> List[Dict]:
        """Procesa todos los archivos XML en la carpeta"""
        lineas_reggis = []
        archivos_xml = list(self.carpeta_xml.glob("*.xml"))

        logger.info(f"SEABOARD: Se encontraron {len(archivos_xml)} archivos XML")

        for xml_file in archivos_xml:
            try:
                invoice_xml = self.extraer_invoice_de_attached_document(xml_file)

                if invoice_xml:
                    extractor = FacturaExtractorSeaboard(invoice_xml)
                    lineas = extractor.extraer_datos()
                    lineas_reggis.extend(lineas)

            except Exception as e:
                logger.error(f"Error procesando {xml_file.name}: {str(e)}")

        return lineas_reggis

    def crear_carpeta_salida(self) -> Path:
        """Crea la carpeta de salida para los resultados"""
        carpeta_resultados_base = self.plantilla_excel.parent / "Resultados_SEABOARD"
        carpeta_resultados_base.mkdir(exist_ok=True)

        nombre_carpeta_xml = self.carpeta_xml.name
        carpeta_salida = carpeta_resultados_base / nombre_carpeta_xml
        carpeta_salida.mkdir(exist_ok=True)

        return carpeta_salida

    def escribir_reggis(self, lineas: List[Dict]) -> Path:
        """Escribe las líneas procesadas en el archivo Excel"""
        wb = openpyxl.load_workbook(self.plantilla_excel)
        ws = wb.active

        fila_inicial = ws.max_row + 1 if ws.max_row > 1 else 2

        for idx, linea in enumerate(lineas, start=fila_inicial):
            ws.cell(row=idx, column=1, value=linea['numero_factura'])
            ws.cell(row=idx, column=2, value=linea['nombre_producto'])
            ws.cell(row=idx, column=3, value=linea['codigo_subyacente'])
            ws.cell(row=idx, column=4, value=linea['unidad_medida'])
            ws.cell(row=idx, column=5, value=linea['cantidad'])
            ws.cell(row=idx, column=6, value=linea['precio_unitario'])
            ws.cell(row=idx, column=7, value=linea['fecha_factura'])
            ws.cell(row=idx, column=8, value=linea['fecha_pago'])
            ws.cell(row=idx, column=9, value=linea['nit_comprador'])
            ws.cell(row=idx, column=10, value=linea['nombre_comprador'])
            ws.cell(row=idx, column=11, value=linea['nit_vendedor'])
            ws.cell(row=idx, column=12, value=linea['nombre_vendedor'])
            ws.cell(row=idx, column=13, value=linea['principal'])
            ws.cell(row=idx, column=14, value=linea['municipio'])
            ws.cell(row=idx, column=15, value=linea['iva'])
            ws.cell(row=idx, column=16, value=linea['descripcion'])
            ws.cell(row=idx, column=17, value=linea['activa_factura'])
            ws.cell(row=idx, column=18, value=linea['activa_bodega'])
            ws.cell(row=idx, column=19, value=linea['incentivo'])
            ws.cell(row=idx, column=20, value=linea['cantidad_original'])
            ws.cell(row=idx, column=21, value=linea['moneda'])
            ws.cell(row=idx, column=22, value=linea['total_sin_iva'])
            ws.cell(row=idx, column=23, value=linea['total_iva'])
            ws.cell(row=idx, column=24, value=linea['total_con_iva'])

        salida = self.carpeta_salida / "REGGIS_Procesado_SEABOARD.xlsx"
        wb.save(salida)
        return salida

    def procesar(self) -> Path:
        """Ejecuta el procesamiento completo"""
        self.carpeta_salida = self.crear_carpeta_salida()
        lineas_reggis = self.procesar_archivos_xml()

        if not lineas_reggis:
            raise Exception("No se procesaron líneas. Verifique los archivos XML.")

        self.escribir_reggis(lineas_reggis)
        logger.info(f"SEABOARD: Proceso completado - {len(lineas_reggis)} líneas")

        return self.carpeta_salida
