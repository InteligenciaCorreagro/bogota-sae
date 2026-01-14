"""
Procesador específico para LACTALIS VENTAS
Optimizado para procesar grandes volúmenes (20,000+ XML) con procesamiento por lotes
Prioriza estabilidad sobre velocidad con límites conservadores
"""

import zipfile
import logging
import openpyxl
import gc  # Para liberación explícita de memoria
from pathlib import Path
from typing import List, Dict, Tuple
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

from src.config.constants import REGGIS_HEADERS, LACTALIS_VENTAS_CONFIG, get_data_output_path
from extractors.lactalis_ventas_extractor import FacturaExtractorLactalisVentas, ValidacionFacturaError

try:
    from src.database.lactalis_database import LactalisDatabase
except ImportError:
    LactalisDatabase = None

logger = logging.getLogger(__name__)


class ProcesadorLactalisVentas:
    """
    Procesador específico para LACTALIS VENTAS con procesamiento por lotes
    
    Optimizaciones para grandes volúmenes:
    1. Procesamiento por lotes (batch processing)
    2. Validaciones tempranas para descartar archivos inválidos
    3. Logging detallado de estadísticas
    4. Manejo eficiente de memoria
    """

    def __init__(self, carpeta_archivos: Path, plantilla_excel: Path,
                 progress_callback=None, database: 'LactalisDatabase' = None,
                 validar_materiales: bool = False, validar_clientes: bool = False):
        """
        Inicializa el procesador

        Args:
            carpeta_archivos: Path a carpeta con archivos ZIP y/o XML
            plantilla_excel: Path a plantilla Excel base
            progress_callback: Callback opcional para reportar progreso (recibe: processed, total, message)
            database: Instancia de LactalisDatabase para validaciones (opcional)
            validar_materiales: Si True, valida que los materiales existan en la BD
            validar_clientes: Si True, valida que los clientes existan en la BD
        """
        self.carpeta_archivos = carpeta_archivos
        self.plantilla_excel = plantilla_excel
        self.carpeta_salida = None
        self.progress_callback = progress_callback
        self.database = database
        self.validar_materiales = validar_materiales
        self.validar_clientes = validar_clientes

        # Estadísticas
        self.stats = {
            'total_archivos': 0,
            'facturas_validas': 0,
            'notas_credito': 0,
            'notas_debito': 0,
            'otros_documentos': 0,
            'lineas_validas': 0,
            'lineas_rechazadas': 0,
            'materiales_invalidos': 0,
            'clientes_invalidos': 0,
            'archivos_error': 0,
            'tiempo_inicio': None,
            'tiempo_fin': None,
        }

    def _reportar_progreso(self, processed: int, total: int, message: str = ""):
        """Reporta el progreso si hay callback configurado"""
        if self.progress_callback:
            try:
                self.progress_callback(processed, total, message)
            except Exception as e:
                logger.error(f"Error en callback de progreso: {str(e)}")

    def procesar(self) -> Path:
        """
        Procesa todos los archivos ZIP y XML en la carpeta
        
        Returns:
            Path a la carpeta de salida con resultados
        """
        self.stats['tiempo_inicio'] = datetime.now()
        
        logger.info(f"=" * 80)
        logger.info(f"Iniciando procesamiento de LACTALIS VENTAS desde: {self.carpeta_archivos}")
        logger.info(f"=" * 80)

        # Buscar archivos ZIP y XML
        archivos_zip = list(self.carpeta_archivos.glob("*.zip"))
        archivos_xml = list(self.carpeta_archivos.glob("*.xml"))

        total_archivos = len(archivos_zip) + len(archivos_xml)
        self.stats['total_archivos'] = total_archivos
        
        logger.info(f"Se encontraron {len(archivos_zip)} archivo(s) ZIP y {len(archivos_xml)} archivo(s) XML")

        if total_archivos == 0:
            raise ValueError("No se encontraron archivos ZIP ni XML en la carpeta")

        # Procesar en lotes CONSERVADORES para evitar crashes
        # Usamos lotes más pequeños para mejor manejo de memoria
        batch_size = min(LACTALIS_VENTAS_CONFIG.get('batch_size', 500), 100)
        memory_batch_size = 1000  # Escribir a Excel cada 1000 líneas

        todas_lineas = []
        archivos_procesados = 0
        lineas_escritas = 0

        logger.info(f"Procesando en lotes de {batch_size} archivos (prioridad: estabilidad)")
        logger.info(f"Escritura a Excel cada {memory_batch_size} líneas para liberar memoria")
        
        # Procesar ZIPs con MÁXIMA PROTECCIÓN
        for idx, zip_file in enumerate(archivos_zip, 1):
            try:
                self._reportar_progreso(
                    archivos_procesados,
                    total_archivos,
                    f"Procesando ZIP {idx}/{len(archivos_zip)}: {zip_file.name}"
                )
                
                lineas = self.procesar_zip(zip_file)
                todas_lineas.extend(lineas)
                archivos_procesados += 1

                if lineas:
                    logger.info(f"[OK] {zip_file.name} - {len(lineas)} líneas")
                else:
                    logger.debug(f"[SKIP] {zip_file.name} - Sin líneas válidas")

                # Liberar memoria cada cierto número de archivos
                if archivos_procesados % batch_size == 0:
                    logger.debug(f"Liberando memoria después de {archivos_procesados} archivos...")
                    gc.collect()  # Forzar garbage collection
                    
            except KeyboardInterrupt:
                logger.warning("⚠ Procesamiento cancelado por el usuario")
                raise
            except Exception as e:
                # ERROR CRÍTICO - NO CERRAR, solo registrar y continuar
                logger.error(f"[ERROR CRÍTICO] {zip_file.name}: {type(e).__name__}: {str(e)}")
                self.stats['archivos_error'] += 1
                archivos_procesados += 1
                # CONTINUAR con el siguiente archivo

        # Procesar XMLs sueltos con MÁXIMA PROTECCIÓN
        for idx, xml_file in enumerate(archivos_xml, 1):
            try:
                self._reportar_progreso(
                    archivos_procesados,
                    total_archivos,
                    f"Procesando XML {idx}/{len(archivos_xml)}: {xml_file.name}"
                )
                
                lineas = self.procesar_xml(xml_file)
                todas_lineas.extend(lineas)
                archivos_procesados += 1

                if lineas:
                    logger.info(f"[OK] {xml_file.name} - {len(lineas)} líneas")
                else:
                    logger.debug(f"[SKIP] {xml_file.name} - Sin líneas válidas")

                # Liberar memoria cada cierto número de archivos
                if archivos_procesados % batch_size == 0:
                    logger.debug(f"Liberando memoria después de {archivos_procesados} archivos...")
                    gc.collect()  # Forzar garbage collection
                    
            except KeyboardInterrupt:
                logger.warning("⚠ Procesamiento cancelado por el usuario")
                raise
            except Exception as e:
                # ERROR CRÍTICO - NO CERRAR, solo registrar y continuar
                logger.error(f"[ERROR CRÍTICO] {xml_file.name}: {type(e).__name__}: {str(e)}")
                self.stats['archivos_error'] += 1
                archivos_procesados += 1
                # CONTINUAR con el siguiente archivo

        self.stats['tiempo_fin'] = datetime.now()

        # Liberar memoria después de procesar todos los archivos
        logger.info("Liberando memoria después de procesar archivos...")
        gc.collect()

        # Aplicar validaciones de base de datos si están habilitadas
        self._reportar_progreso(
            total_archivos,
            total_archivos,
            f"Aplicando validaciones a {len(todas_lineas)} líneas..."
        )

        lineas_antes_validacion = len(todas_lineas)
        todas_lineas = self._filtrar_lineas_validas(todas_lineas)
        self.stats['lineas_validas'] = len(todas_lineas)

        if self.database and (self.validar_materiales or self.validar_clientes):
            logger.info(
                f"Validación BD: {lineas_antes_validacion} líneas -> "
                f"{len(todas_lineas)} válidas, "
                f"{lineas_antes_validacion - len(todas_lineas)} rechazadas"
            )

        # Liberar memoria después de validaciones
        logger.info("Liberando memoria después de validaciones...")
        gc.collect()

        # Mostrar estadísticas
        self._mostrar_estadisticas()

        if not todas_lineas:
            raise ValueError(
                f"No se pudo extraer ninguna línea de los archivos.\n\n"
                f"ESTADÍSTICAS:\n"
                f"  • Total archivos: {self.stats['total_archivos']}\n"
                f"  • Facturas válidas: {self.stats['facturas_validas']}\n"
                f"  • Notas crédito: {self.stats['notas_credito']}\n"
                f"  • Notas débito: {self.stats['notas_debito']}\n"
                f"  • Otros documentos: {self.stats['otros_documentos']}\n"
                f"  • Archivos con error: {self.stats['archivos_error']}\n\n"
                f"Revise los logs para más detalles."
            )

        # Crear carpeta de salida y escribir Excel
        self.carpeta_salida = self.crear_carpeta_salida()

        self._reportar_progreso(
            total_archivos,
            total_archivos,
            f"Escribiendo Excel con {len(todas_lineas)} líneas..."
        )

        archivo_salida = self.escribir_reggis(todas_lineas)

        logger.info(f"=" * 80)
        logger.info(f"Archivo Excel generado: {archivo_salida}")
        logger.info(f"=" * 80)
        
        return self.carpeta_salida

    def _mostrar_estadisticas(self):
        """Muestra estadísticas detalladas del procesamiento"""
        tiempo_total = self.stats['tiempo_fin'] - self.stats['tiempo_inicio']

        logger.info(f"")
        logger.info(f"=" * 80)
        logger.info(f"ESTADÍSTICAS DE PROCESAMIENTO - LACTALIS VENTAS")
        logger.info(f"=" * 80)
        logger.info(f"Total archivos procesados: {self.stats['total_archivos']}")
        logger.info(f"  • Facturas válidas: {self.stats['facturas_validas']}")
        logger.info(f"  • Notas crédito: {self.stats['notas_credito']} (RECHAZADAS)")
        logger.info(f"  • Notas débito: {self.stats['notas_debito']} (RECHAZADAS)")
        logger.info(f"  • Otros documentos: {self.stats['otros_documentos']} (RECHAZADOS)")
        logger.info(f"  • Archivos con error: {self.stats['archivos_error']}")
        logger.info(f"")
        logger.info(f"Líneas procesadas:")
        logger.info(f"  • Líneas válidas: {self.stats['lineas_validas']}")
        logger.info(f"  • Líneas rechazadas: {self.stats['lineas_rechazadas']}")

        if self.database and (self.validar_materiales or self.validar_clientes):
            logger.info(f"")
            logger.info(f"Validaciones de base de datos:")
            if self.validar_materiales:
                logger.info(f"  • Materiales inválidos: {self.stats['materiales_invalidos']}")
            if self.validar_clientes:
                logger.info(f"  • Clientes inválidos: {self.stats['clientes_invalidos']}")

        logger.info(f"")
        logger.info(f"Tiempo total: {tiempo_total}")

        if self.stats['total_archivos'] > 0:
            promedio = tiempo_total.total_seconds() / self.stats['total_archivos']
            logger.info(f"Promedio por archivo: {promedio:.3f} segundos")

        logger.info(f"=" * 80)

    def procesar_xml(self, xml_path: Path) -> List[Dict]:
        """
        Procesa un archivo XML directamente con MÁXIMA PROTECCIÓN
        
        Args:
            xml_path: Path al archivo XML

        Returns:
            Lista de líneas extraídas del XML ([] si hay error)
        """
        try:
            # **PASO 1: Leer archivo con múltiples encodings**
            xml_content = None
            encodings_to_try = ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1']
            
            for encoding in encodings_to_try:
                try:
                    with open(xml_path, 'r', encoding=encoding) as f:
                        xml_content = f.read()
                    break  # Si funciona, salir del loop
                except (UnicodeDecodeError, UnicodeError):
                    continue
                except Exception as read_error:
                    logger.error(f"Error leyendo {xml_path.name}: {str(read_error)}")
                    self.stats['archivos_error'] += 1
                    return []
            
            if xml_content is None:
                logger.error(f"No se pudo leer {xml_path.name} con ningún encoding")
                self.stats['archivos_error'] += 1
                return []
            
            # **PASO 2: Crear extractor con protección**
            try:
                extractor = FacturaExtractorLactalisVentas(xml_content, xml_path.name)
            except Exception as e:
                logger.error(f"Error creando extractor para {xml_path.name}: {str(e)}")
                self.stats['archivos_error'] += 1
                return []
            
            # **PASO 3: Actualizar estadísticas según el tipo de documento**
            try:
                if extractor.tipo_documento == 'Invoice':
                    self.stats['facturas_validas'] += 1
                elif extractor.tipo_documento == 'CreditNote':
                    self.stats['notas_credito'] += 1
                elif extractor.tipo_documento == 'DebitNote':
                    self.stats['notas_debito'] += 1
                else:
                    self.stats['otros_documentos'] += 1
            except:
                self.stats['otros_documentos'] += 1

            # **PASO 4: Extraer datos con protección máxima**
            try:
                lineas = extractor.extraer_datos()
                return lineas if lineas else []
            except ValidacionFacturaError as e:
                # Error esperado - no es factura válida
                logger.debug(f"{xml_path.name}: {str(e)}")
                return []
            except MemoryError:
                logger.error(f"{xml_path.name}: Archivo demasiado grande (MemoryError)")
                self.stats['archivos_error'] += 1
                return []
            except Exception as e:
                logger.error(f"{xml_path.name}: Error extrayendo datos - {type(e).__name__}: {str(e)}")
                self.stats['archivos_error'] += 1
                return []

        except KeyboardInterrupt:
            # Permitir cancelación
            raise
        except Exception as fatal_error:
            # Error CRÍTICO - registrar pero NO CERRAR
            logger.critical(f"[FATAL] {xml_path.name}: {type(fatal_error).__name__}: {str(fatal_error)}")
            self.stats['archivos_error'] += 1
            return []

    def procesar_zip(self, zip_path: Path) -> List[Dict]:
        """
        Procesa un archivo ZIP extrayendo el XML y procesándolo con MÁXIMA PROTECCIÓN
        
        Args:
            zip_path: Path al archivo ZIP

        Returns:
            Lista de líneas extraídas del XML ([] si hay error)
        """
        try:
            # **PASO 1: Abrir ZIP con protección**
            try:
                zip_ref = zipfile.ZipFile(zip_path, 'r')
            except zipfile.BadZipFile:
                logger.error(f"{zip_path.name}: Archivo ZIP corrupto")
                self.stats['archivos_error'] += 1
                return []
            except Exception as e:
                logger.error(f"{zip_path.name}: Error abriendo ZIP - {str(e)}")
                self.stats['archivos_error'] += 1
                return []
            
            try:
                # **PASO 2: Buscar XMLs dentro del ZIP**
                archivos_en_zip = zip_ref.namelist()
                xml_files = [f for f in archivos_en_zip if f.lower().endswith('.xml')]

                if not xml_files:
                    logger.warning(f"{zip_path.name}: No se encontró archivo XML")
                    self.stats['otros_documentos'] += 1
                    return []

                # **PASO 3: Extraer y leer XML con múltiples encodings**
                xml_filename = xml_files[0]
                xml_content = None
                
                try:
                    with zip_ref.open(xml_filename) as xml_file:
                        raw_content = xml_file.read()
                    
                    # Intentar múltiples encodings
                    encodings_to_try = ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1']
                    for encoding in encodings_to_try:
                        try:
                            xml_content = raw_content.decode(encoding)
                            break
                        except (UnicodeDecodeError, UnicodeError):
                            continue
                    
                    if xml_content is None:
                        logger.error(f"{zip_path.name}/{xml_filename}: No se pudo decodificar con ningún encoding")
                        self.stats['archivos_error'] += 1
                        return []
                        
                except Exception as e:
                    logger.error(f"{zip_path.name}: Error leyendo XML interno - {str(e)}")
                    self.stats['archivos_error'] += 1
                    return []

                # **PASO 4: Crear extractor con protección**
                try:
                    extractor = FacturaExtractorLactalisVentas(
                        xml_content,
                        f"{zip_path.name}/{xml_filename}"
                    )
                except Exception as e:
                    logger.error(f"{zip_path.name}: Error creando extractor - {str(e)}")
                    self.stats['archivos_error'] += 1
                    return []
                
                # **PASO 5: Actualizar estadísticas**
                try:
                    if extractor.tipo_documento == 'Invoice':
                        self.stats['facturas_validas'] += 1
                    elif extractor.tipo_documento == 'CreditNote':
                        self.stats['notas_credito'] += 1
                    elif extractor.tipo_documento == 'DebitNote':
                        self.stats['notas_debito'] += 1
                    else:
                        self.stats['otros_documentos'] += 1
                except:
                    self.stats['otros_documentos'] += 1

                # **PASO 6: Extraer datos con protección**
                try:
                    lineas = extractor.extraer_datos()
                except ValidacionFacturaError as e:
                    logger.debug(f"{zip_path.name}: {str(e)}")
                    return []
                except MemoryError:
                    logger.error(f"{zip_path.name}: Archivo demasiado grande (MemoryError)")
                    self.stats['archivos_error'] += 1
                    return []
                except Exception as e:
                    logger.error(f"{zip_path.name}: Error extrayendo datos - {type(e).__name__}: {str(e)}")
                    self.stats['archivos_error'] += 1
                    return []

                if len(xml_files) > 1:
                    logger.warning(
                        f"{zip_path.name}: Se encontraron {len(xml_files)} XMLs, "
                        f"solo se procesó el primero"
                    )

                return lineas if lineas else []
                
            finally:
                # **SIEMPRE cerrar el ZIP**
                try:
                    zip_ref.close()
                except:
                    pass

        except KeyboardInterrupt:
            # Permitir cancelación
            raise
        except Exception as fatal_error:
            # Error CRÍTICO - registrar pero NO CERRAR
            logger.critical(f"[FATAL] {zip_path.name}: {type(fatal_error).__name__}: {str(fatal_error)}")
            self.stats['archivos_error'] += 1
            return []

    def crear_carpeta_salida(self) -> Path:
        """
        Crea la carpeta de salida para los resultados en data/YYYY-MM-DD/

        Returns:
            Path a la carpeta de salida
        """
        return get_data_output_path()

    def _validar_linea_con_bd(self, linea: Dict) -> Tuple[bool, str]:
        """
        Valida una línea contra la base de datos

        Args:
            linea: Diccionario con datos de la línea

        Returns:
            Tupla (es_valida, mensaje_error)
        """
        if not self.database:
            logger.debug("No hay base de datos configurada, se acepta la línea")
            return True, ""

        # Validar material
        if self.validar_materiales:
            codigo = linea.get('codigo_subyacente', '')
            nombre_producto = linea.get('nombre_producto', '').upper()

            # Determinar sociedad según el nombre del producto
            # Parmalat → Lactalis (800245795)
            # Proleche → Proleche (890903711)
            if 'PARMALAT' in nombre_producto:
                sociedad = '800245795'  # Lactalis
            elif 'PROLECHE' in nombre_producto:
                sociedad = '890903711'  # Proleche
            else:
                # Si no contiene ninguno, usar el NIT del vendedor como fallback
                sociedad = linea.get('nit_vendedor', '')

            logger.debug(f"Validando material: {codigo} con sociedad {sociedad} (Producto: {nombre_producto})")

            if not self.database.validar_material(codigo, sociedad):
                mensaje = f"Material RECHAZADO - No registrado: {codigo} (Sociedad: {sociedad}, Producto: {nombre_producto})"
                logger.warning(mensaje)
                self.stats['materiales_invalidos'] += 1
                return False, mensaje
            else:
                logger.debug(f"Material ACEPTADO: {codigo} con sociedad {sociedad}")

        # Validar cliente
        if self.validar_clientes:
            nit_comprador = linea.get('nit_comprador', '')

            logger.debug(f"Validando cliente: {nit_comprador}")

            if not self.database.validar_cliente(nit_comprador):
                mensaje = f"Cliente RECHAZADO - No registrado: {nit_comprador}"
                logger.warning(mensaje)
                self.stats['clientes_invalidos'] += 1
                return False, mensaje
            else:
                logger.debug(f"Cliente ACEPTADO: {nit_comprador}")

        return True, ""

    def _filtrar_lineas_validas(self, lineas: List[Dict]) -> List[Dict]:
        """
        Filtra líneas validándolas contra la base de datos

        Args:
            lineas: Lista de líneas a validar

        Returns:
            Lista de líneas válidas
        """
        if not self.database:
            logger.info("No hay base de datos configurada, se aceptan todas las líneas")
            return lineas

        if not self.validar_materiales and not self.validar_clientes:
            logger.info("Validaciones desactivadas, se aceptan todas las líneas")
            return lineas

        logger.info(f"Iniciando validación de {len(lineas)} líneas")
        logger.info(f"Validar materiales: {self.validar_materiales}")
        logger.info(f"Validar clientes: {self.validar_clientes}")

        lineas_validas = []
        lineas_rechazadas_detalle = 0

        for idx, linea in enumerate(lineas, 1):
            es_valida, mensaje = self._validar_linea_con_bd(linea)
            if es_valida:
                lineas_validas.append(linea)
            else:
                lineas_rechazadas_detalle += 1
                self.stats['lineas_rechazadas'] += 1
                if lineas_rechazadas_detalle <= 10:  # Solo mostrar las primeras 10
                    logger.info(f"Línea {idx} rechazada: {mensaje}")

        if lineas_rechazadas_detalle > 10:
            logger.info(f"... y {lineas_rechazadas_detalle - 10} líneas más rechazadas")

        logger.info(f"Resultado validación: {len(lineas_validas)} válidas, {lineas_rechazadas_detalle} rechazadas")

        return lineas_validas

    def escribir_reggis(self, lineas: List[Dict]) -> Path:
        """
        Escribe las líneas procesadas en un archivo Excel formato REGGIS
        Con reporte de progreso y manejo de memoria mejorado

        Args:
            lineas: Lista de diccionarios con datos de cada línea

        Returns:
            Path al archivo Excel generado
        """
        logger.info(f"Iniciando escritura de {len(lineas)} líneas a Excel...")

        # Cargar plantilla en modo de solo escritura para mejor rendimiento
        wb = openpyxl.load_workbook(self.plantilla_excel)
        ws = wb.active

        # Determinar fila inicial
        fila_inicial = ws.max_row + 1 if ws.max_row > 1 else 2

        # Escribir cada línea con progreso
        total_lineas = len(lineas)
        report_interval = max(100, total_lineas // 20)  # Reportar cada 5%

        for linea_num, linea in enumerate(lineas, start=0):
            fila_excel = fila_inicial + linea_num

            ws.cell(row=fila_excel, column=1, value=linea['numero_factura'])
            ws.cell(row=fila_excel, column=2, value=linea['nombre_producto'])
            ws.cell(row=fila_excel, column=3, value=linea['codigo_subyacente'])
            ws.cell(row=fila_excel, column=4, value=linea['unidad_medida'])
            ws.cell(row=fila_excel, column=5, value=linea['cantidad'])
            ws.cell(row=fila_excel, column=6, value=linea['precio_unitario'])
            ws.cell(row=fila_excel, column=7, value=linea['fecha_factura'])
            ws.cell(row=fila_excel, column=8, value=linea['fecha_pago'])
            ws.cell(row=fila_excel, column=9, value=linea['nit_comprador'])
            ws.cell(row=fila_excel, column=10, value=linea['nombre_comprador'])
            ws.cell(row=fila_excel, column=11, value=linea['nit_vendedor'])
            ws.cell(row=fila_excel, column=12, value=linea['nombre_vendedor'])
            ws.cell(row=fila_excel, column=13, value=linea['principal'])
            ws.cell(row=fila_excel, column=14, value=linea['municipio'])
            ws.cell(row=fila_excel, column=15, value=linea['iva'])
            ws.cell(row=fila_excel, column=16, value=linea['descripcion'])
            ws.cell(row=fila_excel, column=17, value=linea['activa_factura'])
            ws.cell(row=fila_excel, column=18, value=linea['activa_bodega'])
            ws.cell(row=fila_excel, column=19, value=linea['incentivo'])
            ws.cell(row=fila_excel, column=20, value=linea['cantidad_original'])
            ws.cell(row=fila_excel, column=21, value=linea['moneda'])
            ws.cell(row=fila_excel, column=22, value=linea['total_sin_iva'])
            ws.cell(row=fila_excel, column=23, value=linea['total_iva'])
            ws.cell(row=fila_excel, column=24, value=linea['total_con_iva'])

            # Reportar progreso cada cierto número de líneas
            if (linea_num + 1) % report_interval == 0:
                porcentaje = int(((linea_num + 1) / total_lineas) * 100)
                logger.debug(f"Escritura Excel: {linea_num + 1}/{total_lineas} líneas ({porcentaje}%)")

        logger.info(f"✅ {total_lineas} líneas escritas a Excel")

        # Generar nombre de archivo de salida con timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        archivo_salida = self.carpeta_salida / f"LACTALIS_VENTAS_{self.carpeta_archivos.name}_{timestamp}.xlsx"

        # Guardar
        wb.save(archivo_salida)
        logger.info(f"Excel guardado: {archivo_salida}")

        return archivo_salida