"""
Procesador específico para CASA DEL AGRICULTOR
"""

import xml.etree.ElementTree as ET
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import logging
import zipfile
import re
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Optional

from config.constants import CURRENCY_CODE_MAP, UNIT_MAP

logger = logging.getLogger(__name__)


class ProcesadorCasaDelAgricultor:
    """Procesador específico para CASA DEL AGRICULTOR"""

    def __init__(self, carpeta_zip: Path, carpeta_salida: Path):
        self.carpeta_zip = carpeta_zip
        self.carpeta_salida = carpeta_salida
        self.processed_lines = []

    def extract_xml_from_zip(self, zip_path: Path) -> Optional[str]:
        """Extrae el contenido XML de un archivo ZIP"""
        try:
            with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                xml_files = [f for f in zip_ref.namelist() if f.endswith('.xml')]
                if xml_files:
                    xml_content = zip_ref.read(xml_files[0])
                    return xml_content.decode('utf-8')
        except Exception as e:
            logger.error(f"Error al extraer XML de ZIP: {str(e)}")
        return None

    def parse_invoice_xml(self, xml_content: str) -> List[Dict]:
        """Parsea el XML de la factura y extrae los datos"""
        try:
            cdata_match = re.search(r'<!\[CDATA\[(.*?)\]\]>', xml_content, re.DOTALL)
            if cdata_match:
                invoice_xml = cdata_match.group(1)
            else:
                invoice_xml = xml_content

            invoice_xml = re.sub(r'\sxmlns[^=]*="[^"]*"', '', invoice_xml)
            invoice_xml = re.sub(r'<([a-zA-Z]+):([a-zA-Z]+)', r'<\2', invoice_xml)
            invoice_xml = re.sub(r'</([a-zA-Z]+):([a-zA-Z]+)', r'</\2', invoice_xml)

            root = ET.fromstring(invoice_xml)

            lines = []

            supplier_elem = root.find('.//AccountingSupplierParty')
            if supplier_elem:
                supplier_name = supplier_elem.findtext('.//RegistrationName') or ''
                supplier_nit = supplier_elem.findtext('.//CompanyID') or ''
                supplier_city = supplier_elem.findtext('.//CityName') or ''
            else:
                supplier_name = supplier_nit = supplier_city = ''

            customer_elem = root.find('.//AccountingCustomerParty')
            if customer_elem:
                customer_name = customer_elem.findtext('.//RegistrationName') or ''
                customer_nit = customer_elem.findtext('.//CompanyID') or ''
            else:
                customer_name = customer_nit = ''

            invoice_number = root.findtext('.//ID') or ''
            invoice_date = root.findtext('.//IssueDate') or ''
            payment_date = root.findtext('.//PaymentDueDate') or ''

            currency_elem = root.find('.//DocumentCurrencyCode')
            currency_code = currency_elem.text if currency_elem is not None else 'COP'
            currency_id = CURRENCY_CODE_MAP.get(currency_code, '1')

            for line in root.findall('.//InvoiceLine'):
                line_id = line.findtext('.//ID') or ''

                qty_elem = line.find('.//InvoicedQuantity')
                if qty_elem is not None:
                    quantity = float(qty_elem.text or 0)
                    unit_code = qty_elem.get('unitCode') or ''
                else:
                    quantity = 0
                    unit_code = ''

                desc_elem = line.find('.//Description')
                description = desc_elem.text if desc_elem is not None else ''

                code_elem = line.find('.//StandardItemIdentification/ID')
                if code_elem is None:
                    code_elem = line.find('.//Item/ID')
                code = code_elem.text if code_elem is not None else ''

                price_elem = line.find('.//PriceAmount')
                price = float(price_elem.text or 0) if price_elem is not None else 0

                total_elem = line.find('.//LineExtensionAmount')
                line_total = float(total_elem.text or 0) if total_elem is not None else 0

                iva_percent = 0
                iva_amount = 0
                tax_total = line.find('.//TaxTotal')
                if tax_total is not None:
                    percent_elem = tax_total.find('.//Percent')
                    if percent_elem is not None:
                        iva_percent = float(percent_elem.text or 0)

                    iva_elem = tax_total.find('.//TaxAmount')
                    if iva_elem is not None:
                        iva_amount = float(iva_elem.text or 0)

                total_with_iva = line_total + iva_amount

                lines.append({
                    'line_id': line_id,
                    'code': code,
                    'description': description,
                    'quantity': quantity,
                    'unit': unit_code,
                    'price': price,
                    'line_total': line_total,
                    'invoice_number': invoice_number,
                    'invoice_date': invoice_date,
                    'payment_date': payment_date,
                    'supplier_name': supplier_name,
                    'supplier_nit': supplier_nit,
                    'supplier_city': supplier_city,
                    'customer_name': customer_name,
                    'customer_nit': customer_nit,
                    'currency_id': currency_id,
                    'iva_percent': iva_percent,
                    'iva_amount': iva_amount,
                    'total_with_iva': total_with_iva
                })

            return lines

        except Exception as e:
            logger.error(f"Error al parsear XML: {str(e)}")
            return []

    def apply_conversion_rules(self, line: Dict) -> Dict:
        """Aplica las reglas de conversión de unidades"""
        original_qty = line['quantity']
        original_unit = line['unit']
        converted_qty = original_qty
        converted_unit = original_unit
        conversion_note = ""

        # Conversión de libras a kilogramos
        if original_unit == 'LBR':
            converted_qty = original_qty / 2
            converted_unit = 'KG'
            conversion_note = f"Convertido de {original_qty} LBR a {converted_qty:.5f} KG"

        # Conversión de GRAMOS en descripción (busca "GRAMOS" o "GRAMO")
        gram_match = re.search(r'(\d+)\s*GRAMOS?', line['description'], re.IGNORECASE)
        if gram_match:
            grams = float(gram_match.group(1))
            converted_qty = (grams * original_qty) / 1000
            converted_unit = 'KG'
            conversion_note = f"Convertido: ({grams} gr × {original_qty}) ÷ 1000 = {converted_qty:.5f} KG"

        # Conversión de GRS en descripción (busca "GRS" o "grs")
        grs_match = re.search(r'(\d+)\s*GRS', line['description'], re.IGNORECASE)
        if grs_match:
            grams = float(grs_match.group(1))
            converted_qty = (grams * original_qty) / 1000
            converted_unit = 'KG'
            conversion_note = f"Convertido: ({grams} grs × {original_qty}) ÷ 1000 = {converted_qty:.5f} KG"

        # Mapeo de unidades estándar
        converted_unit = UNIT_MAP.get(converted_unit, converted_unit)

        return {
            **line,
            'converted_quantity': converted_qty,
            'converted_unit': converted_unit,
            'conversion_note': conversion_note
        }

    def create_reggis_excel(self, output_path: Path):
        """Crea el archivo Excel con formato REGGIS"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Facturas Procesadas"

        headers = [
            'N° Factura', 'Nombre Producto', 'Codigo Subyacente',
            'Unidad Medida en Kg,Un,Lt', 'Cantidad', 'Precio Unitario',
            'Fecha Factura', 'Fecha Pago', 'Nit Comprador', 'Nombre Comprador',
            'Nit Vendedor', 'Nombre Vendedor', 'Principal V,C', 'Municipio',
            'Iva', 'Descripción', 'Activa Factura', 'Activa Bodega', 'Incentivo',
            'Cantidad Original', 'Moneda', 'Total Sin IVA', 'Total IVA', 'Total Con IVA'
        ]

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for idx, line in enumerate(self.processed_lines, start=2):
            cantidad_convertida = f"{line['converted_quantity']:.5f}".replace('.', ',')
            cantidad_original = f"{line['quantity']:.5f}".replace('.', ',')
            precio_unitario = f"{line['price']:.5f}".replace('.', ',')

            row_data = [
                line['invoice_number'], line['description'], line['code'],
                line['converted_unit'], cantidad_convertida, precio_unitario,
                line['invoice_date'], line['payment_date'], line['customer_nit'],
                line['customer_name'], line['supplier_nit'], line['supplier_name'],
                'V', line['supplier_city'], line['iva_percent'], line['conversion_note'],
                '1', '1', '', cantidad_original, line['currency_id'],
                round(line['line_total'], 2), round(line['iva_amount'], 2),
                round(line['total_with_iva'], 2)
            ]

            for col, value in enumerate(row_data, start=1):
                cell = ws.cell(row=idx, column=col, value=value)
                cell.alignment = Alignment(vertical='center')

        column_widths = [15, 40, 15, 20, 15, 15, 12, 12, 15, 30, 15, 30, 12, 20, 8, 50, 12, 12, 10, 18, 10, 15, 12, 15]
        for col, width in enumerate(column_widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

        wb.save(output_path)

    def procesar(self) -> Path:
        """Ejecuta el procesamiento completo"""
        zip_files = list(self.carpeta_zip.glob("*.zip"))

        if not zip_files:
            raise Exception("No se encontraron archivos ZIP en la carpeta seleccionada")

        logger.info(f"CASA DEL AGRICULTOR: Encontrados {len(zip_files)} archivos ZIP")

        for zip_file in zip_files:
            logger.info(f"Procesando: {zip_file.name}")

            xml_content = self.extract_xml_from_zip(zip_file)
            if not xml_content:
                continue

            lines = self.parse_invoice_xml(xml_content)
            if not lines:
                logger.warning(f"No se extrajeron líneas de {zip_file.name}")
                continue

            for line in lines:
                processed_line = self.apply_conversion_rules(line)
                self.processed_lines.append(processed_line)

                if processed_line['conversion_note']:
                    logger.info(f"  {processed_line['conversion_note']}")

        if not self.processed_lines:
            raise Exception("No se procesaron líneas. Verifique los archivos ZIP.")

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = self.carpeta_salida / f"Facturas_REGGIS_CASA_DEL_AGRICULTOR_{timestamp}.xlsx"

        self.create_reggis_excel(output_file)
        logger.info(f"CASA DEL AGRICULTOR: Proceso completado - {len(self.processed_lines)} líneas")

        return self.carpeta_salida
