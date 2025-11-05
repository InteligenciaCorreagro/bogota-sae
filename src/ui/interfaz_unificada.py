"""
Interfaz que gestiona todos los clientes con botón Volver (PyQt6)
"""

import os
import platform
import logging
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path

from PyQt6.QtWidgets import (QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
                              QLabel, QPushButton, QProgressBar, QFileDialog,
                              QMessageBox, QListWidget, QDialog)
from PyQt6.QtCore import Qt, QThread, pyqtSignal
from PyQt6.QtGui import QFont

from config.constants import REGGIS_HEADERS
from utils.sharepoint_detector import DetectorSharePoint
from processors.seaboard_processor import ProcesadorSeaboard
from processors.casa_del_agricultor_processor import ProcesadorCasaDelAgricultor

logger = logging.getLogger(__name__)


class ProcesamientoThread(QThread):
    """Thread para ejecutar procesamiento en segundo plano"""
    finished = pyqtSignal(bool, str, object)  # success, message, result

    def __init__(self, procesador_func, *args):
        super().__init__()
        self.procesador_func = procesador_func
        self.args = args

    def run(self):
        try:
            result = self.procesador_func(*self.args)
            self.finished.emit(True, "Proceso completado exitosamente", result)
        except Exception as e:
            self.finished.emit(False, f"Error: {str(e)}", None)


class InterfazUnificada(QMainWindow):
    """Interfaz que gestiona todos los clientes con botón Volver"""

    def __init__(self, cliente: str, app):
        super().__init__()
        self.cliente = cliente
        self.app = app
        self.carpeta_entrada = None
        self.carpetas_sharepoint = []
        self.request_return = False
        self.procesamiento_thread = None

        if cliente == "SEABOARD":
            self.detectar_sharepoint()

        self.setup_ui()
        self.centrar_ventana()

    def centrar_ventana(self):
        """Centra la ventana en la pantalla"""
        screen = self.app.primaryScreen().geometry()
        window_geo = self.frameGeometry()
        center_point = screen.center()
        window_geo.moveCenter(center_point)
        self.move(window_geo.topLeft())

    def detectar_sharepoint(self):
        """Detecta carpetas de SharePoint sincronizadas"""
        try:
            self.carpetas_sharepoint = DetectorSharePoint.encontrar_carpetas_sharepoint()
            if self.carpetas_sharepoint:
                logger.info(f"Se encontraron {len(self.carpetas_sharepoint)} carpetas de SharePoint")
        except Exception as e:
            logger.error(f"Error detectando SharePoint: {str(e)}")
            self.carpetas_sharepoint = []

    def setup_ui(self):
        """Configura la interfaz de usuario"""
        self.setWindowTitle(f"Procesador de Facturas - {self.cliente}")
        self.setMinimumSize(900, 600)

        # Widget central
        central_widget = QWidget()
        self.setCentralWidget(central_widget)

        # Layout principal
        main_layout = QVBoxLayout()
        main_layout.setContentsMargins(20, 20, 20, 20)
        main_layout.setSpacing(15)
        central_widget.setLayout(main_layout)

        # Botón Volver
        btn_volver = QPushButton("← Volver")
        btn_volver.setMaximumWidth(100)
        btn_volver.setStyleSheet("""
            QPushButton {
                background-color: #95a5a6;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 8px;
            }
            QPushButton:hover {
                background-color: #7f8c8d;
            }
        """)
        btn_volver.clicked.connect(self.volver_al_selector)
        main_layout.addWidget(btn_volver, alignment=Qt.AlignmentFlag.AlignLeft)

        # Título
        titulo = QLabel(f"PROCESADOR - {self.cliente}")
        titulo.setFont(QFont("Arial", 18, QFont.Weight.Bold))
        titulo.setAlignment(Qt.AlignmentFlag.AlignCenter)
        titulo.setStyleSheet("color: #2c3e50; padding: 10px;")
        main_layout.addWidget(titulo)

        # Layout de botones según cliente
        if self.cliente == "SEABOARD":
            self.setup_botones_seaboard(main_layout)
        elif self.cliente == "CASA_DEL_AGRICULTOR":
            self.setup_botones_casa(main_layout)
        elif self.cliente == "LACTALIS_COMPRAS":
            self.setup_botones_lactalis(main_layout)

        # Barra de progreso
        self.progress = QProgressBar()
        self.progress.setMinimum(0)
        self.progress.setMaximum(0)  # Modo indeterminado
        self.progress.setVisible(False)
        self.progress.setStyleSheet("""
            QProgressBar {
                border: 2px solid #bdc3c7;
                border-radius: 5px;
                text-align: center;
            }
            QProgressBar::chunk {
                background-color: #3498db;
            }
        """)
        main_layout.addWidget(self.progress)

        # Label de estado
        self.estado_label = QLabel("")
        self.estado_label.setFont(QFont("Arial", 10))
        self.estado_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.estado_label.setStyleSheet("padding: 10px;")
        main_layout.addWidget(self.estado_label)

        # Espaciador
        main_layout.addStretch()

        # Botón Cerrar
        btn_cerrar = QPushButton("Cerrar")
        btn_cerrar.setMaximumWidth(100)
        btn_cerrar.setStyleSheet("""
            QPushButton {
                background-color: #e74c3c;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 8px;
            }
            QPushButton:hover {
                background-color: #c0392b;
            }
        """)
        btn_cerrar.clicked.connect(self.close)

        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        btn_layout.addWidget(btn_cerrar)
        main_layout.addLayout(btn_layout)

    def setup_botones_seaboard(self, layout):
        """Configura los botones específicos para SEABOARD"""
        if self.carpetas_sharepoint:
            info_sp = QLabel(f"Se detectaron {len(self.carpetas_sharepoint)} carpeta(s) de SharePoint sincronizada(s)")
            info_sp.setStyleSheet("color: #27ae60; padding: 5px;")
            info_sp.setFont(QFont("Arial", 10))
            layout.addWidget(info_sp)

            btn_sharepoint = QPushButton("BUSCAR EN SHAREPOINT SINCRONIZADO")
            btn_sharepoint.setMinimumHeight(60)
            btn_sharepoint.setFont(QFont("Arial", 12, QFont.Weight.Bold))
            btn_sharepoint.setStyleSheet("""
                QPushButton {
                    background-color: #0078D4;
                    color: white;
                    border: none;
                    border-radius: 8px;
                    padding: 15px;
                }
                QPushButton:hover {
                    background-color: #005a9e;
                }
            """)
            btn_sharepoint.clicked.connect(self.seleccionar_desde_sharepoint)
            layout.addWidget(btn_sharepoint)

        btn_local = QPushButton("BUSCAR EN CARPETA LOCAL")
        btn_local.setMinimumHeight(60)
        btn_local.setFont(QFont("Arial", 12, QFont.Weight.Bold))
        btn_local.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border: none;
                border-radius: 8px;
                padding: 15px;
            }
            QPushButton:hover {
                background-color: #388E3C;
            }
        """)
        btn_local.clicked.connect(self.seleccionar_y_procesar_seaboard)
        layout.addWidget(btn_local)

    def setup_botones_casa(self, layout):
        """Configura los botones específicos para CASA DEL AGRICULTOR"""
        info = QLabel("Seleccione la carpeta que contiene los archivos ZIP de facturas")
        info.setFont(QFont("Arial", 10))
        info.setStyleSheet("padding: 5px;")
        layout.addWidget(info)

        btn_procesar = QPushButton("SELECCIONAR CARPETA CON ARCHIVOS ZIP")
        btn_procesar.setMinimumHeight(60)
        btn_procesar.setFont(QFont("Arial", 12, QFont.Weight.Bold))
        btn_procesar.setStyleSheet("""
            QPushButton {
                background-color: #27ae60;
                color: white;
                border: none;
                border-radius: 8px;
                padding: 15px;
            }
            QPushButton:hover {
                background-color: #1e8449;
            }
        """)
        btn_procesar.clicked.connect(self.seleccionar_y_procesar_casa)
        layout.addWidget(btn_procesar)

    def setup_botones_lactalis(self, layout):
        """Configura los botones específicos para LACTALIS COMPRAS"""
        info = QLabel("Seleccione la carpeta que contiene los archivos XML de Lactalis Compras")
        info.setFont(QFont("Arial", 10))
        info.setStyleSheet("padding: 5px;")
        layout.addWidget(info)

        btn_procesar = QPushButton("SELECCIONAR CARPETA CON ARCHIVOS XML")
        btn_procesar.setMinimumHeight(60)
        btn_procesar.setFont(QFont("Arial", 12, QFont.Weight.Bold))
        btn_procesar.setStyleSheet("""
            QPushButton {
                background-color: #e74c3c;
                color: white;
                border: none;
                border-radius: 8px;
                padding: 15px;
            }
            QPushButton:hover {
                background-color: #c0392b;
            }
        """)
        btn_procesar.clicked.connect(self.seleccionar_y_procesar_lactalis)
        layout.addWidget(btn_procesar)

    def volver_al_selector(self):
        """Marca la intención de volver y cierra la ventana actual"""
        self.request_return = True
        self.close()

    def seleccionar_desde_sharepoint(self):
        """Muestra un diálogo para seleccionar una carpeta de SharePoint"""
        if not self.carpetas_sharepoint:
            QMessageBox.information(self, "No hay carpetas", "No se detectaron carpetas de SharePoint")
            return

        dialog = QDialog(self)
        dialog.setWindowTitle("Seleccionar Carpeta de SharePoint")
        dialog.setMinimumSize(700, 500)

        layout = QVBoxLayout()

        titulo = QLabel("Carpetas de SharePoint Detectadas")
        titulo.setFont(QFont("Arial", 14, QFont.Weight.Bold))
        layout.addWidget(titulo)

        listbox = QListWidget()
        for carpeta in self.carpetas_sharepoint:
            listbox.addItem(str(carpeta))
        layout.addWidget(listbox)

        btn_seleccionar = QPushButton("Procesar Carpeta Seleccionada")
        btn_seleccionar.setStyleSheet("""
            QPushButton {
                background-color: #0078D4;
                color: white;
                padding: 10px;
                border-radius: 5px;
            }
        """)

        def seleccionar():
            if listbox.currentRow() >= 0:
                carpeta_sel = self.carpetas_sharepoint[listbox.currentRow()]
                dialog.close()
                self.procesar_carpeta_xml(carpeta_sel)
            else:
                QMessageBox.warning(dialog, "Selección requerida", "Seleccione una carpeta")

        btn_seleccionar.clicked.connect(seleccionar)
        layout.addWidget(btn_seleccionar)

        dialog.setLayout(layout)
        dialog.exec()

    def seleccionar_y_procesar_seaboard(self):
        """Permite seleccionar una carpeta local con archivos XML"""
        carpeta = QFileDialog.getExistingDirectory(
            self,
            "Seleccione la carpeta con archivos XML",
            "",
            QFileDialog.Option.ShowDirsOnly
        )

        if carpeta:
            self.procesar_carpeta_xml(Path(carpeta))

    def procesar_carpeta_xml(self, carpeta: Path):
        """Procesa la carpeta con archivos XML"""
        archivos_xml = list(carpeta.glob("*.xml"))

        if not archivos_xml:
            QMessageBox.critical(self, "Sin archivos", "No se encontraron archivos XML")
            return

        respuesta = QMessageBox.question(
            self,
            "Confirmar",
            f"Se encontraron {len(archivos_xml)} archivo(s) XML.\n\n¿Procesar ahora?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )

        if respuesta == QMessageBox.StandardButton.Yes:
            self.carpeta_entrada = carpeta
            self.iniciar_procesamiento_seaboard()

    def seleccionar_y_procesar_casa(self):
        """Permite seleccionar una carpeta con archivos ZIP"""
        carpeta = QFileDialog.getExistingDirectory(
            self,
            "Seleccione la carpeta con archivos ZIP",
            "",
            QFileDialog.Option.ShowDirsOnly
        )

        if not carpeta:
            return

        self.carpeta_entrada = Path(carpeta)
        zip_files = list(self.carpeta_entrada.glob("*.zip"))

        if not zip_files:
            QMessageBox.critical(self, "Sin archivos", "No se encontraron archivos ZIP")
            return

        respuesta = QMessageBox.question(
            self,
            "Confirmar",
            f"Se encontraron {len(zip_files)} archivo(s) ZIP.\n\n¿Procesar ahora?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )

        if respuesta == QMessageBox.StandardButton.Yes:
            self.iniciar_procesamiento_casa()

    def seleccionar_y_procesar_lactalis(self):
        """Permite seleccionar una carpeta con archivos XML de Lactalis"""
        carpeta = QFileDialog.getExistingDirectory(
            self,
            "Seleccione la carpeta con archivos XML de Lactalis",
            "",
            QFileDialog.Option.ShowDirsOnly
        )

        if not carpeta:
            return

        self.carpeta_entrada = Path(carpeta)
        xml_files = list(self.carpeta_entrada.glob("*.xml"))

        if not xml_files:
            QMessageBox.critical(self, "Sin archivos", "No se encontraron archivos XML")
            return

        respuesta = QMessageBox.question(
            self,
            "Confirmar",
            f"Se encontraron {len(xml_files)} archivo(s) XML.\n\n¿Procesar ahora?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )

        if respuesta == QMessageBox.StandardButton.Yes:
            self.iniciar_procesamiento_lactalis()

    def buscar_o_crear_plantilla(self) -> Path:
        """Busca o crea la plantilla REGGIS"""
        script_dir = Path(__file__).parent.parent.parent
        plantilla = script_dir / "Plantilla_REGGIS.xlsx"

        if not plantilla.exists():
            self.crear_plantilla_base(plantilla)

        return plantilla

    def crear_plantilla_base(self, ruta: Path):
        """Crea una plantilla base de Excel"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Datos"

        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col, header in enumerate(REGGIS_HEADERS, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        wb.save(ruta)

    def iniciar_procesamiento_seaboard(self):
        """Inicia el procesamiento para SEABOARD"""
        self.progress.setVisible(True)
        self.estado_label.setText("Procesando archivos...")
        self.estado_label.setStyleSheet("color: #f39c12; padding: 10px;")

        plantilla = self.buscar_o_crear_plantilla()
        procesador = ProcesadorSeaboard(self.carpeta_entrada, plantilla)

        self.procesamiento_thread = ProcesamientoThread(procesador.procesar)
        self.procesamiento_thread.finished.connect(self.procesamiento_finalizado)
        self.procesamiento_thread.start()

    def iniciar_procesamiento_casa(self):
        """Inicia el procesamiento para CASA DEL AGRICULTOR"""
        self.progress.setVisible(True)
        self.estado_label.setText("Procesando archivos...")
        self.estado_label.setStyleSheet("color: #f39c12; padding: 10px;")

        carpeta_salida = self.carpeta_entrada.parent / "Resultados_CASA_DEL_AGRICULTOR"
        carpeta_salida.mkdir(exist_ok=True)

        procesador = ProcesadorCasaDelAgricultor(self.carpeta_entrada, carpeta_salida)

        self.procesamiento_thread = ProcesamientoThread(procesador.procesar)
        self.procesamiento_thread.finished.connect(self.procesamiento_finalizado)
        self.procesamiento_thread.start()

    def iniciar_procesamiento_lactalis(self):
        """Inicia el procesamiento para LACTALIS COMPRAS"""
        self.progress.setVisible(True)
        self.estado_label.setText("Procesando archivos de Lactalis...")
        self.estado_label.setStyleSheet("color: #f39c12; padding: 10px;")

        # Por ahora usamos el mismo procesador que SEABOARD (se puede personalizar después)
        plantilla = self.buscar_o_crear_plantilla()
        procesador = ProcesadorSeaboard(self.carpeta_entrada, plantilla)

        self.procesamiento_thread = ProcesamientoThread(procesador.procesar)
        self.procesamiento_thread.finished.connect(self.procesamiento_finalizado)
        self.procesamiento_thread.start()

    def procesamiento_finalizado(self, success, message, result):
        """Callback cuando el procesamiento termina"""
        self.progress.setVisible(False)

        if success:
            self.estado_label.setText("Proceso completado exitosamente")
            self.estado_label.setStyleSheet("color: #27ae60; padding: 10px;")

            respuesta = QMessageBox.question(
                self,
                "Éxito",
                f"{message}\n\nArchivos guardados en:\n{result.name}\n\n¿Abrir carpeta?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )

            if respuesta == QMessageBox.StandardButton.Yes:
                self.abrir_carpeta(result)
        else:
            self.estado_label.setText("Error en el procesamiento")
            self.estado_label.setStyleSheet("color: #e74c3c; padding: 10px;")
            QMessageBox.critical(self, "Error", message)

    def abrir_carpeta(self, carpeta: Path):
        """Abre la carpeta en el explorador de archivos del sistema"""
        if platform.system() == 'Windows':
            os.startfile(carpeta)
        elif platform.system() == 'Darwin':
            os.system(f'open "{carpeta}"')
        else:
            os.system(f'xdg-open "{carpeta}"')
