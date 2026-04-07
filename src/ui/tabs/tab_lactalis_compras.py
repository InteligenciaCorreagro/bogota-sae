"""
Tab para procesamiento de facturas LACTALIS COMPRAS
Procesa archivos ZIP y XML con configuración fija según especificaciones de Lactalis
"""

import os
import platform
import logging
import openpyxl
from pathlib import Path
from openpyxl.styles import Font, PatternFill, Alignment

from PyQt6.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QLabel,
                              QPushButton, QProgressBar, QFileDialog,
                              QMessageBox, QGroupBox)
from PyQt6.QtCore import Qt, QThread, pyqtSignal
from PyQt6.QtGui import QFont

from config.constants import REGGIS_HEADERS
from processors.lactalis_processor import ProcesadorLactalis

logger = logging.getLogger(__name__)


class ProcesamientoThread(QThread):
    """Thread para ejecutar procesamiento en segundo plano"""
    finished = pyqtSignal(bool, str, object)  # success, message, result

    def __init__(self, procesador_func, *args):
        super().__init__()
        self.procesador_func = procesador_func
        self.args = args

    def run(self):
        try:
            result = self.procesador_func(*self.args)
            self.finished.emit(True, "Proceso completado exitosamente", result)
        except Exception as e:
            logger.error(f"Error en procesamiento: {str(e)}", exc_info=True)
            self.finished.emit(False, f"Error: {str(e)}", None)


class TabLactalisCompras(QWidget):
    """
    Tab para procesamiento de facturas LACTALIS COMPRAS

    Procesamiento simplificado con configuración fija:
    - Producto: Siempre "LECHE CRUDA"
    - Código Subyacente: Siempre "SPN-1"
    - Unidad: Siempre "Lt" (litros)
    - Activa Factura/Bodega: Siempre "1"
    - Comprador: Siempre Lactalis (NIT fijo)
    """

    def __init__(self):
        super().__init__()
        self.carpeta_entrada = None
        self.procesamiento_thread = None
        self.setup_ui()

    def setup_ui(self):
        """Configura la interfaz del tab"""
        # Layout principal
        main_layout = QVBoxLayout()
        main_layout.setContentsMargins(20, 20, 20, 20)
        main_layout.setSpacing(15)

        # --- Título ---
        titulo = QLabel("🥛 PROCESADOR LACTALIS COMPRAS")
        titulo.setFont(QFont("Arial", 16, QFont.Weight.Bold))
        titulo.setAlignment(Qt.AlignmentFlag.AlignCenter)
        titulo.setStyleSheet("color: #e74c3c; padding: 10px;")
        main_layout.addWidget(titulo)

        # --- Descripción ---
        descripcion = QLabel(
            "Procesa archivos ZIP y XML de facturas de Lactalis Compras.\n"
            "Configuración fija: LECHE CRUDA • SPN-1 • Litros • Activa=1"
        )
        descripcion.setFont(QFont("Arial", 9))
        descripcion.setAlignment(Qt.AlignmentFlag.AlignCenter)
        descripcion.setStyleSheet("color: #555; padding: 5px;")
        descripcion.setWordWrap(True)
        main_layout.addWidget(descripcion)

        # --- Group Box: Selección de Archivos ---
        group_archivos = QGroupBox("📁 Seleccionar Archivos")
        group_archivos.setFont(QFont("Arial", 10, QFont.Weight.Bold))
        archivos_layout = QVBoxLayout()
        archivos_layout.setSpacing(10)

        # Información
        info_label = QLabel(
            "ℹ️ Seleccione la carpeta que contiene archivos ZIP y/o XML.\n"
            "Cada ZIP debe contener un archivo XML; los XML sueltos también se procesan."
        )
        info_label.setFont(QFont("Arial", 9))
        info_label.setStyleSheet("color: #34495e; padding: 5px;")
        info_label.setWordWrap(True)
        archivos_layout.addWidget(info_label)

        # Botón seleccionar carpeta
        btn_seleccionar = QPushButton("📂 SELECCIONAR CARPETA CON ARCHIVOS")
        btn_seleccionar.setMinimumHeight(60)
        btn_seleccionar.setFont(QFont("Arial", 11, QFont.Weight.Bold))
        btn_seleccionar.setStyleSheet("""
            QPushButton {
                background-color: #e74c3c;
                color: white;
                border: none;
                border-radius: 8px;
                padding: 15px;
            }
            QPushButton:hover {
                background-color: #c0392b;
            }
            QPushButton:pressed {
                background-color: #a93226;
            }
        """)
        btn_seleccionar.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_seleccionar.clicked.connect(self.seleccionar_carpeta)
        archivos_layout.addWidget(btn_seleccionar)

        group_archivos.setLayout(archivos_layout)
        main_layout.addWidget(group_archivos)

        # --- Barra de Progreso ---
        self.progress = QProgressBar()
        self.progress.setMinimum(0)
        self.progress.setMaximum(0)  # Modo indeterminado
        self.progress.setVisible(False)
        self.progress.setStyleSheet("""
            QProgressBar {
                border: 2px solid #bdc3c7;
                border-radius: 5px;
                text-align: center;
                height: 25px;
            }
            QProgressBar::chunk {
                background-color: #e74c3c;
            }
        """)
        main_layout.addWidget(self.progress)

        # --- Label de Estado ---
        self.estado_label = QLabel("")
        self.estado_label.setFont(QFont("Arial", 10))
        self.estado_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.estado_label.setStyleSheet("padding: 10px;")
        main_layout.addWidget(self.estado_label)

        # Espaciador
        main_layout.addStretch()

        # --- Información adicional ---
        info_footer = QLabel(
            "💡 Se procesan ZIP y XML sueltos automáticamente.\n"
            "Los resultados se guardarán en formato Excel REGGIS."
        )
        info_footer.setFont(QFont("Arial", 8))
        info_footer.setAlignment(Qt.AlignmentFlag.AlignCenter)
        info_footer.setStyleSheet("color: #7f8c8d; padding: 10px;")
        info_footer.setWordWrap(True)
        main_layout.addWidget(info_footer)

        self.setLayout(main_layout)

    def seleccionar_carpeta(self):
        """Permite seleccionar una carpeta con archivos ZIP y/o XML"""
        carpeta = QFileDialog.getExistingDirectory(
            self,
            "Seleccione la carpeta con archivos ZIP y/o XML de Lactalis Compras",
            "",
            QFileDialog.Option.ShowDirsOnly
        )

        if not carpeta:
            return

        self.carpeta_entrada = Path(carpeta)

        # Buscar archivos ZIP y XML
        zip_files = list(self.carpeta_entrada.glob("*.zip"))
        xml_files = list(self.carpeta_entrada.glob("*.xml"))
        total_archivos = len(zip_files) + len(xml_files)

        if total_archivos == 0:
            QMessageBox.critical(
                self,
                "Sin archivos",
                "No se encontraron archivos ZIP ni XML en la carpeta seleccionada"
            )
            return

        # Confirmar procesamiento
        respuesta = QMessageBox.question(
            self,
            "Confirmar procesamiento",
            f"Se encontraron:\n"
            f"  • {len(zip_files)} archivo(s) ZIP\n"
            f"  • {len(xml_files)} archivo(s) XML\n"
            f"  • {total_archivos} archivos TOTALES\n\n"
            f"Carpeta: {self.carpeta_entrada.name}\n\n"
            f"Configuración fija: LECHE CRUDA • SPN-1 • Litros • Activa=1\n\n"
            f"¿Procesar ahora?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )

        if respuesta == QMessageBox.StandardButton.Yes:
            self.iniciar_procesamiento()

    def buscar_o_crear_plantilla(self) -> Path:
        """Busca o crea la plantilla REGGIS Excel"""
        script_dir = Path(__file__).parent.parent.parent.parent
        plantilla = script_dir / "Plantilla_REGGIS.xlsx"

        if not plantilla.exists():
            logger.info(f"Plantilla no encontrada. Creando en: {plantilla}")
            self.crear_plantilla_base(plantilla)

        return plantilla

    def crear_plantilla_base(self, ruta: Path):
        """Crea una plantilla base de Excel con formato REGGIS"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Datos"

        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col, header in enumerate(REGGIS_HEADERS, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        wb.save(ruta)
        logger.info(f"Plantilla creada exitosamente: {ruta}")

    def iniciar_procesamiento(self):
        """Inicia el procesamiento en segundo plano"""
        self.progress.setVisible(True)
        self.estado_label.setText("⏳ Procesando archivos de LACTALIS COMPRAS...")
        self.estado_label.setStyleSheet("color: #f39c12; padding: 10px; font-weight: bold;")

        # Obtener plantilla
        plantilla = self.buscar_o_crear_plantilla()

        # Usar procesador específico de Lactalis
        procesador = ProcesadorLactalis(self.carpeta_entrada, plantilla)

        # Iniciar thread de procesamiento
        self.procesamiento_thread = ProcesamientoThread(procesador.procesar)
        self.procesamiento_thread.finished.connect(self.procesamiento_finalizado)
        self.procesamiento_thread.start()

    def procesamiento_finalizado(self, success: bool, message: str, result):
        """
        Callback cuando el procesamiento termina

        Args:
            success: True si fue exitoso
            message: Mensaje de resultado
            result: Path de carpeta de resultados o None
        """
        self.progress.setVisible(False)

        if success:
            self.estado_label.setText("✅ Proceso completado exitosamente")
            self.estado_label.setStyleSheet("color: #27ae60; padding: 10px; font-weight: bold;")

            respuesta = QMessageBox.question(
                self,
                "Éxito",
                f"{message}\n\n"
                f"📂 Archivos guardados en:\n{result.name}\n\n"
                f"¿Desea abrir la carpeta de resultados?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )

            if respuesta == QMessageBox.StandardButton.Yes:
                self.abrir_carpeta(result)
        else:
            self.estado_label.setText("❌ Error en el procesamiento")
            self.estado_label.setStyleSheet("color: #e74c3c; padding: 10px; font-weight: bold;")
            QMessageBox.critical(
                self,
                "Error de procesamiento",
                f"Ocurrió un error durante el procesamiento:\n\n{message}"
            )

    def abrir_carpeta(self, carpeta: Path):
        """
        Abre la carpeta de resultados en el explorador del sistema

        Args:
            carpeta: Path a la carpeta a abrir
        """
        try:
            if platform.system() == 'Windows':
                os.startfile(carpeta)
            elif platform.system() == 'Darwin':  # macOS
                os.system(f'open "{carpeta}"')
            else:  # Linux
                os.system(f'xdg-open "{carpeta}"')
        except Exception as e:
            logger.error(f"Error abriendo carpeta: {str(e)}")
            QMessageBox.warning(
                self,
                "Error",
                f"No se pudo abrir la carpeta automáticamente:\n{str(e)}\n\n"
                f"Ubicación: {carpeta}"
            )


# GUÍA DE PERSONALIZACIÓN PARA LACTALIS COMPRAS
# ===============================================
#
# 1. CREAR PROCESADOR ESPECÍFICO:
#    Archivo: src/processors/lactalis_processor.py
#
#    class ProcesadorLactalis:
#        def __init__(self, carpeta_xml, plantilla_excel):
#            ...
#
#        def procesar(self):
#            # Lógica específica de Lactalis
#            pass
#
# 2. CREAR EXTRACTOR ESPECÍFICO:
#    Archivo: src/extractors/lactalis_extractor.py
#
#    class FacturaExtractorLactalis:
#        def extraer_datos(self, xml_path):
#            # Extracción específica de campos Lactalis
#            pass
#
# 3. ACTUALIZAR ESTE ARCHIVO:
#    - Importar ProcesadorLactalis
#    - Reemplazar línea 217: procesador = ProcesadorLactalis(...)
#    - Agregar validaciones específicas
#    - Personalizar campos de configuración
#
# 4. AGREGAR LÓGICA DE FILTROS:
#    - Implementar filtro por fecha en el procesador
#    - Implementar filtro por proveedor
#    - Implementar manejo de facturas anuladas
