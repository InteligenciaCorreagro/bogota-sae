"""
Tab para procesamiento de facturas LACTALIS VENTAS
Optimizado para grandes volúmenes (20,000+ XML) con progreso detallado
"""

import os
import platform
import logging
import openpyxl
from pathlib import Path
from openpyxl.styles import Font, PatternFill, Alignment

from PyQt6.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QLabel,
                              QPushButton, QProgressBar, QFileDialog,
                              QMessageBox, QGroupBox, QTextEdit, QCheckBox,
                              QScrollArea)
from PyQt6.QtCore import Qt, QThread, pyqtSignal
from PyQt6.QtGui import QFont

from src.config.constants import REGGIS_HEADERS
from processors.lactalis_ventas_processor import ProcesadorLactalisVentas
from src.database.lactalis_database import LactalisDatabase
from src.database.excel_importer import ExcelImporter, ExcelImporterError

logger = logging.getLogger(__name__)


class ProcesamientoThread(QThread):
    """Thread para ejecutar procesamiento en segundo plano con reportes de progreso"""
    progress = pyqtSignal(int, int, str)  # processed, total, message
    finished = pyqtSignal(bool, str, object)  # success, message, result

    def __init__(self, procesador_func, *args):
        super().__init__()
        self.procesador_func = procesador_func
        self.args = args

    def run(self):
        try:
            result = self.procesador_func(*self.args)
            self.finished.emit(True, "Proceso completado exitosamente", result)
        except Exception as e:
            logger.error(f"Error en procesamiento: {str(e)}", exc_info=True)
            self.finished.emit(False, f"Error: {str(e)}", None)


class TabLactalisVentas(QWidget):
    """
    Tab para procesamiento de facturas LACTALIS VENTAS
    
    REGLAS DE NEGOCIO IMPLEMENTADAS:
    • Solo procesa FACTURAS (Invoice), NO notas crédito/débito
    • Cantidad debe ser > 0
    • Precio unitario debe ser > 0
    • Total debe ser > 0
    
    OPTIMIZACIONES:
    • Procesamiento por lotes para grandes volúmenes
    • Progreso detallado en tiempo real
    • Estadísticas completas de procesamiento
    • Validaciones tempranas para descartar archivos inválidos
    """

    def __init__(self):
        super().__init__()
        self.carpeta_entrada = None
        self.procesamiento_thread = None
        self.db = None
        self.db_inicializada = False  # Flag para saber si ya se intentó inicializar
        self.db_error = None  # Para guardar error de inicialización si ocurre

        # NO inicializar la BD aquí - usar carga lazy
        # La BD se inicializará la primera vez que se necesite
        logger.info("TabLactalisVentas iniciado - BD se cargará cuando sea necesaria")

        self.setup_ui()

    def setup_ui(self):
        """Configura la interfaz del tab"""
        # Layout principal con scroll para evitar cortes
        root_layout = QVBoxLayout()
        root_layout.setContentsMargins(0, 0, 0, 0)

        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        scroll_area.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)

        content_widget = QWidget()
        main_layout = QVBoxLayout()
        main_layout.setContentsMargins(20, 20, 20, 20)
        main_layout.setSpacing(15)

        # --- Título ---
        titulo = QLabel("💼 PROCESADOR LACTALIS VENTAS")
        titulo.setFont(QFont("Arial", 16, QFont.Weight.Bold))
        titulo.setAlignment(Qt.AlignmentFlag.AlignCenter)
        titulo.setStyleSheet("color: #2980b9; padding: 10px;")
        main_layout.addWidget(titulo)

        # --- Descripción ---
        descripcion = QLabel(
            "Procesa archivos ZIP y XML de FACTURAS de ventas de Lactalis.\n"
            "⚠️ Solo facturas (Invoice) - NO se procesan notas crédito/débito"
        )
        descripcion.setFont(QFont("Arial", 9))
        descripcion.setAlignment(Qt.AlignmentFlag.AlignCenter)
        descripcion.setStyleSheet("color: #555; padding: 5px;")
        descripcion.setWordWrap(True)
        main_layout.addWidget(descripcion)

        # --- Group Box: Gestión de Base de Datos ---
        group_db = QGroupBox("🗄️ Gestión de Base de Datos")
        group_db.setFont(QFont("Arial", 10, QFont.Weight.Bold))
        db_layout = QVBoxLayout()
        db_layout.setSpacing(10)

        # Info de BD - mostrar mensaje inicial, se actualizará después
        db_text = (
            "⚠️ IMPORTANTE: Debes importar materiales y clientes ANTES de procesar\n"
            "Base de datos: Se cargará al hacer clic en 'Importar Materiales' o 'Importar Clientes'"
        )

        db_info = QLabel(db_text)
        db_info.setFont(QFont("Arial", 9))
        db_info.setStyleSheet("color: #34495e; padding: 5px;")
        db_info.setWordWrap(True)
        db_layout.addWidget(db_info)
        self.db_info_label = db_info

        # Botones de importación horizontal
        btn_layout = QHBoxLayout()

        # Botón importar materiales
        btn_importar_materiales = QPushButton("📦 Importar Materiales")
        btn_importar_materiales.setMinimumHeight(40)
        btn_importar_materiales.setFont(QFont("Arial", 10))
        btn_importar_materiales.setStyleSheet("""
            QPushButton {
                background-color: #16a085;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 10px;
            }
            QPushButton:hover {
                background-color: #138d75;
            }
        """)
        btn_importar_materiales.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_importar_materiales.clicked.connect(self.importar_materiales)
        btn_layout.addWidget(btn_importar_materiales)

        # Botón importar clientes
        btn_importar_clientes = QPushButton("👥 Importar Clientes")
        btn_importar_clientes.setMinimumHeight(40)
        btn_importar_clientes.setFont(QFont("Arial", 10))
        btn_importar_clientes.setStyleSheet("""
            QPushButton {
                background-color: #2980b9;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 10px;
            }
            QPushButton:hover {
                background-color: #21618c;
            }
        """)
        btn_importar_clientes.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_importar_clientes.clicked.connect(self.importar_clientes)
        btn_layout.addWidget(btn_importar_clientes)

        db_layout.addLayout(btn_layout)

        # Nota sobre formato
        formato_note = QLabel(
            "💡 Materiales: CODIGO, DESCRIPCION, SOCIEDAD (usar: 'Parmalat' o 'Proleche')\n"
            "💡 Clientes: Cód.Padre, Nombre Código Padre, NIT, Se Registra (NIT o NO NIT)"
        )
        formato_note.setFont(QFont("Arial", 8))
        formato_note.setStyleSheet("color: #7f8c8d; padding: 5px;")
        formato_note.setWordWrap(True)
        db_layout.addWidget(formato_note)

        # Checkboxes para validaciones
        validaciones_layout = QHBoxLayout()

        self.checkbox_validar_materiales = QCheckBox("✓ Validar materiales contra BD (activar solo si importaste materiales)")
        self.checkbox_validar_materiales.setFont(QFont("Arial", 9))
        self.checkbox_validar_materiales.setStyleSheet("color: #2c3e50;")
        self.checkbox_validar_materiales.setChecked(True)  # Activado por defecto para forzar validación
        validaciones_layout.addWidget(self.checkbox_validar_materiales)

        self.checkbox_validar_clientes = QCheckBox("✓ Validar clientes contra BD (activar solo si importaste clientes)")
        self.checkbox_validar_clientes.setFont(QFont("Arial", 9))
        self.checkbox_validar_clientes.setStyleSheet("color: #2c3e50;")
        self.checkbox_validar_clientes.setChecked(True)  # Activado por defecto para forzar validación
        validaciones_layout.addWidget(self.checkbox_validar_clientes)

        db_layout.addLayout(validaciones_layout)

        group_db.setLayout(db_layout)
        main_layout.addWidget(group_db)

        # --- Group Box: Reglas de Negocio ---
        group_reglas = QGroupBox("📋 Reglas de Validación")
        group_reglas.setFont(QFont("Arial", 10, QFont.Weight.Bold))
        reglas_layout = QVBoxLayout()
        reglas_layout.setSpacing(5)

        reglas_text = QLabel(
            "✓ Solo FACTURAS (Invoice) - se rechazan notas crédito/débito\n"
            "✓ Cantidad > 0 (se rechazan líneas con cantidad en 0 o negativa)\n"
            "✓ Precio unitario > 0 (se rechazan líneas con precio en 0 o negativo)\n"
            "✓ Total > 0 (se rechazan líneas con total en 0 o negativo)"
        )
        reglas_text.setFont(QFont("Arial", 9))
        reglas_text.setStyleSheet("color: #27ae60; padding: 5px;")
        reglas_text.setWordWrap(True)
        reglas_layout.addWidget(reglas_text)

        group_reglas.setLayout(reglas_layout)
        main_layout.addWidget(group_reglas)

        # --- Group Box: Selección de Archivos ---
        group_archivos = QGroupBox("📁 Seleccionar Archivos")
        group_archivos.setFont(QFont("Arial", 10, QFont.Weight.Bold))
        archivos_layout = QVBoxLayout()
        archivos_layout.setSpacing(10)

        info_label = QLabel(
            "ℹ️ Seleccione la carpeta con archivos ZIP y/o XML de facturas.\n"
            "Optimizado para procesar 20,000+ archivos"
        )
        info_label.setFont(QFont("Arial", 9))
        info_label.setStyleSheet("color: #34495e; padding: 5px;")
        info_label.setWordWrap(True)
        archivos_layout.addWidget(info_label)

        # Botón seleccionar carpeta
        btn_seleccionar = QPushButton("📂 SELECCIONAR CARPETA CON ARCHIVOS")
        btn_seleccionar.setMinimumHeight(60)
        btn_seleccionar.setFont(QFont("Arial", 11, QFont.Weight.Bold))
        btn_seleccionar.setStyleSheet("""
            QPushButton {
                background-color: #2980b9;
                color: white;
                border: none;
                border-radius: 8px;
                padding: 15px;
            }
            QPushButton:hover {
                background-color: #21618c;
            }
            QPushButton:pressed {
                background-color: #1b4f72;
            }
        """)
        btn_seleccionar.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_seleccionar.clicked.connect(self.seleccionar_carpeta)
        archivos_layout.addWidget(btn_seleccionar)

        group_archivos.setLayout(archivos_layout)
        main_layout.addWidget(group_archivos)

        # --- Barra de Progreso ---
        self.progress = QProgressBar()
        self.progress.setMinimum(0)
        self.progress.setMaximum(100)
        self.progress.setValue(0)
        self.progress.setVisible(False)
        self.progress.setStyleSheet("""
            QProgressBar {
                border: 2px solid #bdc3c7;
                border-radius: 5px;
                text-align: center;
                height: 25px;
            }
            QProgressBar::chunk {
                background-color: #2980b9;
            }
        """)
        main_layout.addWidget(self.progress)

        # --- Label de Estado ---
        self.estado_label = QLabel("")
        self.estado_label.setFont(QFont("Arial", 10))
        self.estado_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.estado_label.setStyleSheet("padding: 10px;")
        main_layout.addWidget(self.estado_label)

        # --- Consola de Log ---
        group_log = QGroupBox("📊 Progreso Detallado")
        group_log.setFont(QFont("Arial", 10, QFont.Weight.Bold))
        log_layout = QVBoxLayout()
        
        self.log_console = QTextEdit()
        self.log_console.setReadOnly(True)
        self.log_console.setMaximumHeight(150)
        self.log_console.setFont(QFont("Courier", 9))
        self.log_console.setStyleSheet("""
            QTextEdit {
                background-color: #2c3e50;
                color: #ecf0f1;
                border: 1px solid #34495e;
                border-radius: 4px;
            }
        """)
        log_layout.addWidget(self.log_console)
        
        group_log.setLayout(log_layout)
        main_layout.addWidget(group_log)

        # Espaciador
        main_layout.addStretch()

        # --- Información adicional ---
        info_footer = QLabel(
            "💡 Optimizado para grandes volúmenes. Procesamiento por lotes.\n"
            "Las facturas inválidas se rechazan automáticamente según las reglas."
        )
        info_footer.setFont(QFont("Arial", 8))
        info_footer.setAlignment(Qt.AlignmentFlag.AlignCenter)
        info_footer.setStyleSheet("color: #7f8c8d; padding: 10px;")
        info_footer.setWordWrap(True)
        main_layout.addWidget(info_footer)

        content_widget.setLayout(main_layout)
        scroll_area.setWidget(content_widget)
        root_layout.addWidget(scroll_area)

        self.setLayout(root_layout)

    def _inicializar_db_si_necesario(self) -> bool:
        """
        Inicializa la base de datos de forma lazy (solo cuando se necesita)

        Returns:
            bool: True si la BD está disponible, False si hubo error
        """
        # Si ya se intentó inicializar, retornar estado
        if self.db_inicializada:
            return self.db is not None

        # Marcar que ya se intentó
        self.db_inicializada = True

        try:
            logger.info("Inicializando base de datos de forma lazy...")
            self.db = LactalisDatabase()
            logger.info("Base de datos inicializada correctamente")

            # Actualizar label de info
            try:
                db_text = (
                    f"Base de datos: {self.db.db_path}\n"
                    f"Materiales: {self.db.contar_materiales()} | Clientes: {self.db.contar_clientes()}"
                )
                self.db_info_label.setText(db_text)
            except:
                pass

            return True

        except Exception as e:
            logger.error(f"Error al inicializar base de datos: {str(e)}", exc_info=True)
            self.db_error = str(e)
            self.db = None

            # Actualizar label de info
            self.db_info_label.setText(
                f"Base de datos: Error de inicialización\n"
                f"Detalles: {str(e)}"
            )

            return False

    def agregar_log(self, mensaje: str):
        """Agrega un mensaje a la consola de log"""
        self.log_console.append(mensaje)
        # Auto-scroll al final
        scrollbar = self.log_console.verticalScrollBar()
        scrollbar.setValue(scrollbar.maximum())

    def seleccionar_carpeta(self):
        """Permite seleccionar una carpeta con archivos ZIP/XML"""
        carpeta = QFileDialog.getExistingDirectory(
            self,
            "Seleccione la carpeta con archivos ZIP y/o XML de Lactalis Ventas",
            "",
            QFileDialog.Option.ShowDirsOnly
        )

        if not carpeta:
            return

        self.carpeta_entrada = Path(carpeta)

        # Buscar archivos ZIP y XML
        zip_files = list(self.carpeta_entrada.glob("*.zip"))
        xml_files = list(self.carpeta_entrada.glob("*.xml"))
        total_archivos = len(zip_files) + len(xml_files)

        if total_archivos == 0:
            QMessageBox.critical(
                self,
                "Sin archivos",
                "No se encontraron archivos ZIP ni XML en la carpeta seleccionada"
            )
            return

        # Confirmar procesamiento
        respuesta = QMessageBox.question(
            self,
            "Confirmar procesamiento",
            f"Se encontraron:\n"
            f"  • {len(zip_files)} archivo(s) ZIP\n"
            f"  • {len(xml_files)} archivo(s) XML\n"
            f"  • {total_archivos} archivos TOTALES\n\n"
            f"Carpeta: {self.carpeta_entrada.name}\n\n"
            f"REGLAS DE VALIDACIÓN:\n"
            f"  ✓ Solo FACTURAS (Invoice)\n"
            f"  ✓ Cantidad > 0\n"
            f"  ✓ Precio unitario > 0\n"
            f"  ✓ Total > 0\n\n"
            f"¿Procesar ahora?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )

        if respuesta == QMessageBox.StandardButton.Yes:
            self.iniciar_procesamiento()

    def buscar_o_crear_plantilla(self) -> Path:
        """Busca o crea la plantilla REGGIS Excel"""
        script_dir = Path(__file__).parent.parent.parent.parent
        plantilla = script_dir / "Plantilla_REGGIS.xlsx"

        if not plantilla.exists():
            logger.info(f"Plantilla no encontrada. Creando en: {plantilla}")
            self.crear_plantilla_base(plantilla)

        return plantilla

    def crear_plantilla_base(self, ruta: Path):
        """Crea una plantilla base de Excel con formato REGGIS"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Datos"

        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col, header in enumerate(REGGIS_HEADERS, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        wb.save(ruta)
        logger.info(f"Plantilla creada exitosamente: {ruta}")

    def actualizar_progreso(self, processed: int, total: int, message: str):
        """Callback para actualizar el progreso"""
        if total > 0:
            porcentaje = int((processed / total) * 100)
            self.progress.setValue(porcentaje)
            
        self.agregar_log(f"[{processed}/{total}] {message}")

    def iniciar_procesamiento(self):
        """Inicia el procesamiento en segundo plano"""
        self.progress.setVisible(True)
        self.progress.setValue(0)
        self.log_console.clear()

        self.estado_label.setText("⏳ Procesando archivos de LACTALIS VENTAS...")
        self.estado_label.setStyleSheet("color: #f39c12; padding: 10px; font-weight: bold;")

        self.agregar_log("=" * 60)
        self.agregar_log("Iniciando procesamiento de LACTALIS VENTAS...")
        self.agregar_log("=" * 60)

        # Obtener plantilla
        plantilla = self.buscar_o_crear_plantilla()

        # Obtener estado de checkboxes
        validar_materiales = self.checkbox_validar_materiales.isChecked()
        validar_clientes = self.checkbox_validar_clientes.isChecked()

        logger.info(
            f"Validaciones Lactalis Ventas -> materiales={validar_materiales}, clientes={validar_clientes}"
        )
        self.agregar_log(
            f"Validaciones activas -> Materiales: {'Sí' if validar_materiales else 'No'} | Clientes: {'Sí' if validar_clientes else 'No'}"
        )

        # Si se requieren validaciones, inicializar BD
        if validar_materiales or validar_clientes:
            self.agregar_log("Inicializando base de datos para validaciones...")
            if not self._inicializar_db_si_necesario():
                QMessageBox.critical(
                    self,
                    "Error de base de datos",
                    f"No se pudo inicializar la base de datos:\n\n{self.db_error}\n\n"
                    f"Las validaciones se deshabilitarán."
                )
                validar_materiales = False
                validar_clientes = False
                self.checkbox_validar_materiales.setChecked(False)
                self.checkbox_validar_clientes.setChecked(False)
            else:
                self.agregar_log("✅ Base de datos lista")

                # CRÍTICO: Verificar que la BD tenga datos
                num_materiales = self.db.contar_materiales()
                num_clientes = self.db.contar_clientes()

                warning_msgs = []
                if validar_materiales and num_materiales == 0:
                    warning_msgs.append(
                        "⚠️ VALIDACIÓN DE MATERIALES ACTIVADA pero la base de datos está VACÍA.\n"
                        f"   NO hay materiales registrados ({num_materiales} materiales).\n"
                        "   TODAS las líneas serán RECHAZADAS.\n"
                        "   Debes importar materiales desde Excel primero."
                    )

                if validar_clientes and num_clientes == 0:
                    warning_msgs.append(
                        "⚠️ VALIDACIÓN DE CLIENTES ACTIVADA pero la base de datos está VACÍA.\n"
                        f"   NO hay clientes registrados ({num_clientes} clientes).\n"
                        "   TODAS las líneas serán RECHAZADAS.\n"
                        "   Debes importar clientes desde Excel primero."
                    )

                if warning_msgs:
                    warning_text = "\n\n".join(warning_msgs)
                    respuesta = QMessageBox.warning(
                        self,
                        "⚠️ Base de datos vacía",
                        f"{warning_text}\n\n"
                        f"📊 Estado actual de la base de datos:\n"
                        f"  • Materiales: {num_materiales}\n"
                        f"  • Clientes: {num_clientes}\n\n"
                        f"¿Deseas continuar de todas formas?\n"
                        f"(Se procesarán los archivos pero probablemente todas las líneas serán rechazadas)",
                        QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
                    )

                    if respuesta == QMessageBox.StandardButton.No:
                        self.agregar_log("❌ Procesamiento cancelado por el usuario")
                        self.progress.setVisible(False)
                        self.estado_label.setText("")
                        return
                else:
                    self.agregar_log(f"✅ BD tiene {num_materiales} materiales y {num_clientes} clientes")
        else:
            logger.info("Validaciones desactivadas para Lactalis Ventas (no se consultará BD)")
            self.agregar_log("⚠️ Validaciones desactivadas: no se consultará la base de datos")

        # Crear procesador con callback de progreso y validaciones
        procesador = ProcesadorLactalisVentas(
            self.carpeta_entrada,
            plantilla,
            progress_callback=None,
            database=self.db if (validar_materiales or validar_clientes) else None,
            validar_materiales=validar_materiales,
            validar_clientes=validar_clientes
        )

        # Iniciar thread de procesamiento
        self.procesamiento_thread = ProcesamientoThread(procesador.procesar)
        self.procesamiento_thread.progress.connect(self.actualizar_progreso)
        procesador.progress_callback = self.procesamiento_thread.progress.emit
        self.procesamiento_thread.finished.connect(self.procesamiento_finalizado)
        self.procesamiento_thread.start()

    def procesamiento_finalizado(self, success: bool, message: str, result):
        """
        Callback cuando el procesamiento termina
        
        Args:
            success: True si fue exitoso
            message: Mensaje de resultado
            result: Path de carpeta de resultados o None
        """
        self.progress.setVisible(False)

        if success:
            self.estado_label.setText("✅ Proceso completado exitosamente")
            self.estado_label.setStyleSheet("color: #27ae60; padding: 10px; font-weight: bold;")
            
            self.agregar_log("=" * 60)
            self.agregar_log("✅ PROCESO COMPLETADO EXITOSAMENTE")
            self.agregar_log("=" * 60)

            respuesta = QMessageBox.question(
                self,
                "Éxito",
                f"{message}\n\n"
                f"📂 Archivos guardados en:\n{result.name}\n\n"
                f"¿Desea abrir la carpeta de resultados?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )

            if respuesta == QMessageBox.StandardButton.Yes:
                self.abrir_carpeta(result)
        else:
            self.estado_label.setText("❌ Error en el procesamiento")
            self.estado_label.setStyleSheet("color: #e74c3c; padding: 10px; font-weight: bold;")
            
            self.agregar_log("=" * 60)
            self.agregar_log("❌ ERROR EN EL PROCESAMIENTO")
            self.agregar_log(f"Error: {message}")
            self.agregar_log("=" * 60)
            
            QMessageBox.critical(
                self,
                "Error de procesamiento",
                f"Ocurrió un error durante el procesamiento:\n\n{message}"
            )

    def importar_materiales(self):
        """Importa materiales desde un archivo Excel"""
        # Inicializar BD si es necesario
        if not self._inicializar_db_si_necesario():
            QMessageBox.critical(
                self,
                "Base de datos no disponible",
                f"No se pudo inicializar la base de datos:\n\n{self.db_error}\n\n"
                f"Por favor revise los logs para más detalles."
            )
            return

        archivo, _ = QFileDialog.getOpenFileName(
            self,
            "Seleccionar archivo Excel de Materiales",
            "",
            "Excel Files (*.xlsx *.xls)"
        )

        if not archivo:
            return

        try:
            logger.info(f"Iniciando importación de materiales desde: {archivo}")
            # Validar formato del archivo
            es_valido, mensaje = ExcelImporter.validar_archivo_materiales(archivo)
            if not es_valido:
                QMessageBox.critical(
                    self,
                    "Formato inválido",
                    f"El archivo no tiene el formato correcto:\n\n{mensaje}\n\n"
                    f"Se esperan los encabezados: CODIGO, DESCRIPCION, SOCIEDAD"
                )
                return

            # Importar materiales
            materiales = ExcelImporter.importar_materiales_desde_excel(archivo)

            if not materiales:
                QMessageBox.warning(
                    self,
                    "Sin datos",
                    "No se encontraron materiales para importar"
                )
                return

            # Guardar en base de datos
            nuevos, existentes, errores = self.db.importar_materiales(materiales)

            # Actualizar label de info
            self.db_info_label.setText(
                f"Base de datos: {self.db.db_path}\n"
                f"Materiales: {self.db.contar_materiales()} | Clientes: {self.db.contar_clientes()}"
            )

            # Mostrar resultado
            QMessageBox.information(
                self,
                "Importación completada",
                f"Materiales importados:\n\n"
                f"✓ Nuevos: {nuevos}\n"
                f"⊙ Ya existentes: {existentes}\n"
                f"✗ Errores: {errores}\n\n"
                f"Total en BD: {self.db.contar_materiales()}"
            )

        except ExcelImporterError as e:
            QMessageBox.critical(
                self,
                "Error de importación",
                f"Error importando materiales:\n\n{str(e)}"
            )
        except Exception as e:
            logger.error(f"Error importando materiales: {str(e)}", exc_info=True)
            QMessageBox.critical(
                self,
                "Error",
                f"Error inesperado:\n\n{str(e)}"
            )

    def importar_clientes(self):
        """Importa clientes desde un archivo Excel"""
        # Inicializar BD si es necesario
        if not self._inicializar_db_si_necesario():
            QMessageBox.critical(
                self,
                "Base de datos no disponible",
                f"No se pudo inicializar la base de datos:\n\n{self.db_error}\n\n"
                f"Por favor revise los logs para más detalles."
            )
            return

        archivo, _ = QFileDialog.getOpenFileName(
            self,
            "Seleccionar archivo Excel de Clientes",
            "",
            "Excel Files (*.xlsx *.xls)"
        )

        if not archivo:
            return

        try:
            logger.info(f"Iniciando importación de clientes desde: {archivo}")
            # Validar formato del archivo
            es_valido, mensaje = ExcelImporter.validar_archivo_clientes(archivo)
            if not es_valido:
                QMessageBox.critical(
                    self,
                    "Formato inválido",
                    f"El archivo no tiene el formato correcto:\n\n{mensaje}\n\n"
                    f"Se esperan los encabezados: Cód.Padre, Nombre Código Padre, NIT, Se Registra"
                )
                return

            # Importar clientes
            clientes = ExcelImporter.importar_clientes_desde_excel(archivo)

            if not clientes:
                QMessageBox.warning(
                    self,
                    "Sin datos",
                    "No se encontraron clientes para importar"
                )
                return

            # Guardar en base de datos
            nuevos, existentes, errores = self.db.importar_clientes(clientes)

            # Actualizar label de info
            self.db_info_label.setText(
                f"Base de datos: {self.db.db_path}\n"
                f"Materiales: {self.db.contar_materiales()} | Clientes: {self.db.contar_clientes()}"
            )

            # Mostrar resultado
            QMessageBox.information(
                self,
                "Importación completada",
                f"Clientes importados:\n\n"
                f"✓ Nuevos: {nuevos}\n"
                f"⊙ Ya existentes: {existentes}\n"
                f"✗ Errores/rechazados: {errores}\n\n"
                f"Total en BD: {self.db.contar_clientes()}\n\n"
                f"Nota: Los clientes con 'Se Registra' = NO NIT no se validan"
            )

        except ExcelImporterError as e:
            QMessageBox.critical(
                self,
                "Error de importación",
                f"Error importando clientes:\n\n{str(e)}"
            )
        except Exception as e:
            logger.error(f"Error importando clientes: {str(e)}", exc_info=True)
            QMessageBox.critical(
                self,
                "Error",
                f"Error inesperado:\n\n{str(e)}"
            )

    def abrir_carpeta(self, carpeta: Path):
        """
        Abre la carpeta de resultados en el explorador del sistema

        Args:
            carpeta: Path a la carpeta a abrir
        """
        try:
            if platform.system() == 'Windows':
                os.startfile(carpeta)
            elif platform.system() == 'Darwin':  # macOS
                os.system(f'open "{carpeta}"')
            else:  # Linux
                os.system(f'xdg-open "{carpeta}"')
        except Exception as e:
            logger.error(f"Error abriendo carpeta: {str(e)}")
            QMessageBox.warning(
                self,
                "Error",
                f"No se pudo abrir la carpeta automáticamente:\n{str(e)}\n\n"
                f"Ubicación: {carpeta}"
            )
