"""
Tab para procesamiento de facturas SEABOARD
Reutiliza la lógica existente de procesadores
"""

import os
import platform
import logging
import openpyxl
from pathlib import Path
from openpyxl.styles import Font, PatternFill, Alignment

from PyQt6.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QLabel,
                              QPushButton, QProgressBar, QFileDialog,
                              QMessageBox, QListWidget, QDialog, QGroupBox,
                              QFormLayout)
from PyQt6.QtCore import Qt, QThread, pyqtSignal
from PyQt6.QtGui import QFont

from config.constants import REGGIS_HEADERS
from utils.sharepoint_detector import DetectorSharePoint
from processors.seaboard_processor import ProcesadorSeaboard

logger = logging.getLogger(__name__)


class ProcesamientoThread(QThread):
    """Thread para ejecutar procesamiento en segundo plano"""
    finished = pyqtSignal(bool, str, object)  # success, message, result

    def __init__(self, procesador_func, *args):
        super().__init__()
        self.procesador_func = procesador_func
        self.args = args

    def run(self):
        try:
            result = self.procesador_func(*self.args)
            self.finished.emit(True, "Proceso completado exitosamente", result)
        except Exception as e:
            logger.error(f"Error en procesamiento: {str(e)}", exc_info=True)
            self.finished.emit(False, f"Error: {str(e)}", None)


class TabSeaboard(QWidget):
    """
    Tab para procesamiento de facturas SEABOARD

    Funcionalidades:
    - Detección automática de carpetas SharePoint
    - Selección de carpeta local con archivos XML
    - Procesamiento en segundo plano con barra de progreso
    - Apertura automática de carpeta de resultados
    """

    def __init__(self):
        super().__init__()
        self.carpeta_entrada = None
        self.carpetas_sharepoint = []
        self.procesamiento_thread = None

        self.detectar_sharepoint()
        self.setup_ui()

    def detectar_sharepoint(self):
        """Detecta carpetas de SharePoint sincronizadas localmente"""
        try:
            self.carpetas_sharepoint = DetectorSharePoint.encontrar_carpetas_sharepoint()
            if self.carpetas_sharepoint:
                logger.info(f"Se encontraron {len(self.carpetas_sharepoint)} carpetas de SharePoint")
        except Exception as e:
            logger.error(f"Error detectando SharePoint: {str(e)}")
            self.carpetas_sharepoint = []

    def setup_ui(self):
        """Configura la interfaz del tab"""
        # Layout principal
        main_layout = QVBoxLayout()
        main_layout.setContentsMargins(20, 20, 20, 20)
        main_layout.setSpacing(15)

        # --- Título ---
        titulo = QLabel("🌐 PROCESADOR SEABOARD")
        titulo.setFont(QFont("Arial", 16, QFont.Weight.Bold))
        titulo.setAlignment(Qt.AlignmentFlag.AlignCenter)
        titulo.setStyleSheet("color: #0078D4; padding: 10px;")
        main_layout.addWidget(titulo)

        # --- Descripción ---
        descripcion = QLabel(
            "Procesa archivos XML de facturas SEABOARD desde SharePoint o carpetas locales.\n"
            "Soporta conversión de moneda (USD/COP) y unidades (TNE → Kg)."
        )
        descripcion.setFont(QFont("Arial", 9))
        descripcion.setAlignment(Qt.AlignmentFlag.AlignCenter)
        descripcion.setStyleSheet("color: #555; padding: 5px;")
        descripcion.setWordWrap(True)
        main_layout.addWidget(descripcion)

        # --- Group Box: Selección de Origen ---
        group_origen = QGroupBox("Seleccionar Origen de Datos")
        group_origen.setFont(QFont("Arial", 10, QFont.Weight.Bold))
        group_layout = QVBoxLayout()
        group_layout.setSpacing(10)

        # Información de SharePoint
        if self.carpetas_sharepoint:
            info_sp = QLabel(f"✅ Se detectaron {len(self.carpetas_sharepoint)} carpeta(s) de SharePoint sincronizada(s)")
            info_sp.setStyleSheet("color: #27ae60; padding: 5px;")
            info_sp.setFont(QFont("Arial", 9))
            group_layout.addWidget(info_sp)

            # Botón SharePoint
            btn_sharepoint = QPushButton("📁 BUSCAR EN SHAREPOINT SINCRONIZADO")
            btn_sharepoint.setMinimumHeight(50)
            btn_sharepoint.setFont(QFont("Arial", 11, QFont.Weight.Bold))
            btn_sharepoint.setStyleSheet("""
                QPushButton {
                    background-color: #0078D4;
                    color: white;
                    border: none;
                    border-radius: 8px;
                    padding: 10px;
                }
                QPushButton:hover {
                    background-color: #005a9e;
                }
                QPushButton:pressed {
                    background-color: #004578;
                }
            """)
            btn_sharepoint.setCursor(Qt.CursorShape.PointingHandCursor)
            btn_sharepoint.clicked.connect(self.seleccionar_desde_sharepoint)
            group_layout.addWidget(btn_sharepoint)
        else:
            info_no_sp = QLabel("ℹ️ No se detectaron carpetas de SharePoint sincronizadas")
            info_no_sp.setStyleSheet("color: #95a5a6; padding: 5px;")
            info_no_sp.setFont(QFont("Arial", 9))
            group_layout.addWidget(info_no_sp)

        # Botón Local
        btn_local = QPushButton("💻 BUSCAR EN CARPETA LOCAL")
        btn_local.setMinimumHeight(50)
        btn_local.setFont(QFont("Arial", 11, QFont.Weight.Bold))
        btn_local.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border: none;
                border-radius: 8px;
                padding: 10px;
            }
            QPushButton:hover {
                background-color: #388E3C;
            }
            QPushButton:pressed {
                background-color: #2e7d32;
            }
        """)
        btn_local.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_local.clicked.connect(self.seleccionar_carpeta_local)
        group_layout.addWidget(btn_local)

        group_origen.setLayout(group_layout)
        main_layout.addWidget(group_origen)

        # --- Barra de Progreso ---
        self.progress = QProgressBar()
        self.progress.setMinimum(0)
        self.progress.setMaximum(0)  # Modo indeterminado
        self.progress.setVisible(False)
        self.progress.setStyleSheet("""
            QProgressBar {
                border: 2px solid #bdc3c7;
                border-radius: 5px;
                text-align: center;
                height: 25px;
            }
            QProgressBar::chunk {
                background-color: #0078D4;
            }
        """)
        main_layout.addWidget(self.progress)

        # --- Label de Estado ---
        self.estado_label = QLabel("")
        self.estado_label.setFont(QFont("Arial", 10))
        self.estado_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.estado_label.setStyleSheet("padding: 10px;")
        main_layout.addWidget(self.estado_label)

        # Espaciador
        main_layout.addStretch()

        # --- Información adicional ---
        info_footer = QLabel(
            "💡 Los archivos procesados se guardarán en formato Excel REGGIS en la carpeta de resultados."
        )
        info_footer.setFont(QFont("Arial", 8))
        info_footer.setAlignment(Qt.AlignmentFlag.AlignCenter)
        info_footer.setStyleSheet("color: #7f8c8d; padding: 10px;")
        info_footer.setWordWrap(True)
        main_layout.addWidget(info_footer)

        self.setLayout(main_layout)

    def seleccionar_desde_sharepoint(self):
        """Muestra diálogo para seleccionar una carpeta de SharePoint"""
        if not self.carpetas_sharepoint:
            QMessageBox.information(self, "Sin carpetas", "No se detectaron carpetas de SharePoint")
            return

        # Crear diálogo personalizado
        dialog = QDialog(self)
        dialog.setWindowTitle("Seleccionar Carpeta de SharePoint")
        dialog.setMinimumSize(700, 500)

        layout = QVBoxLayout()

        titulo = QLabel("📂 Carpetas de SharePoint Detectadas")
        titulo.setFont(QFont("Arial", 14, QFont.Weight.Bold))
        layout.addWidget(titulo)

        # Lista de carpetas
        listbox = QListWidget()
        for carpeta in self.carpetas_sharepoint:
            listbox.addItem(str(carpeta))
        layout.addWidget(listbox)

        # Botón seleccionar
        btn_seleccionar = QPushButton("✓ Procesar Carpeta Seleccionada")
        btn_seleccionar.setMinimumHeight(40)
        btn_seleccionar.setStyleSheet("""
            QPushButton {
                background-color: #0078D4;
                color: white;
                padding: 10px;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #005a9e;
            }
        """)

        def seleccionar():
            if listbox.currentRow() >= 0:
                carpeta_sel = self.carpetas_sharepoint[listbox.currentRow()]
                dialog.close()
                self.procesar_carpeta_xml(carpeta_sel)
            else:
                QMessageBox.warning(dialog, "Selección requerida", "Por favor seleccione una carpeta")

        btn_seleccionar.clicked.connect(seleccionar)
        layout.addWidget(btn_seleccionar)

        dialog.setLayout(layout)
        dialog.exec()

    def seleccionar_carpeta_local(self):
        """Permite seleccionar una carpeta local con archivos XML"""
        carpeta = QFileDialog.getExistingDirectory(
            self,
            "Seleccione la carpeta con archivos XML de SEABOARD",
            "",
            QFileDialog.Option.ShowDirsOnly
        )

        if carpeta:
            self.procesar_carpeta_xml(Path(carpeta))

    def procesar_carpeta_xml(self, carpeta: Path):
        """
        Valida y procesa la carpeta con archivos XML

        Args:
            carpeta: Path a la carpeta con archivos XML
        """
        # Buscar archivos XML
        archivos_xml = list(carpeta.glob("*.xml"))

        if not archivos_xml:
            QMessageBox.critical(
                self,
                "Sin archivos",
                "No se encontraron archivos XML en la carpeta seleccionada"
            )
            return

        # Confirmar procesamiento
        respuesta = QMessageBox.question(
            self,
            "Confirmar procesamiento",
            f"Se encontraron {len(archivos_xml)} archivo(s) XML.\n\n"
            f"Carpeta: {carpeta.name}\n\n"
            f"¿Procesar ahora?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )

        if respuesta == QMessageBox.StandardButton.Yes:
            self.carpeta_entrada = carpeta
            self.iniciar_procesamiento()

    def buscar_o_crear_plantilla(self) -> Path:
        """Busca o crea la plantilla REGGIS Excel"""
        # Buscar plantilla en el directorio raíz del proyecto
        script_dir = Path(__file__).parent.parent.parent.parent
        plantilla = script_dir / "Plantilla_REGGIS.xlsx"

        if not plantilla.exists():
            logger.info(f"Plantilla no encontrada. Creando en: {plantilla}")
            self.crear_plantilla_base(plantilla)

        return plantilla

    def crear_plantilla_base(self, ruta: Path):
        """
        Crea una plantilla base de Excel con formato REGGIS

        Args:
            ruta: Path donde crear la plantilla
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Datos"

        # Estilos de encabezado
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        # Escribir encabezados
        for col, header in enumerate(REGGIS_HEADERS, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        wb.save(ruta)
        logger.info(f"Plantilla creada exitosamente: {ruta}")

    def iniciar_procesamiento(self):
        """Inicia el procesamiento en segundo plano"""
        self.progress.setVisible(True)
        self.estado_label.setText("⏳ Procesando archivos XML de SEABOARD...")
        self.estado_label.setStyleSheet("color: #f39c12; padding: 10px; font-weight: bold;")

        # Obtener plantilla
        plantilla = self.buscar_o_crear_plantilla()

        # Crear procesador
        procesador = ProcesadorSeaboard(self.carpeta_entrada, plantilla)

        # Iniciar thread de procesamiento
        self.procesamiento_thread = ProcesamientoThread(procesador.procesar)
        self.procesamiento_thread.finished.connect(self.procesamiento_finalizado)
        self.procesamiento_thread.start()

    def procesamiento_finalizado(self, success: bool, message: str, result):
        """
        Callback cuando el procesamiento termina

        Args:
            success: True si fue exitoso
            message: Mensaje de resultado
            result: Path de carpeta de resultados o None
        """
        self.progress.setVisible(False)

        if success:
            self.estado_label.setText("✅ Proceso completado exitosamente")
            self.estado_label.setStyleSheet("color: #27ae60; padding: 10px; font-weight: bold;")

            respuesta = QMessageBox.question(
                self,
                "Éxito",
                f"{message}\n\n"
                f"📂 Archivos guardados en:\n{result.name}\n\n"
                f"¿Desea abrir la carpeta de resultados?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )

            if respuesta == QMessageBox.StandardButton.Yes:
                self.abrir_carpeta(result)
        else:
            self.estado_label.setText("❌ Error en el procesamiento")
            self.estado_label.setStyleSheet("color: #e74c3c; padding: 10px; font-weight: bold;")
            QMessageBox.critical(
                self,
                "Error de procesamiento",
                f"Ocurrió un error durante el procesamiento:\n\n{message}"
            )

    def abrir_carpeta(self, carpeta: Path):
        """
        Abre la carpeta de resultados en el explorador del sistema

        Args:
            carpeta: Path a la carpeta a abrir
        """
        try:
            if platform.system() == 'Windows':
                os.startfile(carpeta)
            elif platform.system() == 'Darwin':  # macOS
                os.system(f'open "{carpeta}"')
            else:  # Linux
                os.system(f'xdg-open "{carpeta}"')
        except Exception as e:
            logger.error(f"Error abriendo carpeta: {str(e)}")
            QMessageBox.warning(
                self,
                "Error",
                f"No se pudo abrir la carpeta automáticamente:\n{str(e)}\n\n"
                f"Ubicación: {carpeta}"
            )
