"""
Procesador Unificado de Facturas Electrónicas XML a Excel
Sistema Multi-Cliente: SEABOARD y CASA DEL AGRICULTOR
Autor: Sistema REGGIS
"""

import xml.etree.ElementTree as ET
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path
from datetime import datetime
import logging
from tkinter import Tk, filedialog, messagebox, ttk
import tkinter as tk
from typing import Dict, List, Optional
import os
import platform
import zipfile
import re

# Configuración de logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(f'procesamiento_facturas_{datetime.now().strftime("%Y%m%d_%H%M%S")}.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# Namespaces UBL estándar para Colombia
NAMESPACES = {
    'cac': 'urn:oasis:names:specification:ubl:schema:xsd:CommonAggregateComponents-2',
    'cbc': 'urn:oasis:names:specification:ubl:schema:xsd:CommonBasicComponents-2',
    'ext': 'urn:oasis:names:specification:ubl:schema:xsd:CommonExtensionComponents-2',
    'sts': 'dian:gov:co:facturaelectronica:Structures-2-1',
}


class SelectorCliente:
    """Ventana inicial para seleccionar el cliente"""
    
    def __init__(self):
        self.cliente_seleccionado = None
        self.root = Tk()
        self.root.title("Procesador de Facturas - Selector de Cliente")
        self.root.resizable(True, True)
        self.setup_ui()
        self.centrar_ventana()
    
    def centrar_ventana(self):
        """Centra la ventana en la pantalla"""
        self.root.update_idletasks()
        req_w = self.root.winfo_reqwidth()
        req_h = self.root.winfo_reqheight()
        screen_w = self.root.winfo_screenwidth()
        screen_h = self.root.winfo_screenheight()
        width = min(req_w, screen_w - 120)
        height = min(req_h, screen_h - 120)
        self.root.minsize(width, height)
        x = (screen_w // 2) - (width // 2)
        y = (screen_h // 2) - (height // 2)
        self.root.geometry(f'{width}x{height}+{x}+{y}')
    
    def setup_ui(self):
        """Configura la interfaz de selección"""
        main_frame = ttk.Frame(self.root, padding=(20, 18))
        main_frame.grid(row=0, column=0, sticky=(tk.N, tk.S, tk.E, tk.W))
        main_frame.columnconfigure(0, weight=1)
        
        titulo = ttk.Label(
            main_frame,
            text="PROCESADOR DE FACTURAS ELECTRÓNICAS",
            font=("Arial", 16, "bold"),
            anchor="center",
            wraplength=520
        )
        titulo.grid(row=0, column=0, pady=(6, 8), sticky="ew")
        
        subtitulo = ttk.Label(
            main_frame,
            text="Seleccione el cliente que desea procesar:",
            font=("Arial", 11),
            anchor="center",
            wraplength=520
        )
        subtitulo.grid(row=1, column=0, pady=(0, 12), sticky="ew")
        
        ttk.Separator(main_frame, orient='horizontal').grid(
            row=2, column=0, sticky='ew', pady=(0, 12)
        )
        
        botones_frame = ttk.Frame(main_frame)
        botones_frame.grid(row=3, column=0, pady=(0, 12), sticky="ew")
        botones_frame.columnconfigure(0, weight=1)
        botones_frame.columnconfigure(1, weight=1)
        
        btn_seaboard = tk.Button(
            botones_frame,
            text="🌐 SEABOARD\n(Procesamiento desde SharePoint/Local)",
            command=lambda: self.seleccionar_cliente("SEABOARD"),
            font=("Arial", 12, "bold"),
            bg="#0078D4",
            fg="white",
            padx=18,
            pady=14,
            cursor="hand2",
            relief=tk.RAISED,
            bd=3,
        )
        btn_seaboard.grid(row=0, column=0, padx=(0,10), pady=8, sticky="nsew")
        
        btn_casa = tk.Button(
            botones_frame,
            text="🌾 CASA DEL AGRICULTOR\n(Procesamiento desde archivos ZIP)",
            command=lambda: self.seleccionar_cliente("CASA_DEL_AGRICULTOR"),
            font=("Arial", 12, "bold"),
            bg="#27ae60",
            fg="white",
            padx=18,
            pady=14,
            cursor="hand2",
            relief=tk.RAISED,
            bd=3,
        )
        btn_casa.grid(row=0, column=1, padx=(10,0), pady=8, sticky="nsew")
        
        info_label = ttk.Label(
            main_frame,
            text="💡 Cada cliente tiene su propio flujo de procesamiento optimizado",
            justify=tk.CENTER,
            foreground="gray",
            font=("Arial", 9),
            wraplength=520
        )
        info_label.grid(row=4, column=0, pady=(6, 8), sticky="ew")
        
        btn_cancel = ttk.Button(main_frame, text="Cerrar", command=self.root.destroy)
        btn_cancel.grid(row=5, column=0, pady=(6, 2), sticky="e")
    
    def seleccionar_cliente(self, cliente: str):
        """Guarda la selección y cierra la ventana"""
        self.cliente_seleccionado = cliente
        self.root.destroy()
    
    def ejecutar(self) -> Optional[str]:
        """Muestra el selector y retorna el cliente seleccionado"""
        self.root.mainloop()
        return self.cliente_seleccionado


class DetectorSharePoint:
    """Detector de carpetas sincronizadas de SharePoint/OneDrive"""
    
    @staticmethod
    def encontrar_carpetas_sharepoint() -> List[Path]:
        """Encuentra carpetas de SharePoint/OneDrive sincronizadas en el sistema"""
        carpetas_encontradas = []
        
        if platform.system() == 'Windows':
            user_profile = Path(os.environ.get('USERPROFILE', ''))
            
            rutas_buscar = [
                user_profile / "OneDrive",
                user_profile / "OneDrive - SEABOARD",
                user_profile / "SharePoint",
                user_profile / "SEABOARD",
            ]
            
            drives = [f"{d}:\\" for d in "CDEFGHIJ" if os.path.exists(f"{d}:\\")]
            
            for drive in drives:
                drive_path = Path(drive)
                try:
                    for item in drive_path.iterdir():
                        if item.is_dir():
                            nombre_lower = item.name.lower()
                            if any(x in nombre_lower for x in ['sharepoint', 'onedrive', 'seaboard']):
                                rutas_buscar.append(item)
                except (PermissionError, OSError):
                    continue
            
            for ruta in rutas_buscar:
                if ruta.exists() and ruta.is_dir():
                    carpetas_encontradas.append(ruta)
                    try:
                        for subcarpeta in ruta.rglob("*"):
                            if subcarpeta.is_dir() and 'SEABOARD' in subcarpeta.name.upper():
                                carpetas_encontradas.append(subcarpeta)
                    except (PermissionError, OSError):
                        continue
        
        carpetas_unicas = []
        for carpeta in carpetas_encontradas:
            if carpeta not in carpetas_unicas:
                carpetas_unicas.append(carpeta)
        
        return carpetas_unicas


class FacturaExtractorSeaboard:
    """Extractor de datos de facturas para SEABOARD"""
    
    def __init__(self, xml_content: str):
        self.root = ET.fromstring(xml_content)
        self.ns = NAMESPACES
    
    def _get_text(self, xpath: str, default: str = "") -> str:
        element = self.root.find(xpath, self.ns)
        return element.text.strip() if element is not None and element.text else default
    
    def _get_decimal(self, xpath: str, default: float = 0.0) -> float:
        text = self._get_text(xpath)
        if not text:
            return default
        try:
            return float(text.replace(',', ''))
        except ValueError:
            return default
    
    def _formato_decimal(self, valor: float, decimales: int = 2) -> str:
        valor_str = f"{valor:.{decimales}f}"
        partes = valor_str.split('.')
        parte_entera = partes[0]
        parte_decimal = partes[1] if len(partes) > 1 else '00'
        
        parte_entera_formateada = ''
        for i, digito in enumerate(reversed(parte_entera)):
            if i > 0 and i % 3 == 0:
                parte_entera_formateada = '.' + parte_entera_formateada
            parte_entera_formateada = digito + parte_entera_formateada
        
        return f"{parte_entera_formateada},{parte_decimal}"
    
    def extraer_datos(self) -> List[Dict]:
        """Extrae datos en formato REGGIS para SEABOARD"""
        numero_factura = self._get_text('.//cbc:ID')
        fecha_emision = self._get_text('.//cbc:IssueDate')
        fecha_vencimiento = self._get_text('.//cbc:DueDate')
        
        nit_vendedor = self._get_text('.//cac:AccountingSupplierParty//cbc:CompanyID')
        nombre_vendedor = self._get_text('.//cac:AccountingSupplierParty//cbc:RegistrationName')
        ciudad_vendedor = self._get_text('.//cac:AccountingSupplierParty//cac:PhysicalLocation//cac:Address//cbc:CityName')
        
        nit_comprador = self._get_text('.//cac:AccountingCustomerParty//cbc:CompanyID')
        nombre_comprador = self._get_text('.//cac:AccountingCustomerParty//cbc:RegistrationName')
        
        moneda_documento = self._get_text('.//cbc:DocumentCurrencyCode', 'COP')
        trm = self._get_decimal('.//cac:PaymentExchangeRate//cbc:CalculationRate', 1.0)
        codigo_moneda = {'COP': '1', 'USD': '2', 'EUR': '3'}.get(moneda_documento, '1')
        
        porcentaje_iva = self._get_decimal('.//cac:TaxCategory//cbc:Percent', 5.0)
        
        lineas_procesadas = []
        items = self.root.findall('.//cac:InvoiceLine', self.ns)
        
        for idx, item in enumerate(items, 1):
            try:
                descripcion = item.find('.//cbc:Description', self.ns)
                nombre_producto = descripcion.text if descripcion is not None else ''
                
                codigo = item.find('.//cac:SellersItemIdentification//cbc:ID', self.ns)
                codigo_producto = codigo.text if codigo is not None else ''
                
                cantidad_elem = item.find('.//cbc:InvoicedQuantity', self.ns)
                cantidad_original = float(cantidad_elem.text) if cantidad_elem is not None else 0.0
                unidad_medida_original = cantidad_elem.get('unitCode', '') if cantidad_elem is not None else ''
                
                unidad_medida = unidad_medida_original
                if unidad_medida_original == 'TNE':
                    cantidad_en_kg = cantidad_original * 1000
                    unidad_medida = 'Kg'
                else:
                    cantidad_en_kg = cantidad_original
                
                precio_elem = item.find('.//cac:Price//cbc:PriceAmount', self.ns)
                precio_unitario_xml = float(precio_elem.text) if precio_elem is not None else 0.0
                
                if unidad_medida_original == 'TNE':
                    precio_por_kg = precio_unitario_xml / 1000
                else:
                    precio_por_kg = precio_unitario_xml
                
                if moneda_documento == 'USD':
                    precio_cop = precio_por_kg * trm
                else:
                    precio_cop = precio_por_kg
                
                total_sin_iva_linea = cantidad_en_kg * precio_cop
                iva_linea = total_sin_iva_linea * (porcentaje_iva / 100)
                total_con_iva_linea = total_sin_iva_linea + iva_linea
                
                linea = {
                    'numero_factura': numero_factura,
                    'nombre_producto': nombre_producto,
                    'codigo_subyacente': codigo_producto,
                    'unidad_medida': unidad_medida,
                    'cantidad': self._formato_decimal(cantidad_en_kg, decimales=5),
                    'precio_unitario': self._formato_decimal(precio_cop, decimales=6),
                    'fecha_factura': fecha_emision,
                    'fecha_pago': fecha_vencimiento,
                    'nit_comprador': nit_comprador,
                    'nombre_comprador': nombre_comprador,
                    'nit_vendedor': nit_vendedor,
                    'nombre_vendedor': nombre_vendedor,
                    'principal': 'V',
                    'municipio': ciudad_vendedor,
                    'iva': str(int(porcentaje_iva)),
                    'descripcion': nombre_producto,
                    'activa_factura': 'SI',
                    'activa_bodega': '',
                    'incentivo': '',
                    'cantidad_original': self._formato_decimal(cantidad_original, decimales=5),
                    'moneda': codigo_moneda,
                    'total_sin_iva': total_sin_iva_linea,
                    'total_iva': iva_linea,
                    'total_con_iva': total_con_iva_linea
                }
                
                lineas_procesadas.append(linea)
                
            except Exception as e:
                logger.error(f"Error procesando linea {idx}: {str(e)}")
                continue
        
        return lineas_procesadas


class ProcesadorSeaboard:
    """Procesador específico para SEABOARD"""
    
    def __init__(self, carpeta_xml: Path, plantilla_excel: Path):
        self.carpeta_xml = carpeta_xml
        self.plantilla_excel = plantilla_excel
        self.carpeta_salida = None
    
    def extraer_invoice_de_attached_document(self, xml_path: Path) -> Optional[str]:
        try:
            tree = ET.parse(xml_path)
            root = tree.getroot()
            description = root.find('.//cac:ExternalReference/cbc:Description', NAMESPACES)
            
            if description is not None and description.text:
                return description.text.strip()
            return None
        except Exception as e:
            logger.error(f"Error al extraer factura de {xml_path.name}: {str(e)}")
            return None
    
    def procesar_archivos_xml(self) -> List[Dict]:
        lineas_reggis = []
        archivos_xml = list(self.carpeta_xml.glob("*.xml"))
        
        logger.info(f"SEABOARD: Se encontraron {len(archivos_xml)} archivos XML")
        
        for xml_file in archivos_xml:
            try:
                invoice_xml = self.extraer_invoice_de_attached_document(xml_file)
                
                if invoice_xml:
                    extractor = FacturaExtractorSeaboard(invoice_xml)
                    lineas = extractor.extraer_datos()
                    lineas_reggis.extend(lineas)
                    
            except Exception as e:
                logger.error(f"Error procesando {xml_file.name}: {str(e)}")
        
        return lineas_reggis
    
    def crear_carpeta_salida(self) -> Path:
        carpeta_resultados_base = self.plantilla_excel.parent / "Resultados_SEABOARD"
        carpeta_resultados_base.mkdir(exist_ok=True)
        
        nombre_carpeta_xml = self.carpeta_xml.name
        carpeta_salida = carpeta_resultados_base / nombre_carpeta_xml
        carpeta_salida.mkdir(exist_ok=True)
        
        return carpeta_salida
    
    def escribir_reggis(self, lineas: List[Dict]) -> Path:
        wb = openpyxl.load_workbook(self.plantilla_excel)
        ws = wb.active
        
        fila_inicial = ws.max_row + 1 if ws.max_row > 1 else 2
        
        for idx, linea in enumerate(lineas, start=fila_inicial):
            ws.cell(row=idx, column=1, value=linea['numero_factura'])
            ws.cell(row=idx, column=2, value=linea['nombre_producto'])
            ws.cell(row=idx, column=3, value=linea['codigo_subyacente'])
            ws.cell(row=idx, column=4, value=linea['unidad_medida'])
            ws.cell(row=idx, column=5, value=linea['cantidad'])
            ws.cell(row=idx, column=6, value=linea['precio_unitario'])
            ws.cell(row=idx, column=7, value=linea['fecha_factura'])
            ws.cell(row=idx, column=8, value=linea['fecha_pago'])
            ws.cell(row=idx, column=9, value=linea['nit_comprador'])
            ws.cell(row=idx, column=10, value=linea['nombre_comprador'])
            ws.cell(row=idx, column=11, value=linea['nit_vendedor'])
            ws.cell(row=idx, column=12, value=linea['nombre_vendedor'])
            ws.cell(row=idx, column=13, value=linea['principal'])
            ws.cell(row=idx, column=14, value=linea['municipio'])
            ws.cell(row=idx, column=15, value=linea['iva'])
            ws.cell(row=idx, column=16, value=linea['descripcion'])
            ws.cell(row=idx, column=17, value=linea['activa_factura'])
            ws.cell(row=idx, column=18, value=linea['activa_bodega'])
            ws.cell(row=idx, column=19, value=linea['incentivo'])
            ws.cell(row=idx, column=20, value=linea['cantidad_original'])
            ws.cell(row=idx, column=21, value=linea['moneda'])
            ws.cell(row=idx, column=22, value=linea['total_sin_iva'])
            ws.cell(row=idx, column=23, value=linea['total_iva'])
            ws.cell(row=idx, column=24, value=linea['total_con_iva'])
        
        salida = self.carpeta_salida / "REGGIS_Procesado_SEABOARD.xlsx"
        wb.save(salida)
        return salida
    
    def procesar(self) -> Path:
        self.carpeta_salida = self.crear_carpeta_salida()
        lineas_reggis = self.procesar_archivos_xml()
        
        if not lineas_reggis:
            raise Exception("No se procesaron líneas. Verifique los archivos XML.")
        
        self.escribir_reggis(lineas_reggis)
        logger.info(f"SEABOARD: Proceso completado - {len(lineas_reggis)} líneas")
        
        return self.carpeta_salida


class ProcesadorCasaDelAgricultor:
    """Procesador específico para CASA DEL AGRICULTOR"""
    
    def __init__(self, carpeta_zip: Path, carpeta_salida: Path):
        self.carpeta_zip = carpeta_zip
        self.carpeta_salida = carpeta_salida
        self.processed_lines = []
    
    def extract_xml_from_zip(self, zip_path: Path) -> Optional[str]:
        try:
            with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                xml_files = [f for f in zip_ref.namelist() if f.endswith('.xml')]
                if xml_files:
                    xml_content = zip_ref.read(xml_files[0])
                    return xml_content.decode('utf-8')
        except Exception as e:
            logger.error(f"Error al extraer XML de ZIP: {str(e)}")
        return None
    
    def parse_invoice_xml(self, xml_content: str) -> List[Dict]:
        try:
            cdata_match = re.search(r'<!\[CDATA\[(.*?)\]\]>', xml_content, re.DOTALL)
            if cdata_match:
                invoice_xml = cdata_match.group(1)
            else:
                invoice_xml = xml_content
            
            invoice_xml = re.sub(r'\sxmlns[^=]*="[^"]*"', '', invoice_xml)
            invoice_xml = re.sub(r'<([a-zA-Z]+):([a-zA-Z]+)', r'<\2', invoice_xml)
            invoice_xml = re.sub(r'</([a-zA-Z]+):([a-zA-Z]+)', r'</\2', invoice_xml)
            
            root = ET.fromstring(invoice_xml)
            
            lines = []
            
            supplier_elem = root.find('.//AccountingSupplierParty')
            if supplier_elem:
                supplier_name = supplier_elem.findtext('.//RegistrationName') or ''
                supplier_nit = supplier_elem.findtext('.//CompanyID') or ''
                supplier_city = supplier_elem.findtext('.//CityName') or ''
            else:
                supplier_name = supplier_nit = supplier_city = ''
            
            customer_elem = root.find('.//AccountingCustomerParty')
            if customer_elem:
                customer_name = customer_elem.findtext('.//RegistrationName') or ''
                customer_nit = customer_elem.findtext('.//CompanyID') or ''
            else:
                customer_name = customer_nit = ''
            
            invoice_number = root.findtext('.//ID') or ''
            invoice_date = root.findtext('.//IssueDate') or ''
            payment_date = root.findtext('.//PaymentDueDate') or ''
            
            currency_elem = root.find('.//DocumentCurrencyCode')
            currency_code = currency_elem.text if currency_elem is not None else 'COP'
            currency_map = {'COP': '1', 'USD': '2', 'EUR': '3'}
            currency_id = currency_map.get(currency_code, '1')
            
            for line in root.findall('.//InvoiceLine'):
                line_id = line.findtext('.//ID') or ''
                
                qty_elem = line.find('.//InvoicedQuantity')
                if qty_elem is not None:
                    quantity = float(qty_elem.text or 0)
                    unit_code = qty_elem.get('unitCode') or ''
                else:
                    quantity = 0
                    unit_code = ''
                
                desc_elem = line.find('.//Description')
                description = desc_elem.text if desc_elem is not None else ''
                
                code_elem = line.find('.//StandardItemIdentification/ID')
                if code_elem is None:
                    code_elem = line.find('.//Item/ID')
                code = code_elem.text if code_elem is not None else ''
                
                price_elem = line.find('.//PriceAmount')
                price = float(price_elem.text or 0) if price_elem is not None else 0
                
                total_elem = line.find('.//LineExtensionAmount')
                line_total = float(total_elem.text or 0) if total_elem is not None else 0
                
                iva_percent = 0
                iva_amount = 0
                tax_total = line.find('.//TaxTotal')
                if tax_total is not None:
                    percent_elem = tax_total.find('.//Percent')
                    if percent_elem is not None:
                        iva_percent = float(percent_elem.text or 0)
                    
                    iva_elem = tax_total.find('.//TaxAmount')
                    if iva_elem is not None:
                        iva_amount = float(iva_elem.text or 0)
                
                total_with_iva = line_total + iva_amount
                
                lines.append({
                    'line_id': line_id,
                    'code': code,
                    'description': description,
                    'quantity': quantity,
                    'unit': unit_code,
                    'price': price,
                    'line_total': line_total,
                    'invoice_number': invoice_number,
                    'invoice_date': invoice_date,
                    'payment_date': payment_date,
                    'supplier_name': supplier_name,
                    'supplier_nit': supplier_nit,
                    'supplier_city': supplier_city,
                    'customer_name': customer_name,
                    'customer_nit': customer_nit,
                    'currency_id': currency_id,
                    'iva_percent': iva_percent,
                    'iva_amount': iva_amount,
                    'total_with_iva': total_with_iva
                })
                
            return lines
            
        except Exception as e:
            logger.error(f"Error al parsear XML: {str(e)}")
            return []
    
    def apply_conversion_rules(self, line: Dict) -> Dict:
        original_qty = line['quantity']
        original_unit = line['unit']
        converted_qty = original_qty
        converted_unit = original_unit
        conversion_note = ""
        
        # Conversión de libras a kilogramos
        if original_unit == 'LBR':
            converted_qty = original_qty / 2
            converted_unit = 'KG'
            conversion_note = f"Convertido de {original_qty} LBR a {converted_qty:.5f} KG"
        
        # Conversión de GRAMOS en descripción (busca "GRAMOS" o "GRAMO")
        gram_match = re.search(r'(\d+)\s*GRAMOS?', line['description'], re.IGNORECASE)
        if gram_match:
            grams = float(gram_match.group(1))
            converted_qty = (grams * original_qty) / 1000
            converted_unit = 'KG'
            conversion_note = f"Convertido: ({grams} gr × {original_qty}) ÷ 1000 = {converted_qty:.5f} KG"
        
        # NUEVA VALIDACIÓN: Conversión de GRS en descripción (busca "GRS" o "grs")
        grs_match = re.search(r'(\d+)\s*GRS', line['description'], re.IGNORECASE)
        if grs_match:
            grams = float(grs_match.group(1))
            converted_qty = (grams * original_qty) / 1000
            converted_unit = 'KG'
            conversion_note = f"Convertido: ({grams} grs × {original_qty}) ÷ 1000 = {converted_qty:.5f} KG"
        
        # Mapeo de unidades estándar
        unit_map = {
            'KG': 'Kg', 'KGM': 'Kg', 'LBR': 'Kg',
            'LTR': 'Lt', 'LT': 'Lt',
            'NIU': 'Un', 'EA': 'Un', 'EV': 'Un', 'JR': 'Un', 'UN': 'Un'
        }
        
        converted_unit = unit_map.get(converted_unit, converted_unit)
        
        return {
            **line,
            'converted_quantity': converted_qty,
            'converted_unit': converted_unit,
            'conversion_note': conversion_note
        }
        
    def create_reggis_excel(self, output_path: Path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Facturas Procesadas"
        
        headers = [
            'N° Factura', 'Nombre Producto', 'Codigo Subyacente',
            'Unidad Medida en Kg,Un,Lt', 'Cantidad', 'Precio Unitario',
            'Fecha Factura', 'Fecha Pago', 'Nit Comprador', 'Nombre Comprador',
            'Nit Vendedor', 'Nombre Vendedor', 'Principal V,C', 'Municipio',
            'Iva', 'Descripción', 'Activa Factura', 'Activa Bodega', 'Incentivo',
            'Cantidad Original', 'Moneda', 'Total Sin IVA', 'Total IVA', 'Total Con IVA'
        ]
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for idx, line in enumerate(self.processed_lines, start=2):
            cantidad_convertida = f"{line['converted_quantity']:.5f}".replace('.', ',')
            cantidad_original = f"{line['quantity']:.5f}".replace('.', ',')
            precio_unitario = f"{line['price']:.5f}".replace('.', ',')
            
            row_data = [
                line['invoice_number'], line['description'], line['code'],
                line['converted_unit'], cantidad_convertida, precio_unitario,
                line['invoice_date'], line['payment_date'], line['customer_nit'],
                line['customer_name'], line['supplier_nit'], line['supplier_name'],
                'V', line['supplier_city'], line['iva_percent'], line['conversion_note'],
                '1', '1', '', cantidad_original, line['currency_id'],
                round(line['line_total'], 2), round(line['iva_amount'], 2),
                round(line['total_with_iva'], 2)
            ]
            
            for col, value in enumerate(row_data, start=1):
                cell = ws.cell(row=idx, column=col, value=value)
                cell.alignment = Alignment(vertical='center')

        column_widths = [15, 40, 15, 20, 15, 15, 12, 12, 15, 30, 15, 30, 12, 20, 8, 50, 12, 12, 10, 18, 10, 15, 12, 15]
        for col, width in enumerate(column_widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
        
        wb.save(output_path)
    
    def procesar(self) -> Path:
        zip_files = list(self.carpeta_zip.glob("*.zip"))
        
        if not zip_files:
            raise Exception("No se encontraron archivos ZIP en la carpeta seleccionada")
        
        logger.info(f"CASA DEL AGRICULTOR: Encontrados {len(zip_files)} archivos ZIP")
        
        for zip_file in zip_files:
            logger.info(f"Procesando: {zip_file.name}")
            
            xml_content = self.extract_xml_from_zip(zip_file)
            if not xml_content:
                continue
            
            lines = self.parse_invoice_xml(xml_content)
            if not lines:
                logger.warning(f"No se extrajeron líneas de {zip_file.name}")
                continue
            
            for line in lines:
                processed_line = self.apply_conversion_rules(line)
                self.processed_lines.append(processed_line)
                
                if processed_line['conversion_note']:
                    logger.info(f"  {processed_line['conversion_note']}")
        
        if not self.processed_lines:
            raise Exception("No se procesaron líneas. Verifique los archivos ZIP.")
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = self.carpeta_salida / f"Facturas_REGGIS_CASA_DEL_AGRICULTOR_{timestamp}.xlsx"
        
        self.create_reggis_excel(output_file)
        logger.info(f"CASA DEL AGRICULTOR: Proceso completado - {len(self.processed_lines)} líneas")
        
        return self.carpeta_salida


class InterfazUnificada:
    """Interfaz que gestiona ambos clientes con botón Volver"""
    
    def __init__(self, cliente: str):
        self.cliente = cliente
        self.root = Tk()
        self.root.title(f"Procesador de Facturas - {cliente}")
        self.root.resizable(True, True)
        
        self.carpeta_entrada = None
        self.carpetas_sharepoint = []
        self.request_return = False
        
        if cliente == "SEABOARD":
            self.detectar_sharepoint()
        
        self.setup_ui()
        self.centrar_ventana_por_cliente()
    
    def centrar_ventana_por_cliente(self):
        """Centra la ventana usando el tamaño requerido por los widgets"""
        self.root.update_idletasks()
        req_w = self.root.winfo_reqwidth()
        req_h = self.root.winfo_reqheight()
        screen_w = self.root.winfo_screenwidth()
        screen_h = self.root.winfo_screenheight()
        
        cliente_min_map = {
            "SEABOARD": (820, 620),
            "CASA_DEL_AGRICULTOR": (760, 560)
        }
        default_min = (720, 520)
        min_w, min_h = cliente_min_map.get(self.cliente, default_min)
        
        width = max(req_w, min_w)
        height = max(req_h, min_h)
        
        width = min(width, screen_w - 120)
        height = min(height, screen_h - 120)
        
        self.root.minsize(width, height)
        x = (screen_w // 2) - (width // 2)
        y = (screen_h // 2) - (height // 2)
        self.root.geometry(f'{width}x{height}+{x}+{y}')
    
    def detectar_sharepoint(self):
        try:
            self.carpetas_sharepoint = DetectorSharePoint.encontrar_carpetas_sharepoint()
            if self.carpetas_sharepoint:
                logger.info(f"Se encontraron {len(self.carpetas_sharepoint)} carpetas de SharePoint")
        except Exception as e:
            logger.error(f"Error detectando SharePoint: {str(e)}")
            self.carpetas_sharepoint = []
    
    def setup_ui(self):
        main_frame = ttk.Frame(self.root, padding=(18, 14))
        main_frame.grid(row=0, column=0, sticky=(tk.N, tk.S, tk.E, tk.W))
        main_frame.columnconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        
        # Barra superior con botón Volver
        top_bar = ttk.Frame(main_frame)
        top_bar.grid(row=0, column=0, columnspan=2, sticky="ew", pady=(0, 8))
        top_bar.columnconfigure(0, weight=0)
        top_bar.columnconfigure(1, weight=1)
        
        btn_volver = ttk.Button(top_bar, text="← Volver", command=self.volver_al_selector)
        btn_volver.grid(row=0, column=0, sticky="w")
        
        titulo = ttk.Label(
            main_frame,
            text=f"PROCESADOR - {self.cliente}",
            font=("Arial", 18, "bold"),
            anchor="center",
            wraplength=680
        )
        titulo.grid(row=1, column=0, columnspan=2, pady=(4, 10), sticky="ew")
        
        ttk.Separator(main_frame, orient='horizontal').grid(
            row=2, column=0, columnspan=2, sticky='ew', pady=(0, 12)
        )
        
        botones_frame = ttk.Frame(main_frame)
        botones_frame.grid(row=3, column=0, columnspan=2, pady=(0, 12), sticky="ew")
        botones_frame.columnconfigure(0, weight=1)
        
        if self.cliente == "SEABOARD":
            self.setup_botones_seaboard(botones_frame)
        else:
            self.setup_botones_casa(botones_frame)
        
        # Barra de progreso
        self.progress = ttk.Progressbar(
            main_frame,
            orient='horizontal',
            mode='indeterminate'
        )
        self.progress.grid(row=4, column=0, columnspan=2, pady=(8, 8), sticky="ew")
        
        # Label de estado
        self.estado_label = ttk.Label(
            main_frame,
            text="",
            font=("Arial", 10),
            wraplength=680,
            anchor="w",
            justify="left"
        )
        self.estado_label.grid(row=5, column=0, columnspan=2, pady=(2, 6), sticky="ew")
        
        # Espacio inferior
        bottom_frame = ttk.Frame(main_frame)
        bottom_frame.grid(row=6, column=0, columnspan=2, sticky="ew", pady=(6, 4))
        bottom_frame.columnconfigure(0, weight=1)
        bottom_frame.columnconfigure(1, weight=0)
        
        self.lbl_resultados = ttk.Label(bottom_frame, text="", font=("Arial", 9), foreground="gray")
        self.lbl_resultados.grid(row=0, column=0, sticky="w")
        
        btn_cerrar = ttk.Button(bottom_frame, text="Cerrar", command=self.root.destroy)
        btn_cerrar.grid(row=0, column=1, sticky="e", padx=(6,0))
    
    def setup_botones_seaboard(self, parent):
        if self.carpetas_sharepoint:
            info_sp = ttk.Label(
                parent,
                text=f"Se detectaron {len(self.carpetas_sharepoint)} carpeta(s) de SharePoint sincronizada(s)",
                font=("Arial", 10),
                foreground="green",
                wraplength=640
            )
            info_sp.pack(fill="x", pady=(0,6))
            
            btn_sharepoint = tk.Button(
                parent,
                text="BUSCAR EN SHAREPOINT SINCRONIZADO",
                command=self.seleccionar_desde_sharepoint,
                font=("Arial", 12, "bold"),
                bg="#0078D4",
                fg="white",
                padx=12,
                pady=12,
                cursor="hand2",
                relief=tk.RAISED,
                bd=3
            )
            btn_sharepoint.pack(fill="x", pady=(0,8))
        
        btn_local = tk.Button(
            parent,
            text="BUSCAR EN CARPETA LOCAL",
            command=self.seleccionar_y_procesar_seaboard,
            font=("Arial", 12, "bold"),
            bg="#4CAF50",
            fg="white",
            padx=12,
            pady=12,
            cursor="hand2",
            relief=tk.RAISED,
            bd=3
        )
        btn_local.pack(fill="x", pady=(0,2))
    
    def setup_botones_casa(self, parent):
        info = ttk.Label(
            parent,
            text="Seleccione la carpeta que contiene los archivos ZIP de facturas",
            font=("Arial", 10),
            wraplength=640
        )
        info.pack(fill="x", pady=(0,6))
        
        btn_procesar = tk.Button(
            parent,
            text="SELECCIONAR CARPETA CON ARCHIVOS ZIP",
            command=self.seleccionar_y_procesar_casa,
            font=("Arial", 12, "bold"),
            bg="#27ae60",
            fg="white",
            padx=12,
            pady=12,
            cursor="hand2",
            relief=tk.RAISED,
            bd=3
        )
        btn_procesar.pack(fill="x", pady=(0,2))
    
    def volver_al_selector(self):
        """Marca la intención de volver y cierra la ventana actual"""
        self.request_return = True
        self.root.destroy()
    
    def seleccionar_desde_sharepoint(self):
        if not self.carpetas_sharepoint:
            messagebox.showinfo("No hay carpetas", "No se detectaron carpetas de SharePoint")
            return
        
        ventana = tk.Toplevel(self.root)
        ventana.title("Seleccionar Carpeta de SharePoint")
        ventana.geometry("700x500")
        ventana.transient(self.root)
        ventana.grab_set()
        
        frame = ttk.Frame(ventana, padding="20")
        frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(frame, text="Carpetas de SharePoint Detectadas", font=("Arial", 14, "bold")).pack(pady=10)
        
        list_frame = ttk.Frame(frame)
        list_frame.pack(fill=tk.BOTH, expand=True, pady=10)
        
        scrollbar = ttk.Scrollbar(list_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        listbox = tk.Listbox(list_frame, font=("Arial", 10), yscrollcommand=scrollbar.set, height=15)
        listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.config(command=listbox.yview)
        
        for carpeta in self.carpetas_sharepoint:
            listbox.insert(tk.END, str(carpeta))
        
        def seleccionar():
            sel = listbox.curselection()
            if not sel:
                messagebox.showwarning("Selección requerida", "Seleccione una carpeta")
                return
            
            carpeta_sel = self.carpetas_sharepoint[sel[0]]
            ventana.destroy()
            self.procesar_carpeta_xml(carpeta_sel)
        
        ttk.Button(frame, text="Procesar Carpeta Seleccionada", command=seleccionar).pack(pady=10)
    
    def seleccionar_y_procesar_seaboard(self):
        carpeta = filedialog.askdirectory(title="Seleccione la carpeta con archivos XML")
        
        if not carpeta:
            return
        
        self.procesar_carpeta_xml(Path(carpeta))
    
    def procesar_carpeta_xml(self, carpeta: Path):
        archivos_xml = list(carpeta.glob("*.xml"))
        
        if not archivos_xml:
            messagebox.showerror("Sin archivos", "No se encontraron archivos XML")
            return
        
        respuesta = messagebox.askyesno(
            "Confirmar",
            f"Se encontraron {len(archivos_xml)} archivo(s) XML.\n\n¿Procesar ahora?"
        )
        
        if not respuesta:
            return
        
        self.carpeta_entrada = carpeta
        self.progress.start()
        self.estado_label.config(text=f"Procesando {len(archivos_xml)} archivo(s)...", foreground="orange")
        self.root.update()
        self.root.after(100, self.ejecutar_procesamiento_seaboard)
    
    def seleccionar_y_procesar_casa(self):
        carpeta = filedialog.askdirectory(title="Seleccione la carpeta con archivos ZIP")
        
        if not carpeta:
            return
        
        self.carpeta_entrada = Path(carpeta)
        
        zip_files = list(self.carpeta_entrada.glob("*.zip"))
        
        if not zip_files:
            messagebox.showerror("Sin archivos", "No se encontraron archivos ZIP")
            return
        
        respuesta = messagebox.askyesno(
            "Confirmar",
            f"Se encontraron {len(zip_files)} archivo(s) ZIP.\n\n¿Procesar ahora?"
        )
        
        if not respuesta:
            return
        
        self.progress.start()
        self.estado_label.config(text=f"Procesando {len(zip_files)} archivo(s)...", foreground="orange")
        self.root.update()
        self.root.after(100, self.ejecutar_procesamiento_casa)
    
    def buscar_o_crear_plantilla(self) -> Path:
        script_dir = Path(__file__).parent if '__file__' in globals() else Path.cwd()
        
        plantilla = script_dir / "Plantilla_REGGIS.xlsx"
        
        if not plantilla.exists():
            self.crear_plantilla_base(plantilla)
        
        return plantilla
    
    def crear_plantilla_base(self, ruta: Path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Datos"
        
        headers = [
            'N° Factura', 'Nombre Producto', 'Codigo Subyacente',
            'Unidad Medida en Kg,Un,Lt', 'Cantidad (5 decimales - separdor coma)',
            'Precio Unitario (5 decimales - separdor coma)', 'Fecha Factura Año-Mes-Dia',
            'Fecha Pago Año-Mes-Dia', 'Nit Comprador (Existente)', 'Nombre Comprador',
            'Nit Vendedor (Existente)', 'Nombre Vendedor', 'Principal V,C',
            'Municipio (Nombre Exacto de la Ciudad)', 'Iva (N°%)', 'Descripción',
            'Activa Factura', 'Activa Bodega', 'Incentivo',
            'Cantidad Original (5 decimales - separdor coma)', 'Moneda (1,2,3)',
            'Total Sin IVA', 'Total IVA', 'Total Con IVA'
        ]
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        wb.save(ruta)
    
    def ejecutar_procesamiento_seaboard(self):
        try:
            plantilla = self.buscar_o_crear_plantilla()
            procesador = ProcesadorSeaboard(self.carpeta_entrada, plantilla)
            carpeta_salida = procesador.procesar()
            
            self.progress.stop()
            self.estado_label.config(text="Proceso completado exitosamente", foreground="green")
            
            respuesta = messagebox.askyesno(
                "Éxito",
                f"Proceso completado!\n\nArchivos guardados en:\n{carpeta_salida.name}\n\n¿Abrir carpeta?"
            )
            
            if respuesta:
                self.abrir_carpeta(carpeta_salida)
                
        except Exception as e:
            self.progress.stop()
            self.estado_label.config(text="Error en el procesamiento", foreground="red")
            messagebox.showerror("Error", f"Error: {str(e)}")
    
    def ejecutar_procesamiento_casa(self):
        try:
            carpeta_salida = self.carpeta_entrada.parent / "Resultados_CASA_DEL_AGRICULTOR"
            carpeta_salida.mkdir(exist_ok=True)
            
            procesador = ProcesadorCasaDelAgricultor(self.carpeta_entrada, carpeta_salida)
            procesador.procesar()
            
            self.progress.stop()
            self.estado_label.config(text="Proceso completado exitosamente", foreground="green")
            
            respuesta = messagebox.askyesno(
                "Éxito",
                f"Proceso completado!\n\nArchivos guardados en:\n{carpeta_salida.name}\n\n¿Abrir carpeta?"
            )
            
            if respuesta:
                self.abrir_carpeta(carpeta_salida)
                
        except Exception as e:
            self.progress.stop()
            self.estado_label.config(text="Error en el procesamiento", foreground="red")
            messagebox.showerror("Error", f"Error: {str(e)}")
    
    def abrir_carpeta(self, carpeta: Path):
        if platform.system() == 'Windows':
            os.startfile(carpeta)
        elif platform.system() == 'Darwin':
            os.system(f'open "{carpeta}"')
        else:
            os.system(f'xdg-open "{carpeta}"')
    
    def ejecutar(self):
        self.root.mainloop()


def main():
    """Función principal que inicia la aplicación y permite volver al selector"""
    try:
        while True:
            selector = SelectorCliente()
            cliente = selector.ejecutar()
            
            if not cliente:
                logger.info("No se seleccionó ningún cliente. Aplicación cerrada.")
                break
            
            logger.info(f"Cliente seleccionado: {cliente}")
            
            app = InterfazUnificada(cliente)
            app.ejecutar()
            
            if getattr(app, "request_return", False):
                continue
            else:
                break
        
    except Exception as e:
        try:
            messagebox.showerror("Error Fatal", f"Error al iniciar:\n\n{str(e)}")
        except:
            pass
        logger.error(f"Error fatal: {str(e)}", exc_info=True)


if __name__ == "__main__":
    main()