"""
Procesador Unificado de Facturas Electrónicas XML a Excel
Sistema Multi-Cliente: SEABOARD, CASA DEL AGRICULTOR y LACTALIS
Versión PyQt6 - Interfaz Empresarial Profesional
Autor: Sistema REGGIS
"""

import xml.etree.ElementTree as ET
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path
from datetime import datetime
import logging
import sys
import os
import platform
import zipfile
import re
from typing import Dict, List, Optional

from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QLabel, QFileDialog, QMessageBox, QProgressBar,
    QListWidget, QDialog, QFrame, QSizePolicy
)
from PyQt6.QtCore import Qt, QThread, pyqtSignal, QTimer
from PyQt6.QtGui import QFont, QIcon, QPalette, QColor

# Importar el sistema de actualización
try:
    from auto_updater import check_and_notify_update, perform_update
    AUTO_UPDATE_AVAILABLE = True
except ImportError:
    AUTO_UPDATE_AVAILABLE = False
    logging.warning("Sistema de actualización no disponible")

# Configuración de logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(f'procesamiento_facturas_{datetime.now().strftime("%Y%m%d_%H%M%S")}.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# Namespaces UBL estándar para Colombia
NAMESPACES = {
    'cac': 'urn:oasis:names:specification:ubl:schema:xsd:CommonAggregateComponents-2',
    'cbc': 'urn:oasis:names:specification:ubl:schema:xsd:CommonBasicComponents-2',
    'ext': 'urn:oasis:names:specification:ubl:schema:xsd:CommonExtensionComponents-2',
    'sts': 'dian:gov:co:facturaelectronica:Structures-2-1',
}


# =====================================================================
# IMPORTAR CLASES DEL SISTEMA ORIGINAL
# =====================================================================

class DetectorSharePoint:
    """Detector de carpetas sincronizadas de SharePoint/OneDrive"""

    @staticmethod
    def encontrar_carpetas_sharepoint() -> List[Path]:
        """Encuentra carpetas de SharePoint/OneDrive sincronizadas en el sistema"""
        carpetas_encontradas = []

        if platform.system() == 'Windows':
            user_profile = Path(os.environ.get('USERPROFILE', ''))

            rutas_buscar = [
                user_profile / "OneDrive",
                user_profile / "OneDrive - SEABOARD",
                user_profile / "SharePoint",
                user_profile / "SEABOARD",
            ]

            drives = [f"{d}:\\" for d in "CDEFGHIJ" if os.path.exists(f"{d}:\\")]

            for drive in drives:
                drive_path = Path(drive)
                try:
                    for item in drive_path.iterdir():
                        if item.is_dir():
                            nombre_lower = item.name.lower()
                            if any(x in nombre_lower for x in ['sharepoint', 'onedrive', 'seaboard']):
                                rutas_buscar.append(item)
                except (PermissionError, OSError):
                    continue

            for ruta in rutas_buscar:
                if ruta.exists() and ruta.is_dir():
                    carpetas_encontradas.append(ruta)
                    try:
                        for subcarpeta in ruta.rglob("*"):
                            if subcarpeta.is_dir() and 'SEABOARD' in subcarpeta.name.upper():
                                carpetas_encontradas.append(subcarpeta)
                    except (PermissionError, OSError):
                        continue

        carpetas_unicas = []
        for carpeta in carpetas_encontradas:
            if carpeta not in carpetas_unicas:
                carpetas_unicas.append(carpeta)

        return carpetas_unicas


class FacturaExtractorSeaboard:
    """Extractor de datos de facturas para SEABOARD"""

    def __init__(self, xml_content: str):
        self.root = ET.fromstring(xml_content)
        self.ns = NAMESPACES

    def _get_text(self, xpath: str, default: str = "") -> str:
        element = self.root.find(xpath, self.ns)
        return element.text.strip() if element is not None and element.text else default

    def _get_decimal(self, xpath: str, default: float = 0.0) -> float:
        text = self._get_text(xpath)
        if not text:
            return default
        try:
            return float(text.replace(',', ''))
        except ValueError:
            return default

    def _formato_decimal(self, valor: float, decimales: int = 2) -> str:
        valor_str = f"{valor:.{decimales}f}"
        partes = valor_str.split('.')
        parte_entera = partes[0]
        parte_decimal = partes[1] if len(partes) > 1 else '00'

        parte_entera_formateada = ''
        for i, digito in enumerate(reversed(parte_entera)):
            if i > 0 and i % 3 == 0:
                parte_entera_formateada = '.' + parte_entera_formateada
            parte_entera_formateada = digito + parte_entera_formateada

        return f"{parte_entera_formateada},{parte_decimal}"

    def extraer_datos(self) -> List[Dict]:
        """Extrae datos en formato REGGIS para SEABOARD"""
        numero_factura = self._get_text('.//cbc:ID')
        fecha_emision = self._get_text('.//cbc:IssueDate')
        fecha_vencimiento = self._get_text('.//cbc:DueDate')

        nit_vendedor = self._get_text('.//cac:AccountingSupplierParty//cbc:CompanyID')
        nombre_vendedor = self._get_text('.//cac:AccountingSupplierParty//cbc:RegistrationName')
        ciudad_vendedor = self._get_text('.//cac:AccountingSupplierParty//cac:PhysicalLocation//cac:Address//cbc:CityName')

        nit_comprador = self._get_text('.//cac:AccountingCustomerParty//cbc:CompanyID')
        nombre_comprador = self._get_text('.//cac:AccountingCustomerParty//cbc:RegistrationName')

        moneda_documento = self._get_text('.//cbc:DocumentCurrencyCode', 'COP')
        trm = self._get_decimal('.//cac:PaymentExchangeRate//cbc:CalculationRate', 1.0)
        codigo_moneda = {'COP': '1', 'USD': '2', 'EUR': '3'}.get(moneda_documento, '1')

        porcentaje_iva = self._get_decimal('.//cac:TaxCategory//cbc:Percent', 5.0)

        lineas_procesadas = []
        items = self.root.findall('.//cac:InvoiceLine', self.ns)

        for idx, item in enumerate(items, 1):
            try:
                descripcion = item.find('.//cbc:Description', self.ns)
                nombre_producto = descripcion.text if descripcion is not None else ''

                codigo = item.find('.//cac:SellersItemIdentification//cbc:ID', self.ns)
                codigo_producto = codigo.text if codigo is not None else ''

                cantidad_elem = item.find('.//cbc:InvoicedQuantity', self.ns)
                cantidad_original = float(cantidad_elem.text) if cantidad_elem is not None else 0.0
                unidad_medida_original = cantidad_elem.get('unitCode', '') if cantidad_elem is not None else ''

                unidad_medida = unidad_medida_original
                if unidad_medida_original == 'TNE':
                    cantidad_en_kg = cantidad_original * 1000
                    unidad_medida = 'Kg'
                else:
                    cantidad_en_kg = cantidad_original

                precio_elem = item.find('.//cac:Price//cbc:PriceAmount', self.ns)
                precio_unitario_xml = float(precio_elem.text) if precio_elem is not None else 0.0

                if unidad_medida_original == 'TNE':
                    precio_por_kg = precio_unitario_xml / 1000
                else:
                    precio_por_kg = precio_unitario_xml

                if moneda_documento == 'USD':
                    precio_cop = precio_por_kg * trm
                else:
                    precio_cop = precio_por_kg

                total_sin_iva_linea = cantidad_en_kg * precio_cop
                iva_linea = total_sin_iva_linea * (porcentaje_iva / 100)
                total_con_iva_linea = total_sin_iva_linea + iva_linea

                linea = {
                    'numero_factura': numero_factura,
                    'nombre_producto': nombre_producto,
                    'codigo_subyacente': codigo_producto,
                    'unidad_medida': unidad_medida,
                    'cantidad': self._formato_decimal(cantidad_en_kg, decimales=5),
                    'precio_unitario': self._formato_decimal(precio_cop, decimales=6),
                    'fecha_factura': fecha_emision,
                    'fecha_pago': fecha_vencimiento,
                    'nit_comprador': nit_comprador,
                    'nombre_comprador': nombre_comprador,
                    'nit_vendedor': nit_vendedor,
                    'nombre_vendedor': nombre_vendedor,
                    'principal': 'V',
                    'municipio': ciudad_vendedor,
                    'iva': str(int(porcentaje_iva)),
                    'descripcion': nombre_producto,
                    'activa_factura': 'SI',
                    'activa_bodega': '',
                    'incentivo': '',
                    'cantidad_original': self._formato_decimal(cantidad_original, decimales=5),
                    'moneda': codigo_moneda,
                    'total_sin_iva': total_sin_iva_linea,
                    'total_iva': iva_linea,
                    'total_con_iva': total_con_iva_linea
                }

                lineas_procesadas.append(linea)

            except Exception as e:
                logger.error(f"Error procesando linea {idx}: {str(e)}")
                continue

        return lineas_procesadas


class ProcesadorSeaboard:
    """Procesador específico para SEABOARD"""

    def __init__(self, carpeta_xml: Path, plantilla_excel: Path):
        self.carpeta_xml = carpeta_xml
        self.plantilla_excel = plantilla_excel
        self.carpeta_salida = None

    def extraer_invoice_de_attached_document(self, xml_path: Path) -> Optional[str]:
        try:
            tree = ET.parse(xml_path)
            root = tree.getroot()
            description = root.find('.//cac:ExternalReference/cbc:Description', NAMESPACES)

            if description is not None and description.text:
                return description.text.strip()
            return None
        except Exception as e:
            logger.error(f"Error al extraer factura de {xml_path.name}: {str(e)}")
            return None

    def procesar_archivos_xml(self) -> List[Dict]:
        lineas_reggis = []
        archivos_xml = list(self.carpeta_xml.glob("*.xml"))

        logger.info(f"SEABOARD: Se encontraron {len(archivos_xml)} archivos XML")

        for xml_file in archivos_xml:
            try:
                invoice_xml = self.extraer_invoice_de_attached_document(xml_file)

                if invoice_xml:
                    extractor = FacturaExtractorSeaboard(invoice_xml)
                    lineas = extractor.extraer_datos()
                    lineas_reggis.extend(lineas)

            except Exception as e:
                logger.error(f"Error procesando {xml_file.name}: {str(e)}")

        return lineas_reggis

    def crear_carpeta_salida(self) -> Path:
        carpeta_resultados_base = self.plantilla_excel.parent / "Resultados_SEABOARD"
        carpeta_resultados_base.mkdir(exist_ok=True)

        nombre_carpeta_xml = self.carpeta_xml.name
        carpeta_salida = carpeta_resultados_base / nombre_carpeta_xml
        carpeta_salida.mkdir(exist_ok=True)

        return carpeta_salida

    def escribir_reggis(self, lineas: List[Dict]) -> Path:
        wb = openpyxl.load_workbook(self.plantilla_excel)
        ws = wb.active

        fila_inicial = ws.max_row + 1 if ws.max_row > 1 else 2

        for idx, linea in enumerate(lineas, start=fila_inicial):
            ws.cell(row=idx, column=1, value=linea['numero_factura'])
            ws.cell(row=idx, column=2, value=linea['nombre_producto'])
            ws.cell(row=idx, column=3, value=linea['codigo_subyacente'])
            ws.cell(row=idx, column=4, value=linea['unidad_medida'])
            ws.cell(row=idx, column=5, value=linea['cantidad'])
            ws.cell(row=idx, column=6, value=linea['precio_unitario'])
            ws.cell(row=idx, column=7, value=linea['fecha_factura'])
            ws.cell(row=idx, column=8, value=linea['fecha_pago'])
            ws.cell(row=idx, column=9, value=linea['nit_comprador'])
            ws.cell(row=idx, column=10, value=linea['nombre_comprador'])
            ws.cell(row=idx, column=11, value=linea['nit_vendedor'])
            ws.cell(row=idx, column=12, value=linea['nombre_vendedor'])
            ws.cell(row=idx, column=13, value=linea['principal'])
            ws.cell(row=idx, column=14, value=linea['municipio'])
            ws.cell(row=idx, column=15, value=linea['iva'])
            ws.cell(row=idx, column=16, value=linea['descripcion'])
            ws.cell(row=idx, column=17, value=linea['activa_factura'])
            ws.cell(row=idx, column=18, value=linea['activa_bodega'])
            ws.cell(row=idx, column=19, value=linea['incentivo'])
            ws.cell(row=idx, column=20, value=linea['cantidad_original'])
            ws.cell(row=idx, column=21, value=linea['moneda'])
            ws.cell(row=idx, column=22, value=linea['total_sin_iva'])
            ws.cell(row=idx, column=23, value=linea['total_iva'])
            ws.cell(row=idx, column=24, value=linea['total_con_iva'])

        salida = self.carpeta_salida / "REGGIS_Procesado_SEABOARD.xlsx"
        wb.save(salida)
        return salida

    def procesar(self) -> Path:
        self.carpeta_salida = self.crear_carpeta_salida()
        lineas_reggis = self.procesar_archivos_xml()

        if not lineas_reggis:
            raise Exception("No se procesaron líneas. Verifique los archivos XML.")

        self.escribir_reggis(lineas_reggis)
        logger.info(f"SEABOARD: Proceso completado - {len(lineas_reggis)} líneas")

        return self.carpeta_salida


class ProcesadorCasaDelAgricultor:
    """Procesador específico para CASA DEL AGRICULTOR"""

    def __init__(self, carpeta_zip: Path, carpeta_salida: Path):
        self.carpeta_zip = carpeta_zip
        self.carpeta_salida = carpeta_salida
        self.processed_lines = []

    def extract_xml_from_zip(self, zip_path: Path) -> Optional[str]:
        try:
            with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                xml_files = [f for f in zip_ref.namelist() if f.endswith('.xml')]
                if xml_files:
                    xml_content = zip_ref.read(xml_files[0])
                    return xml_content.decode('utf-8')
        except Exception as e:
            logger.error(f"Error al extraer XML de ZIP: {str(e)}")
        return None

    def parse_invoice_xml(self, xml_content: str) -> List[Dict]:
        try:
            cdata_match = re.search(r'<!\[CDATA\[(.*?)\]\]>', xml_content, re.DOTALL)
            if cdata_match:
                invoice_xml = cdata_match.group(1)
            else:
                invoice_xml = xml_content

            invoice_xml = re.sub(r'\sxmlns[^=]*="[^"]*"', '', invoice_xml)
            invoice_xml = re.sub(r'<([a-zA-Z]+):([a-zA-Z]+)', r'<\2', invoice_xml)
            invoice_xml = re.sub(r'</([a-zA-Z]+):([a-zA-Z]+)', r'</\2', invoice_xml)

            root = ET.fromstring(invoice_xml)

            lines = []

            supplier_elem = root.find('.//AccountingSupplierParty')
            if supplier_elem:
                supplier_name = supplier_elem.findtext('.//RegistrationName') or ''
                supplier_nit = supplier_elem.findtext('.//CompanyID') or ''
                supplier_city = supplier_elem.findtext('.//CityName') or ''
            else:
                supplier_name = supplier_nit = supplier_city = ''

            customer_elem = root.find('.//AccountingCustomerParty')
            if customer_elem:
                customer_name = customer_elem.findtext('.//RegistrationName') or ''
                customer_nit = customer_elem.findtext('.//CompanyID') or ''
            else:
                customer_name = customer_nit = ''

            invoice_number = root.findtext('.//ID') or ''
            invoice_date = root.findtext('.//IssueDate') or ''
            payment_date = root.findtext('.//PaymentDueDate') or ''

            currency_elem = root.find('.//DocumentCurrencyCode')
            currency_code = currency_elem.text if currency_elem is not None else 'COP'
            currency_map = {'COP': '1', 'USD': '2', 'EUR': '3'}
            currency_id = currency_map.get(currency_code, '1')

            for line in root.findall('.//InvoiceLine'):
                line_id = line.findtext('.//ID') or ''

                qty_elem = line.find('.//InvoicedQuantity')
                if qty_elem is not None:
                    quantity = float(qty_elem.text or 0)
                    unit_code = qty_elem.get('unitCode') or ''
                else:
                    quantity = 0
                    unit_code = ''

                desc_elem = line.find('.//Description')
                description = desc_elem.text if desc_elem is not None else ''

                code_elem = line.find('.//StandardItemIdentification/ID')
                if code_elem is None:
                    code_elem = line.find('.//Item/ID')
                code = code_elem.text if code_elem is not None else ''

                price_elem = line.find('.//PriceAmount')
                price = float(price_elem.text or 0) if price_elem is not None else 0

                total_elem = line.find('.//LineExtensionAmount')
                line_total = float(total_elem.text or 0) if total_elem is not None else 0

                iva_percent = 0
                iva_amount = 0
                tax_total = line.find('.//TaxTotal')
                if tax_total is not None:
                    percent_elem = tax_total.find('.//Percent')
                    if percent_elem is not None:
                        iva_percent = float(percent_elem.text or 0)

                    iva_elem = tax_total.find('.//TaxAmount')
                    if iva_elem is not None:
                        iva_amount = float(iva_elem.text or 0)

                total_with_iva = line_total + iva_amount

                lines.append({
                    'line_id': line_id,
                    'code': code,
                    'description': description,
                    'quantity': quantity,
                    'unit': unit_code,
                    'price': price,
                    'line_total': line_total,
                    'invoice_number': invoice_number,
                    'invoice_date': invoice_date,
                    'payment_date': payment_date,
                    'supplier_name': supplier_name,
                    'supplier_nit': supplier_nit,
                    'supplier_city': supplier_city,
                    'customer_name': customer_name,
                    'customer_nit': customer_nit,
                    'currency_id': currency_id,
                    'iva_percent': iva_percent,
                    'iva_amount': iva_amount,
                    'total_with_iva': total_with_iva
                })

            return lines

        except Exception as e:
            logger.error(f"Error al parsear XML: {str(e)}")
            return []

    def apply_conversion_rules(self, line: Dict) -> Dict:
        original_qty = line['quantity']
        original_unit = line['unit']
        converted_qty = original_qty
        converted_unit = original_unit
        conversion_note = ""

        # Conversión de libras a kilogramos
        if original_unit == 'LBR':
            converted_qty = original_qty / 2
            converted_unit = 'KG'
            conversion_note = f"Convertido de {original_qty} LBR a {converted_qty:.5f} KG"

        # Conversión de GRAMOS en descripción
        gram_match = re.search(r'(\d+)\s*GRAMOS?', line['description'], re.IGNORECASE)
        if gram_match:
            grams = float(gram_match.group(1))
            converted_qty = (grams * original_qty) / 1000
            converted_unit = 'KG'
            conversion_note = f"Convertido: ({grams} gr × {original_qty}) ÷ 1000 = {converted_qty:.5f} KG"

        # Conversión de GRS en descripción
        grs_match = re.search(r'(\d+)\s*GRS', line['description'], re.IGNORECASE)
        if grs_match:
            grams = float(grs_match.group(1))
            converted_qty = (grams * original_qty) / 1000
            converted_unit = 'KG'
            conversion_note = f"Convertido: ({grams} grs × {original_qty}) ÷ 1000 = {converted_qty:.5f} KG"

        # Mapeo de unidades estándar
        unit_map = {
            'KG': 'Kg', 'KGM': 'Kg', 'LBR': 'Kg',
            'LTR': 'Lt', 'LT': 'Lt',
            'NIU': 'Un', 'EA': 'Un', 'EV': 'Un', 'JR': 'Un', 'UN': 'Un'
        }

        converted_unit = unit_map.get(converted_unit, converted_unit)

        return {
            **line,
            'converted_quantity': converted_qty,
            'converted_unit': converted_unit,
            'conversion_note': conversion_note
        }

    def create_reggis_excel(self, output_path: Path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Facturas Procesadas"

        headers = [
            'N° Factura', 'Nombre Producto', 'Codigo Subyacente',
            'Unidad Medida en Kg,Un,Lt', 'Cantidad', 'Precio Unitario',
            'Fecha Factura', 'Fecha Pago', 'Nit Comprador', 'Nombre Comprador',
            'Nit Vendedor', 'Nombre Vendedor', 'Principal V,C', 'Municipio',
            'Iva', 'Descripción', 'Activa Factura', 'Activa Bodega', 'Incentivo',
            'Cantidad Original', 'Moneda', 'Total Sin IVA', 'Total IVA', 'Total Con IVA'
        ]

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for idx, line in enumerate(self.processed_lines, start=2):
            cantidad_convertida = f"{line['converted_quantity']:.5f}".replace('.', ',')
            cantidad_original = f"{line['quantity']:.5f}".replace('.', ',')
            precio_unitario = f"{line['price']:.5f}".replace('.', ',')

            row_data = [
                line['invoice_number'], line['description'], line['code'],
                line['converted_unit'], cantidad_convertida, precio_unitario,
                line['invoice_date'], line['payment_date'], line['customer_nit'],
                line['customer_name'], line['supplier_nit'], line['supplier_name'],
                'V', line['supplier_city'], line['iva_percent'], line['conversion_note'],
                '1', '1', '', cantidad_original, line['currency_id'],
                round(line['line_total'], 2), round(line['iva_amount'], 2),
                round(line['total_with_iva'], 2)
            ]

            for col, value in enumerate(row_data, start=1):
                cell = ws.cell(row=idx, column=col, value=value)
                cell.alignment = Alignment(vertical='center')

        column_widths = [15, 40, 15, 20, 15, 15, 12, 12, 15, 30, 15, 30, 12, 20, 8, 50, 12, 12, 10, 18, 10, 15, 12, 15]
        for col, width in enumerate(column_widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

        wb.save(output_path)

    def procesar(self) -> Path:
        zip_files = list(self.carpeta_zip.glob("*.zip"))

        if not zip_files:
            raise Exception("No se encontraron archivos ZIP en la carpeta seleccionada")

        logger.info(f"CASA DEL AGRICULTOR: Encontrados {len(zip_files)} archivos ZIP")

        for zip_file in zip_files:
            logger.info(f"Procesando: {zip_file.name}")

            xml_content = self.extract_xml_from_zip(zip_file)
            if not xml_content:
                continue

            lines = self.parse_invoice_xml(xml_content)
            if not lines:
                logger.warning(f"No se extrajeron líneas de {zip_file.name}")
                continue

            for line in lines:
                processed_line = self.apply_conversion_rules(line)
                self.processed_lines.append(processed_line)

                if processed_line['conversion_note']:
                    logger.info(f"  {processed_line['conversion_note']}")

        if not self.processed_lines:
            raise Exception("No se procesaron líneas. Verifique los archivos ZIP.")

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = self.carpeta_salida / f"Facturas_REGGIS_CASA_DEL_AGRICULTOR_{timestamp}.xlsx"

        self.create_reggis_excel(output_file)
        logger.info(f"CASA DEL AGRICULTOR: Proceso completado - {len(self.processed_lines)} líneas")

        return self.carpeta_salida


class ProcesadorLactalis:
    """
    Procesador específico para LACTALIS

    NOTA: Esta es una estructura base. El procesamiento específico
    debe ser configurado según los requisitos de Lactalis.
    """

    def __init__(self, carpeta_entrada: Path, carpeta_salida: Path):
        self.carpeta_entrada = carpeta_entrada
        self.carpeta_salida = carpeta_salida
        self.processed_lines = []

    def procesar(self) -> Path:
        """
        Método de procesamiento principal para Lactalis

        TODO: Implementar lógica específica de Lactalis
        """
        logger.info("LACTALIS: Iniciando procesamiento...")

        # PLACEHOLDER: Aquí va la lógica específica de Lactalis
        # Por ahora, retornamos la carpeta de salida

        raise NotImplementedError(
            "El procesamiento de Lactalis aún no está implementado.\n"
            "Por favor, contacte al desarrollador para configurar este módulo."
        )


# =====================================================================
# INTERFAZ PyQt6 PROFESIONAL
# =====================================================================

class WorkerThread(QThread):
    """Thread worker para procesar sin bloquear la UI"""
    finished = pyqtSignal(bool, str, object)  # success, message, result
    progress = pyqtSignal(str)

    def __init__(self, func, *args, **kwargs):
        super().__init__()
        self.func = func
        self.args = args
        self.kwargs = kwargs

    def run(self):
        try:
            result = self.func(*self.args, **self.kwargs)
            self.finished.emit(True, "Proceso completado exitosamente", result)
        except Exception as e:
            self.finished.emit(False, str(e), None)


class UpdateCheckerThread(QThread):
    """Thread para verificar actualizaciones sin bloquear la UI"""
    update_available = pyqtSignal(bool, str, str)  # has_update, version, error

    def run(self):
        if not AUTO_UPDATE_AVAILABLE:
            self.update_available.emit(False, "", "Sistema de actualización no disponible")
            return

        try:
            has_update, new_version, error = check_and_notify_update()
            self.update_available.emit(has_update, new_version or "", error or "")
        except Exception as e:
            self.update_available.emit(False, "", str(e))


class UpdateDialog(QDialog):
    """Diálogo para mostrar actualizaciones disponibles"""

    def __init__(self, new_version: str, parent=None):
        super().__init__(parent)
        self.new_version = new_version
        self.should_update = False
        self.init_ui()

    def init_ui(self):
        self.setWindowTitle("Actualización Disponible")
        self.setMinimumWidth(500)
        self.setMinimumHeight(250)

        layout = QVBoxLayout()
        layout.setSpacing(20)

        # Título
        title = QLabel("¡Nueva actualización disponible!")
        title_font = QFont("Segoe UI", 14, QFont.Weight.Bold)
        title.setFont(title_font)
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(title)

        # Información de versión
        version_info = QLabel(f"Versión disponible: <b>{self.new_version}</b>")
        version_info.setFont(QFont("Segoe UI", 11))
        version_info.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(version_info)

        # Mensaje
        message = QLabel(
            "Se ha detectado una nueva versión del sistema.\n\n"
            "¿Desea descargar e instalar la actualización ahora?\n"
            "La aplicación se reiniciará después de la actualización."
        )
        message.setFont(QFont("Segoe UI", 10))
        message.setAlignment(Qt.AlignmentFlag.AlignCenter)
        message.setWordWrap(True)
        layout.addWidget(message)

        # Espacio
        layout.addStretch()

        # Botones
        button_layout = QHBoxLayout()

        btn_update = QPushButton("Actualizar Ahora")
        btn_update.setFont(QFont("Segoe UI", 10, QFont.Weight.Bold))
        btn_update.setStyleSheet("""
            QPushButton {
                background-color: #0078D4;
                color: white;
                border: none;
                padding: 10px 20px;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #106EBE;
            }
        """)
        btn_update.clicked.connect(self.accept_update)

        btn_later = QPushButton("Más Tarde")
        btn_later.setFont(QFont("Segoe UI", 10))
        btn_later.setStyleSheet("""
            QPushButton {
                background-color: #6c757d;
                color: white;
                border: none;
                padding: 10px 20px;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #5a6268;
            }
        """)
        btn_later.clicked.connect(self.reject)

        button_layout.addWidget(btn_update)
        button_layout.addWidget(btn_later)

        layout.addLayout(button_layout)

        self.setLayout(layout)

    def accept_update(self):
        self.should_update = True
        self.accept()


class ClientSelectorWindow(QMainWindow):
    """Ventana principal de selección de cliente con diseño empresarial"""

    def __init__(self):
        super().__init__()
        self.selected_client = None
        self.init_ui()
        self.check_for_updates()

    def init_ui(self):
        self.setWindowTitle("Sistema SAE - Procesador de Facturas Electrónicas")
        self.setMinimumSize(900, 600)

        # Aplicar estilo empresarial
        self.setStyleSheet("""
            QMainWindow {
                background-color: #f5f5f5;
            }
            QLabel {
                color: #333333;
            }
            QPushButton {
                font-size: 11pt;
                font-weight: bold;
                padding: 15px;
                border-radius: 8px;
                border: none;
            }
        """)

        # Widget central
        central_widget = QWidget()
        self.setCentralWidget(central_widget)

        # Layout principal
        main_layout = QVBoxLayout(central_widget)
        main_layout.setSpacing(20)
        main_layout.setContentsMargins(40, 40, 40, 40)

        # Encabezado con logo/título
        header_frame = QFrame()
        header_frame.setStyleSheet("""
            QFrame {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #0078D4, stop:1 #106EBE);
                border-radius: 10px;
                padding: 20px;
            }
        """)
        header_layout = QVBoxLayout(header_frame)

        title = QLabel("SISTEMA SAE")
        title.setFont(QFont("Segoe UI", 28, QFont.Weight.Bold))
        title.setStyleSheet("color: white;")
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)

        subtitle = QLabel("Procesador de Facturas Electrónicas XML")
        subtitle.setFont(QFont("Segoe UI", 14))
        subtitle.setStyleSheet("color: #E8F4FF;")
        subtitle.setAlignment(Qt.AlignmentFlag.AlignCenter)

        header_layout.addWidget(title)
        header_layout.addWidget(subtitle)

        main_layout.addWidget(header_frame)

        # Instrucciones
        instructions = QLabel("Seleccione el cliente que desea procesar:")
        instructions.setFont(QFont("Segoe UI", 12))
        instructions.setAlignment(Qt.AlignmentFlag.AlignCenter)
        main_layout.addWidget(instructions)

        # Contenedor de botones de clientes
        buttons_container = QWidget()
        buttons_layout = QHBoxLayout(buttons_container)
        buttons_layout.setSpacing(20)

        # Botón SEABOARD
        btn_seaboard = QPushButton("🌐 SEABOARD\n\nProcesamiento desde\nSharePoint/Local")
        btn_seaboard.setMinimumHeight(150)
        btn_seaboard.setStyleSheet("""
            QPushButton {
                background-color: #0078D4;
                color: white;
            }
            QPushButton:hover {
                background-color: #106EBE;
            }
            QPushButton:pressed {
                background-color: #005A9E;
            }
        """)
        btn_seaboard.clicked.connect(lambda: self.select_client("SEABOARD"))

        # Botón Casa del Agricultor
        btn_casa = QPushButton("🌾 CASA DEL AGRICULTOR\n\nProcesamiento desde\narchivos ZIP")
        btn_casa.setMinimumHeight(150)
        btn_casa.setStyleSheet("""
            QPushButton {
                background-color: #27ae60;
                color: white;
            }
            QPushButton:hover {
                background-color: #229954;
            }
            QPushButton:pressed {
                background-color: #1e8449;
            }
        """)
        btn_casa.clicked.connect(lambda: self.select_client("CASA_DEL_AGRICULTOR"))

        # Botón Lactalis
        btn_lactalis = QPushButton("🥛 LACTALIS\n\nProcesamiento\nLactalis")
        btn_lactalis.setMinimumHeight(150)
        btn_lactalis.setStyleSheet("""
            QPushButton {
                background-color: #8e44ad;
                color: white;
            }
            QPushButton:hover {
                background-color: #7d3c98;
            }
            QPushButton:pressed {
                background-color: #6c3483;
            }
        """)
        btn_lactalis.clicked.connect(lambda: self.select_client("LACTALIS"))

        buttons_layout.addWidget(btn_seaboard)
        buttons_layout.addWidget(btn_casa)
        buttons_layout.addWidget(btn_lactalis)

        main_layout.addWidget(buttons_container)

        # Información adicional
        info = QLabel("💡 Cada cliente tiene su propio flujo de procesamiento optimizado")
        info.setFont(QFont("Segoe UI", 10))
        info.setStyleSheet("color: #666666;")
        info.setAlignment(Qt.AlignmentFlag.AlignCenter)
        main_layout.addWidget(info)

        main_layout.addStretch()

        # Barra de estado para actualizaciones
        self.status_label = QLabel("")
        self.status_label.setFont(QFont("Segoe UI", 9))
        self.status_label.setStyleSheet("color: #0078D4;")
        self.status_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        main_layout.addWidget(self.status_label)

        # Botón cerrar
        btn_close = QPushButton("Cerrar")
        btn_close.setMaximumWidth(150)
        btn_close.setStyleSheet("""
            QPushButton {
                background-color: #dc3545;
                color: white;
                padding: 10px;
            }
            QPushButton:hover {
                background-color: #c82333;
            }
        """)
        btn_close.clicked.connect(self.close)

        close_layout = QHBoxLayout()
        close_layout.addStretch()
        close_layout.addWidget(btn_close)
        main_layout.addLayout(close_layout)

        # Centrar ventana
        self.center()

    def center(self):
        """Centra la ventana en la pantalla"""
        screen = QApplication.primaryScreen().geometry()
        window_geometry = self.frameGeometry()
        center_point = screen.center()
        window_geometry.moveCenter(center_point)
        self.move(window_geometry.topLeft())

    def check_for_updates(self):
        """Verifica actualizaciones disponibles al iniciar"""
        if not AUTO_UPDATE_AVAILABLE:
            return

        self.status_label.setText("Verificando actualizaciones...")
        self.update_checker = UpdateCheckerThread()
        self.update_checker.update_available.connect(self.on_update_check_complete)
        self.update_checker.start()

    def on_update_check_complete(self, has_update: bool, version: str, error: str):
        """Callback cuando se completa la verificación de actualizaciones"""
        if error:
            self.status_label.setText(f"Error al verificar actualizaciones: {error}")
            return

        if has_update:
            self.status_label.setText(f"Nueva versión disponible: {version}")
            # Mostrar diálogo de actualización
            dialog = UpdateDialog(version, self)
            if dialog.exec() and dialog.should_update:
                self.perform_update()
        else:
            self.status_label.setText("Sistema actualizado")
            # Ocultar mensaje después de 3 segundos
            QTimer.singleShot(3000, lambda: self.status_label.setText(""))

    def perform_update(self):
        """Realiza la actualización del sistema"""
        msg = QMessageBox(self)
        msg.setIcon(QMessageBox.Icon.Information)
        msg.setWindowTitle("Actualizando")
        msg.setText("Descargando e instalando actualización...\n\nEsto puede tomar unos momentos.")
        msg.setStandardButtons(QMessageBox.StandardButton.NoButton)
        msg.show()
        QApplication.processEvents()

        try:
            success, error = perform_update()
            msg.close()

            if success:
                QMessageBox.information(
                    self,
                    "Actualización Completada",
                    "La actualización se ha instalado correctamente.\n\n"
                    "Por favor, reinicie la aplicación para aplicar los cambios."
                )
                self.close()
            else:
                QMessageBox.critical(
                    self,
                    "Error en la Actualización",
                    f"No se pudo completar la actualización:\n\n{error}"
                )
        except Exception as e:
            msg.close()
            QMessageBox.critical(
                self,
                "Error",
                f"Error inesperado durante la actualización:\n\n{str(e)}"
            )

    def select_client(self, client: str):
        """Selecciona un cliente y abre su ventana de procesamiento"""
        self.selected_client = client
        logger.info(f"Cliente seleccionado: {client}")

        # Crear y mostrar ventana del procesador
        self.processor_window = ProcessorWindow(client)
        self.processor_window.show()
        self.hide()


class ProcessorWindow(QMainWindow):
    """Ventana de procesamiento para cada cliente"""

    def __init__(self, client: str):
        super().__init__()
        self.client = client
        self.carpeta_entrada = None
        self.carpetas_sharepoint = []
        self.worker = None

        if client == "SEABOARD":
            self.detect_sharepoint()

        self.init_ui()

    def init_ui(self):
        self.setWindowTitle(f"Procesador de Facturas - {self.client}")
        self.setMinimumSize(800, 600)

        # Estilo
        self.setStyleSheet("""
            QMainWindow {
                background-color: #f5f5f5;
            }
        """)

        # Widget central
        central_widget = QWidget()
        self.setCentralWidget(central_widget)

        # Layout principal
        main_layout = QVBoxLayout(central_widget)
        main_layout.setSpacing(20)
        main_layout.setContentsMargins(30, 30, 30, 30)

        # Encabezado
        header_colors = {
            "SEABOARD": "#0078D4",
            "CASA_DEL_AGRICULTOR": "#27ae60",
            "LACTALIS": "#8e44ad"
        }

        header_frame = QFrame()
        header_frame.setStyleSheet(f"""
            QFrame {{
                background-color: {header_colors.get(self.client, '#0078D4')};
                border-radius: 10px;
                padding: 20px;
            }}
        """)
        header_layout = QVBoxLayout(header_frame)

        title = QLabel(f"PROCESADOR - {self.client}")
        title.setFont(QFont("Segoe UI", 20, QFont.Weight.Bold))
        title.setStyleSheet("color: white;")
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        header_layout.addWidget(title)

        main_layout.addWidget(header_frame)

        # Botones según el cliente
        if self.client == "SEABOARD":
            self.setup_seaboard_ui(main_layout)
        elif self.client == "CASA_DEL_AGRICULTOR":
            self.setup_casa_ui(main_layout)
        elif self.client == "LACTALIS":
            self.setup_lactalis_ui(main_layout)

        # Barra de progreso
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        self.progress_bar.setTextVisible(False)
        self.progress_bar.setStyleSheet("""
            QProgressBar {
                border: 2px solid #cccccc;
                border-radius: 5px;
                text-align: center;
                height: 25px;
            }
            QProgressBar::chunk {
                background-color: #0078D4;
            }
        """)
        main_layout.addWidget(self.progress_bar)

        # Label de estado
        self.status_label = QLabel("")
        self.status_label.setFont(QFont("Segoe UI", 10))
        self.status_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.status_label.setWordWrap(True)
        main_layout.addWidget(self.status_label)

        main_layout.addStretch()

        # Botones inferiores
        bottom_layout = QHBoxLayout()

        btn_back = QPushButton("← Volver")
        btn_back.setStyleSheet("""
            QPushButton {
                background-color: #6c757d;
                color: white;
                padding: 10px 20px;
                border-radius: 5px;
                font-size: 10pt;
            }
            QPushButton:hover {
                background-color: #5a6268;
            }
        """)
        btn_back.clicked.connect(self.go_back)

        btn_close = QPushButton("Cerrar")
        btn_close.setStyleSheet("""
            QPushButton {
                background-color: #dc3545;
                color: white;
                padding: 10px 20px;
                border-radius: 5px;
                font-size: 10pt;
            }
            QPushButton:hover {
                background-color: #c82333;
            }
        """)
        btn_close.clicked.connect(self.close)

        bottom_layout.addWidget(btn_back)
        bottom_layout.addStretch()
        bottom_layout.addWidget(btn_close)

        main_layout.addLayout(bottom_layout)

        # Centrar ventana
        self.center()

    def setup_seaboard_ui(self, layout):
        """Configura UI específica para SEABOARD"""
        if self.carpetas_sharepoint:
            info = QLabel(f"✓ Se detectaron {len(self.carpetas_sharepoint)} carpeta(s) de SharePoint sincronizada(s)")
            info.setFont(QFont("Segoe UI", 11))
            info.setStyleSheet("color: #27ae60; font-weight: bold;")
            info.setAlignment(Qt.AlignmentFlag.AlignCenter)
            layout.addWidget(info)

            btn_sharepoint = QPushButton("🔍 BUSCAR EN SHAREPOINT SINCRONIZADO")
            btn_sharepoint.setMinimumHeight(60)
            btn_sharepoint.setStyleSheet("""
                QPushButton {
                    background-color: #0078D4;
                    color: white;
                    font-size: 12pt;
                    font-weight: bold;
                    border-radius: 8px;
                }
                QPushButton:hover {
                    background-color: #106EBE;
                }
            """)
            btn_sharepoint.clicked.connect(self.select_from_sharepoint)
            layout.addWidget(btn_sharepoint)

        btn_local = QPushButton("📁 BUSCAR EN CARPETA LOCAL")
        btn_local.setMinimumHeight(60)
        btn_local.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                font-size: 12pt;
                font-weight: bold;
                border-radius: 8px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
        """)
        btn_local.clicked.connect(self.select_and_process_seaboard)
        layout.addWidget(btn_local)

    def setup_casa_ui(self, layout):
        """Configura UI específica para Casa del Agricultor"""
        info = QLabel("Seleccione la carpeta que contiene los archivos ZIP de facturas")
        info.setFont(QFont("Segoe UI", 11))
        info.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(info)

        btn_process = QPushButton("📦 SELECCIONAR CARPETA CON ARCHIVOS ZIP")
        btn_process.setMinimumHeight(60)
        btn_process.setStyleSheet("""
            QPushButton {
                background-color: #27ae60;
                color: white;
                font-size: 12pt;
                font-weight: bold;
                border-radius: 8px;
            }
            QPushButton:hover {
                background-color: #229954;
            }
        """)
        btn_process.clicked.connect(self.select_and_process_casa)
        layout.addWidget(btn_process)

    def setup_lactalis_ui(self, layout):
        """Configura UI específica para Lactalis"""
        info = QLabel("⚠ MÓDULO EN CONFIGURACIÓN")
        info.setFont(QFont("Segoe UI", 14, QFont.Weight.Bold))
        info.setStyleSheet("color: #e74c3c;")
        info.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(info)

        description = QLabel(
            "El módulo de procesamiento para Lactalis está en desarrollo.\n\n"
            "Por favor, contacte al administrador del sistema\n"
            "para configurar las reglas de procesamiento específicas."
        )
        description.setFont(QFont("Segoe UI", 11))
        description.setAlignment(Qt.AlignmentFlag.AlignCenter)
        description.setWordWrap(True)
        layout.addWidget(description)

        btn_placeholder = QPushButton("📋 CONFIGURAR PROCESAMIENTO")
        btn_placeholder.setMinimumHeight(60)
        btn_placeholder.setStyleSheet("""
            QPushButton {
                background-color: #8e44ad;
                color: white;
                font-size: 12pt;
                font-weight: bold;
                border-radius: 8px;
            }
            QPushButton:hover {
                background-color: #7d3c98;
            }
        """)
        btn_placeholder.clicked.connect(self.show_lactalis_config)
        layout.addWidget(btn_placeholder)

    def center(self):
        """Centra la ventana en la pantalla"""
        screen = QApplication.primaryScreen().geometry()
        window_geometry = self.frameGeometry()
        center_point = screen.center()
        window_geometry.moveCenter(center_point)
        self.move(window_geometry.topLeft())

    def detect_sharepoint(self):
        """Detecta carpetas de SharePoint"""
        try:
            self.carpetas_sharepoint = DetectorSharePoint.encontrar_carpetas_sharepoint()
            if self.carpetas_sharepoint:
                logger.info(f"Encontradas {len(self.carpetas_sharepoint)} carpetas de SharePoint")
        except Exception as e:
            logger.error(f"Error detectando SharePoint: {e}")

    def select_from_sharepoint(self):
        """Muestra diálogo para seleccionar carpeta de SharePoint"""
        if not self.carpetas_sharepoint:
            QMessageBox.information(self, "Sin carpetas", "No se detectaron carpetas de SharePoint")
            return

        dialog = QDialog(self)
        dialog.setWindowTitle("Seleccionar Carpeta de SharePoint")
        dialog.setMinimumSize(600, 400)

        layout = QVBoxLayout(dialog)

        title = QLabel("Carpetas de SharePoint Detectadas")
        title.setFont(QFont("Segoe UI", 12, QFont.Weight.Bold))
        layout.addWidget(title)

        list_widget = QListWidget()
        list_widget.setFont(QFont("Segoe UI", 10))
        for carpeta in self.carpetas_sharepoint:
            list_widget.addItem(str(carpeta))
        layout.addWidget(list_widget)

        btn_select = QPushButton("Procesar Carpeta Seleccionada")
        btn_select.setStyleSheet("""
            QPushButton {
                background-color: #0078D4;
                color: white;
                padding: 10px;
                font-size: 11pt;
                font-weight: bold;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #106EBE;
            }
        """)

        def on_select():
            if list_widget.currentRow() >= 0:
                carpeta = self.carpetas_sharepoint[list_widget.currentRow()]
                dialog.accept()
                self.process_xml_folder(carpeta)

        btn_select.clicked.connect(on_select)
        layout.addWidget(btn_select)

        dialog.exec()

    def select_and_process_seaboard(self):
        """Selecciona carpeta local y procesa para SEABOARD"""
        folder = QFileDialog.getExistingDirectory(
            self,
            "Seleccione la carpeta con archivos XML",
            "",
            QFileDialog.Option.ShowDirsOnly
        )

        if folder:
            self.process_xml_folder(Path(folder))

    def process_xml_folder(self, folder: Path):
        """Procesa carpeta de XMLs para SEABOARD"""
        xml_files = list(folder.glob("*.xml"))

        if not xml_files:
            QMessageBox.critical(self, "Sin archivos", "No se encontraron archivos XML en la carpeta")
            return

        reply = QMessageBox.question(
            self,
            "Confirmar",
            f"Se encontraron {len(xml_files)} archivo(s) XML.\n\n¿Procesar ahora?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )

        if reply == QMessageBox.StandardButton.Yes:
            self.carpeta_entrada = folder
            self.start_processing_seaboard()

    def select_and_process_casa(self):
        """Selecciona carpeta y procesa para Casa del Agricultor"""
        folder = QFileDialog.getExistingDirectory(
            self,
            "Seleccione la carpeta con archivos ZIP",
            "",
            QFileDialog.Option.ShowDirsOnly
        )

        if not folder:
            return

        folder_path = Path(folder)
        zip_files = list(folder_path.glob("*.zip"))

        if not zip_files:
            QMessageBox.critical(self, "Sin archivos", "No se encontraron archivos ZIP en la carpeta")
            return

        reply = QMessageBox.question(
            self,
            "Confirmar",
            f"Se encontraron {len(zip_files)} archivo(s) ZIP.\n\n¿Procesar ahora?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )

        if reply == QMessageBox.StandardButton.Yes:
            self.carpeta_entrada = folder_path
            self.start_processing_casa()

    def show_lactalis_config(self):
        """Muestra información de configuración para Lactalis"""
        QMessageBox.information(
            self,
            "Lactalis - En Desarrollo",
            "Este módulo está en desarrollo.\n\n"
            "Para configurar el procesamiento de Lactalis, "
            "por favor contacte al administrador del sistema con los siguientes detalles:\n\n"
            "• Formato de archivos de entrada (XML, ZIP, otro)\n"
            "• Estructura de datos esperada\n"
            "• Reglas de conversión específicas\n"
            "• Formato de salida deseado"
        )

    def start_processing_seaboard(self):
        """Inicia procesamiento para SEABOARD"""
        self.progress_bar.setVisible(True)
        self.progress_bar.setRange(0, 0)  # Modo indeterminado
        self.status_label.setText("Procesando facturas de SEABOARD...")
        self.status_label.setStyleSheet("color: #f39c12; font-weight: bold;")

        plantilla = self.get_or_create_template()

        def process():
            procesador = ProcesadorSeaboard(self.carpeta_entrada, plantilla)
            return procesador.procesar()

        self.worker = WorkerThread(process)
        self.worker.finished.connect(self.on_processing_complete)
        self.worker.start()

    def start_processing_casa(self):
        """Inicia procesamiento para Casa del Agricultor"""
        self.progress_bar.setVisible(True)
        self.progress_bar.setRange(0, 0)
        self.status_label.setText("Procesando facturas de Casa del Agricultor...")
        self.status_label.setStyleSheet("color: #f39c12; font-weight: bold;")

        carpeta_salida = self.carpeta_entrada.parent / "Resultados_CASA_DEL_AGRICULTOR"
        carpeta_salida.mkdir(exist_ok=True)

        def process():
            procesador = ProcesadorCasaDelAgricultor(self.carpeta_entrada, carpeta_salida)
            return procesador.procesar()

        self.worker = WorkerThread(process)
        self.worker.finished.connect(self.on_processing_complete)
        self.worker.start()

    def on_processing_complete(self, success: bool, message: str, result):
        """Callback cuando el procesamiento termina"""
        self.progress_bar.setVisible(False)

        if success:
            self.status_label.setText("✓ Proceso completado exitosamente")
            self.status_label.setStyleSheet("color: #27ae60; font-weight: bold;")

            reply = QMessageBox.question(
                self,
                "Éxito",
                f"{message}\n\nArchivos guardados en:\n{result.name if result else 'carpeta de salida'}\n\n¿Abrir carpeta?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )

            if reply == QMessageBox.StandardButton.Yes and result:
                self.open_folder(result)
        else:
            self.status_label.setText("✗ Error en el procesamiento")
            self.status_label.setStyleSheet("color: #e74c3c; font-weight: bold;")
            QMessageBox.critical(self, "Error", f"Error:\n\n{message}")

    def get_or_create_template(self) -> Path:
        """Obtiene o crea la plantilla de Excel"""
        script_dir = Path(__file__).parent
        plantilla = script_dir / "Plantilla_REGGIS.xlsx"

        if not plantilla.exists():
            self.create_template(plantilla)

        return plantilla

    def create_template(self, path: Path):
        """Crea plantilla base de Excel"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Datos"

        headers = [
            'N° Factura', 'Nombre Producto', 'Codigo Subyacente',
            'Unidad Medida en Kg,Un,Lt', 'Cantidad (5 decimales - separdor coma)',
            'Precio Unitario (5 decimales - separdor coma)', 'Fecha Factura Año-Mes-Dia',
            'Fecha Pago Año-Mes-Dia', 'Nit Comprador (Existente)', 'Nombre Comprador',
            'Nit Vendedor (Existente)', 'Nombre Vendedor', 'Principal V,C',
            'Municipio (Nombre Exacto de la Ciudad)', 'Iva (N°%)', 'Descripción',
            'Activa Factura', 'Activa Bodega', 'Incentivo',
            'Cantidad Original (5 decimales - separdor coma)', 'Moneda (1,2,3)',
            'Total Sin IVA', 'Total IVA', 'Total Con IVA'
        ]

        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        wb.save(path)

    def open_folder(self, folder: Path):
        """Abre carpeta en el explorador del sistema"""
        if platform.system() == 'Windows':
            os.startfile(folder)
        elif platform.system() == 'Darwin':
            os.system(f'open "{folder}"')
        else:
            os.system(f'xdg-open "{folder}"')

    def go_back(self):
        """Vuelve al selector de clientes"""
        self.close()
        # Reabrir ventana de selección
        selector = ClientSelectorWindow()
        selector.show()


def main():
    """Función principal de la aplicación"""
    app = QApplication(sys.argv)

    # Configurar estilo global
    app.setStyle("Fusion")

    # Paleta de colores profesional
    palette = QPalette()
    palette.setColor(QPalette.ColorRole.Window, QColor(245, 245, 245))
    palette.setColor(QPalette.ColorRole.WindowText, QColor(51, 51, 51))
    palette.setColor(QPalette.ColorRole.Base, QColor(255, 255, 255))
    palette.setColor(QPalette.ColorRole.AlternateBase, QColor(245, 245, 245))
    palette.setColor(QPalette.ColorRole.ToolTipBase, QColor(255, 255, 255))
    palette.setColor(QPalette.ColorRole.ToolTipText, QColor(51, 51, 51))
    palette.setColor(QPalette.ColorRole.Text, QColor(51, 51, 51))
    palette.setColor(QPalette.ColorRole.Button, QColor(245, 245, 245))
    palette.setColor(QPalette.ColorRole.ButtonText, QColor(51, 51, 51))
    palette.setColor(QPalette.ColorRole.Highlight, QColor(0, 120, 212))
    palette.setColor(QPalette.ColorRole.HighlightedText, QColor(255, 255, 255))
    app.setPalette(palette)

    # Mostrar ventana principal
    window = ClientSelectorWindow()
    window.show()

    sys.exit(app.exec())


if __name__ == "__main__":
    main()
